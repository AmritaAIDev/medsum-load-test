"""Load testing routes: Excel upload, SSE run, and results export."""

from __future__ import annotations

import base64
import io
import json
import logging
import queue
import re
import threading
import time
import uuid
import wave
from concurrent.futures import ThreadPoolExecutor, as_completed
from datetime import datetime

import requests
from flask import Blueprint, Response, jsonify, request, send_file, stream_with_context
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill

from medsum_testing.backend.services.config_loader import get_config
from medsum_testing.backend.services.drive_service import (
    download_file,
    get_drive_service,
    list_drive_children,
)

load_test_bp = Blueprint("load_test", __name__)
log = logging.getLogger("medsum_load_test")

SOAP_TEMPLATE = {
    "subjective": {
        "chief_complaint": "",
        "history_of_present_illness": "",
        "past_medical_history": "",
        "medications": "",
        "allergies": "",
        "social_history": "",
        "family_history": "",
    },
    "objective": {
        "vitals": {
            "blood_pressure": "",
            "heart_rate": "",
            "respiratory_rate": "",
            "temperature": "",
        },
        "physical_exam": {"heart": []},
    },
    "assessment": [],
    "plan": {
        "medications": [
            {
                "drug_name": "",
                "dose": "",
                "schedule": "",
                "duration": "",
                "instructions": "",
                "snomed_ct_id": "",
            }
        ],
        "activity": "",
        "investigations": "",
        "education": "",
        "follow_up": "",
    },
    "summary": "",
}

DISCHARGE_TEMPLATE = {
    "Patient_Details": {
        "Name": "",
        "Age": "",
        "Gender": "",
        "Blood_Group": "",
        "ABHA_ID": "",
    },
    "Admission_Details": {
        "Date_of_Admission": "",
        "Date_of_Discharge": "",
        "Ward": "",
        "Bed_Number": "",
        "Consultant": "",
        "Department": "",
    },
    "Presenting_Complaints": "",
    "History_of_Presenting_Illness": "",
    "Past_Medical_History": "",
    "Personal_History": "",
    "Family_History": "",
    "Examination_Findings": {"General_Examination": "", "Systemic_Examination": ""},
    "Investigations": "",
    "Diagnosis": {"Primary_Diagnosis": "", "Secondary_Diagnosis": ""},
    "Course_in_the_Hospital_and_Discussion": "",
    "Operative_or_Procedure_Findings": "",
    "Prognosis_on_Discharge": "",
    "Diet_Recommendation": "",
    "Physiscal_Activity": "",
    "Discharge_Medication": {
        "Serial_Number": [],
        "Drug_Name": [],
        "Dose": [],
        "Schedule": [],
        "Duration": [],
        "Days": [],
        "Instructions": [],
    },
    "Follow_Up": "",
    "summary": "",
}

KNOWN_META_KEYS = {
    "doctor_details",
    "patient_demographics",
    "session",
    "transcription",
    "transcription-time",
    "translation-time",
    "llm-time",
    "audio_length",
    "total-time",
    "medicines_backend",
}

HEADER_ALIASES = {
    "phone": "phone",
    "phone_number": "phone",
    "password": "password",
    "patient_id": "patient_id",
    "patient": "patient_id",
}

FOLDER_ID_RE = re.compile(r"/folders/([a-zA-Z0-9_-]+)")


def _make_wav(duration: float = 2.0, rate: int = 16000) -> bytes:
    buf = io.BytesIO()
    with wave.open(buf, "wb") as wf:
        wf.setnchannels(1)
        wf.setsampwidth(2)
        wf.setframerate(rate)
        wf.writeframes(b"\x00\x00" * int(rate * duration))
    return buf.getvalue()


def _extract_summary(body: dict) -> str:
    text = body.get("summary", "")
    if not text:
        for key in (
            "medical_summary",
            "subjective",
            "Diagnosis",
            "Course_in_the_Hospital_and_Discussion",
            "Medicines",
        ):
            value = body.get(key)
            if value:
                text = value if isinstance(value, str) else json.dumps(value, ensure_ascii=False)
                break
    if not text:
        content = {k: v for k, v in body.items() if k not in KNOWN_META_KEYS}
        text = json.dumps(content, ensure_ascii=False)
    return text


def _norm_header(value) -> str:
    raw = str(value or "").strip().lower().replace(" ", "_").replace("-", "_")
    return HEADER_ALIASES.get(raw, raw)


def _cell_str(value) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    if isinstance(value, int):
        return str(value)
    return str(value).strip()


def _extract_folder_id(drive_link: str) -> str | None:
    match = FOLDER_ID_RE.search(drive_link or "")
    if match:
        return match.group(1)
    stripped = (drive_link or "").strip()
    if re.fullmatch(r"[a-zA-Z0-9_-]+", stripped):
        return stripped
    return None


def _list_folder_files(service, folder_id: str) -> list[dict]:
    """List files in a Drive folder, including one level of subfolders."""
    children = list_drive_children(
        service,
        folder_id,
        fields="files(id, name, mimeType, fileExtension)",
    )
    files: list[dict] = []
    for child in children:
        if child.get("mimeType") == "application/vnd.google-apps.folder":
            files.extend(
                list_drive_children(
                    service,
                    child["id"],
                    fields="files(id, name, mimeType, fileExtension)",
                )
            )
        else:
            files.append(child)
    return files


def _match_named_file(files: list[dict], patient_id: str, kind: str) -> dict | None:
    """Match {patient_id}_audio.* or {patient_id}_ground_truth.* (case-insensitive)."""
    prefix = f"{patient_id}_{kind}".lower()
    for item in files:
        name = (item.get("name") or "").lower()
        stem = name.rsplit(".", 1)[0] if "." in name else name
        if stem == prefix or name.startswith(prefix + "."):
            return item
    return None


def _urls_from_config(config: dict) -> tuple[str, str]:
    backends = config.get("backends") or {}
    django = (backends.get("django_base_url") or "").rstrip("/")
    flask_url = (backends.get("flask_transcribe_url") or "").rstrip("/")
    if flask_url and not flask_url.endswith("/transcribe"):
        flask_url += "/transcribe"
    return django, flask_url


def _emit(q: queue.Queue, payload: dict) -> None:
    q.put(payload)


def _run_row(
    row_index: int,
    row: dict,
    drive_files: list[dict],
    drive_svc,
    cfg: dict,
    q: queue.Queue,
    audio_override: tuple[str, bytes] | None = None,
    gt_override: tuple[str, bytes] | None = None,
) -> str:
    """Run one doctor–patient session. Returns 'pass' or 'fail'."""
    phone = str(row.get("phone") or "").strip()
    password = str(row.get("password") or "")
    patient_id = str(row.get("patient_id") or "").strip()
    django = cfg["django_url"]
    flask_url = cfg["flask_url"]
    language = cfg.get("language", "en")
    llm = cfg["llm"]
    stt_model = cfg["stt_model"]
    translate_model = cfg["translate_model"]
    template_id = cfg["template_id"]
    template = cfg["template"]
    audio_duration = float(cfg["duration"])
    test_run_id = cfg["test_run_id"]
    mode = cfg.get("mode", "drive")
    wall_start = time.time()

    def running(stage: str) -> None:
        _emit(
            q,
            {
                "type": "row_update",
                "row_index": row_index,
                "status": "running",
                "stage": stage,
            },
        )

    def fail(error: str, extra: dict | None = None) -> str:
        elapsed = round(time.time() - wall_start, 2)
        timing = {"status": "fail", "error": str(error), "patient_id": patient_id}
        if extra:
            timing.update(extra)
        _emit(
            q,
            {
                "type": "row_update",
                "row_index": row_index,
                "status": "fail",
                "elapsed": elapsed,
                "error": str(error),
                "timing_data": timing,
            },
        )
        return "fail"

    doctor_times: dict = {}
    step1_time = step1b_time = patient_data_time = None
    doctor_id = None
    token = None

    try:
        running("Loading audio" if mode == "manual" else "Fetching audio from Drive")
        if gt_override:
            log.info(
                "Using uploaded ground truth for patient %s (%s)",
                patient_id,
                gt_override[0],
            )
        elif mode == "drive" and drive_svc:
            gt_file = _match_named_file(drive_files, patient_id, "ground_truth")
            if gt_file:
                try:
                    download_file(gt_file["id"], drive_svc)
                except Exception as gt_exc:
                    log.warning(
                        "Ground truth download failed for patient %s: %s",
                        patient_id,
                        gt_exc,
                    )

        if audio_override:
            audio_name, audio_bytes = audio_override
        elif mode == "drive" and drive_svc:
            audio_file = _match_named_file(drive_files, patient_id, "audio")
            if audio_file:
                audio_bytes = download_file(audio_file["id"], drive_svc)
                audio_name = audio_file.get("name") or f"{patient_id}_audio.wav"
            else:
                audio_bytes = _make_wav(audio_duration)
                audio_name = "silent.wav"
        else:
            audio_bytes = _make_wav(audio_duration)
            audio_name = "silent.wav"

        running("Logging in...")
        t0 = time.time()
        login_resp = requests.post(
            f"{django}/api/login/",
            json={"phone_number": phone, "password": password},
            timeout=30,
        )
        step1_time = round(time.time() - t0, 5)
        doctor_times["step1_time"] = step1_time
        if login_resp.status_code != 200:
            return fail(
                f"Login failed {login_resp.status_code}: {login_resp.text[:150]}",
                doctor_times,
            )
        body = login_resp.json()
        doctor_id = body["user"]["id"]
        token = body["access"]

        running("Fetching doctor profile...")
        doctor_name = department = hospital = ""
        t0 = time.time()
        try:
            profile_resp = requests.get(
                f"{django}/api/user/update/{doctor_id}/",
                headers={"Authorization": f"Bearer {token}"},
                timeout=15,
            )
            step1b_time = round(time.time() - t0, 5)
            doctor_times["step1b_time"] = step1b_time
            if profile_resp.status_code == 200:
                prof = profile_resp.json()
                doctor_name = f"Dr {prof.get('firstname', '')} {prof.get('lastname', '')}".strip()
                department = prof.get("department", "")
                hospital = prof.get("hospital_name", "")
        except Exception as profile_exc:
            step1b_time = round(time.time() - t0, 5)
            doctor_times["step1b_time"] = step1b_time
            log.warning("Doctor profile failed: %s", profile_exc)

        sess = requests.Session()
        sess.headers["Authorization"] = f"Bearer {token}"

        running("Fetching patient data...")
        patient_name = age = gender = ""
        t0 = time.time()
        try:
            patient_resp = sess.get(f"{django}/api/patient-data/{patient_id}/", timeout=15)
            patient_data_time = round(time.time() - t0, 5)
            if patient_resp.status_code == 200:
                pdata = patient_resp.json()
                patient_name = pdata.get("patient_name", "")
                age = str(pdata.get("age", ""))
                gender = pdata.get("gender", "")
        except Exception as patient_exc:
            patient_data_time = round(time.time() - t0, 5)
            log.warning("Patient data failed: %s", patient_exc)

        running("Transcribing...")
        clock_start = datetime.now()
        t0 = time.time()
        transcribe_resp = requests.post(
            flask_url,
            json={
                "audio_base64": base64.b64encode(audio_bytes).decode(),
                "doctor_name": doctor_name,
                "doctor_department": department,
                "hospital_name": hospital,
                "patient_id": str(patient_id),
                "patient_name": patient_name,
                "age": str(age),
                "gender": gender,
                "template": json.dumps(template),
                "template_id": template_id,
                "language": language,
                "llm": llm,
                "stt_model": stt_model,
                "translate_model": translate_model,
            },
            timeout=180,
        )
        step4_time = time.time() - t0
        if transcribe_resp.status_code != 200:
            return fail(
                f"transcribe {transcribe_resp.status_code}: {transcribe_resp.text[:200]}",
                {**doctor_times, "patient_metadata_time": patient_data_time},
            )
        b4 = transcribe_resp.json()
        transcription = b4.get("transcription", "")
        summary_text = _extract_summary(b4)
        flask_total = b4.get("total-time") or 0
        stt_time = b4.get("transcription-time") or 0
        tr_time = b4.get("translation-time") or 0
        llm_time = b4.get("llm-time") or 0
        audio_processing_time = (
            round(flask_total - (stt_time + tr_time + llm_time), 5) if flask_total else None
        )
        audio_length = b4.get("audio_length") or audio_duration

        running("Uploading audio...")
        client_sid = f"LOADTEST_{uuid.uuid4()}"
        t0 = time.time()
        audio_resp = sess.post(
            f"{django}/api/audio-data/",
            data={
                "user_id": str(doctor_id),
                "patient_id": str(patient_id),
                "language": language,
                "file_duration": str(audio_duration),
                "session_id": client_sid,
            },
            files={"audio": (audio_name, audio_bytes, "audio/wav")},
            timeout=30,
        )
        step5_time = time.time() - t0
        if audio_resp.status_code != 201:
            return fail(
                f"audio-data {audio_resp.status_code}: {audio_resp.text[:200]}",
                {**doctor_times, "patient_metadata_time": patient_data_time},
            )
        b5 = audio_resp.json()
        if "audio_id" not in b5:
            return fail(f"audio_id missing: {b5}", doctor_times)
        audio_id = b5["audio_id"]
        session_id = b5.get("session_id") or client_sid
        del audio_bytes

        running("Storing summary...")
        t0 = time.time()
        summary_resp = sess.post(
            f"{django}/api/summary-data/",
            json={
                "user_id": doctor_id,
                "patient_id": int(patient_id) if str(patient_id).isdigit() else patient_id,
                "audio_id": audio_id,
                "session_id": session_id,
                "summary": summary_text,
                "summary_length": len(summary_text),
                "transcription": transcription,
                "template_id": template_id,
                "original_summary": summary_text,
                "is_approved": "No",
                "is_update": "False",
                "google_doc_link": None,
            },
            timeout=30,
        )
        step6_time = time.time() - t0
        if summary_resp.status_code != 201:
            return fail(
                f"summary-data {summary_resp.status_code}: {summary_resp.text[:200]}",
                {**doctor_times, "patient_metadata_time": patient_data_time},
            )
        summary_id = summary_resp.json().get("summary_id")
        clock_end = datetime.now()

        timing_data = {
            "status": "pass",
            "doctor_id": doctor_id,
            "patient_id": patient_id,
            "audio_id": audio_id,
            "summary_id": summary_id,
            "audio_duration": audio_length,
            "language": language,
            "step1_time": step1_time,
            "step1b_time": step1b_time,
            "patient_metadata_time": patient_data_time,
            "transcription_time": b4.get("transcription-time"),
            "translation_time": b4.get("translation-time"),
            "llm_time": b4.get("llm-time"),
            "flask_total_time": b4.get("total-time"),
            "audio_processing_time": audio_processing_time,
            "transcribe_rtt": round(step4_time, 5),
            "user_percieved_summary_latency": round(step4_time, 5),
            "audio_upload_time": round(step5_time, 5),
            "summary_store_time": round(step6_time, 5),
            "clock_start": clock_start.strftime("%H:%M:%S.%f")[:-3],
            "clock_end": clock_end.strftime("%H:%M:%S.%f")[:-3],
        }

        running("Saving result...")
        try:
            save_resp = sess.post(
                f"{django}/api/load-test/result/",
                json={
                    "test_run_id": test_run_id,
                    "doctor_id": doctor_id,
                    "patient_id": int(patient_id) if str(patient_id).isdigit() else patient_id,
                    "audio_id": audio_id,
                    "summary_id": summary_id,
                    "language": language,
                    "llm": llm,
                    "stt_model": stt_model,
                    "translate_model": translate_model,
                    "audio_duration_s": audio_length,
                    "login_time": step1_time,
                    "doctor_profile_time": step1b_time,
                    "patient_metadata_time": patient_data_time,
                    "audio_upload_time": round(step5_time, 5),
                    "summary_store_time": round(step6_time, 5),
                    "user_percieved_summary_latency": round(step4_time, 5),
                    "transcription_time": b4.get("transcription-time"),
                    "translation_time": b4.get("translation-time"),
                    "llm_time": b4.get("llm-time"),
                    "flask_total_time": b4.get("total-time"),
                    "audio_processing_time": audio_processing_time,
                    "status": "pass",
                    "error_message": None,
                },
                timeout=10,
            )
            if save_resp.status_code != 201:
                log.warning("LoadTestResult save %s: %s", save_resp.status_code, save_resp.text[:120])
        except Exception as save_exc:
            log.warning("LoadTestResult save failed: %s", save_exc)

        elapsed = round(time.time() - wall_start, 2)
        _emit(
            q,
            {
                "type": "row_update",
                "row_index": row_index,
                "status": "pass",
                "elapsed": elapsed,
                "stage": "Done",
                "transcription": transcription,
                "summary": summary_text,
                "timing_data": timing_data,
            },
        )
        return "pass"

    except Exception as exc:
        return fail(str(exc), doctor_times)


@load_test_bp.route("/api/load-test/upload-excel", methods=["POST"])
def upload_excel():
    uploaded = request.files.get("file")
    if not uploaded:
        return jsonify({"error": "No file uploaded"}), 400
    filename = (uploaded.filename or "").lower()
    if not filename.endswith(".xlsx"):
        return jsonify({"error": "File must be a .xlsx spreadsheet"}), 400

    try:
        workbook = load_workbook(uploaded, read_only=True, data_only=True)
        sheet = workbook.active
        rows_iter = sheet.iter_rows(values_only=True)
        header_row = next(rows_iter, None)
        if not header_row:
            return jsonify({"error": "Spreadsheet is empty"}), 400

        headers = [_norm_header(cell) for cell in header_row]
        try:
            phone_i = headers.index("phone")
            password_i = headers.index("password")
            patient_i = headers.index("patient_id")
        except ValueError:
            return jsonify(
                {"error": "Expected columns: phone, password, patient_id"}
            ), 400

        parsed = []
        for row in rows_iter:
            if not row:
                continue
            phone = _cell_str(row[phone_i] if phone_i < len(row) else None)
            password = _cell_str(row[password_i] if password_i < len(row) else None)
            patient_id = _cell_str(row[patient_i] if patient_i < len(row) else None)
            if not phone or not patient_id:
                continue
            if not patient_id.isdigit():
                continue
            parsed.append({"phone": phone, "password": password, "patient_id": patient_id})

        workbook.close()
        return jsonify({"rows": parsed})
    except Exception as exc:
        return jsonify({"error": str(exc)}), 400


def _parse_run_payload() -> tuple[list, str, str]:
    """Parse rows/mode/drive_link from multipart form or JSON."""
    if request.form and request.form.get("rows") is not None:
        try:
            rows = json.loads(request.form.get("rows") or "[]")
        except json.JSONDecodeError as exc:
            raise ValueError(f"Invalid rows JSON: {exc}") from exc
        mode = (request.form.get("mode") or "drive").strip() or "drive"
        drive_link = (request.form.get("drive_link") or "").strip()
        return rows, mode, drive_link

    body = request.get_json(silent=True) or {}
    rows = body.get("rows") or []
    mode = (body.get("mode") or "drive").strip() or "drive"
    drive_link = (body.get("drive_link") or "").strip()
    return rows, mode, drive_link


def _uploaded_pair(kind: str, index: int) -> tuple[str, bytes] | None:
    files = request.files.getlist(f"{kind}_{index}")
    if not files:
        single = request.files.get(f"{kind}_{index}")
        files = [single] if single else []
    for item in files:
        if item and item.filename:
            return (item.filename, item.read())
    return None


@load_test_bp.route("/api/load-test/verify-drive", methods=["POST"])
def verify_drive():
    body = request.get_json(silent=True) or {}
    folder_id = _extract_folder_id(body.get("drive_link") or "")
    if not folder_id:
        return jsonify({"error": "Could not extract Google Drive folder ID from the link"}), 400
    try:
        config = get_config()
        service = get_drive_service(config)
        files = _list_folder_files(service, folder_id)
        return jsonify({"file_count": len(files)})
    except Exception as exc:
        return jsonify({"error": str(exc)}), 400


@load_test_bp.route("/api/load-test/excel-template", methods=["GET"])
def excel_template():
    wb = Workbook()
    ws = wb.active
    ws.title = "Doctors"
    headers = ["phone", "password", "patient_id"]
    header_fill = PatternFill(start_color="6C5CE7", end_color="6C5CE7", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    grey_font = Font(color="9CA3AF")

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    examples = [
        ("9876543210", "your_password", "101"),
        ("9876543211", "your_password", "102"),
        ("9876543212", "your_password", "103"),
    ]
    for row_idx, values in enumerate(examples, 2):
        for col_idx, value in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = grey_font

    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 16

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(
        buf,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name="lt_template.xlsx",
    )


@load_test_bp.route("/api/load-test/run", methods=["POST"])
def run_load_test():
    try:
        rows, mode, drive_link = _parse_run_payload()
    except ValueError as exc:
        return jsonify({"error": str(exc)}), 400

    if not isinstance(rows, list) or not rows:
        return jsonify({"error": "No doctor–patient rows provided"}), 400
    if mode not in ("drive", "manual"):
        return jsonify({"error": "mode must be 'drive' or 'manual'"}), 400

    try:
        config = get_config()
    except Exception as exc:
        return jsonify({"error": f"Config error: {exc}"}), 500

    django_url, flask_url = _urls_from_config(config)
    if not django_url or not flask_url:
        return jsonify({"error": "django_base_url and flask_transcribe_url must be set in config"}), 500

    llm_settings = config.get("llm_settings") or {}
    template_type = str(llm_settings.get("template_type") or "soap").lower()
    cfg = {
        "django_url": django_url,
        "flask_url": flask_url,
        "language": "en",
        "llm": llm_settings.get("llm_model", "OpenAI"),
        "stt_model": llm_settings.get("stt_model", "Bhasini"),
        "translate_model": llm_settings.get("translation_type", "Bhasini"),
        "template_id": llm_settings.get("template_id", 1),
        "template": DISCHARGE_TEMPLATE if template_type == "discharge" else SOAP_TEMPLATE,
        "duration": 2.0,
        "test_run_id": str(uuid.uuid4()),
        "mode": mode,
    }

    drive_svc = None
    drive_files: list[dict] = []
    if mode == "drive":
        folder_id = _extract_folder_id(drive_link)
        if not folder_id:
            return jsonify({"error": "Could not extract Google Drive folder ID from the link"}), 400
        try:
            drive_svc = get_drive_service(config)
            drive_files = _list_folder_files(drive_svc, folder_id)
        except Exception as exc:
            return jsonify({"error": f"Google Drive error: {exc}"}), 400

    manual_audio: dict[int, tuple[str, bytes]] = {}
    manual_gt: dict[int, tuple[str, bytes]] = {}
    if mode == "manual":
        for idx in range(len(rows)):
            audio_pair = _uploaded_pair("audio", idx)
            if audio_pair:
                manual_audio[idx] = audio_pair
            gt_pair = _uploaded_pair("gt", idx)
            if gt_pair:
                manual_gt[idx] = gt_pair

    q: queue.Queue = queue.Queue()
    passed = [0]

    def run_all():
        _emit(q, {"type": "run_id", "run_id": cfg["test_run_id"]})
        workers = max(1, len(rows))
        with ThreadPoolExecutor(max_workers=workers) as pool:
            futures = [
                pool.submit(
                    _run_row,
                    idx,
                    row,
                    drive_files,
                    drive_svc,
                    cfg,
                    q,
                    manual_audio.get(idx),
                    manual_gt.get(idx),
                )
                for idx, row in enumerate(rows)
            ]
            for future in as_completed(futures):
                try:
                    if future.result() == "pass":
                        passed[0] += 1
                except Exception as exc:
                    log.error("Row worker crashed: %s", exc)
        _emit(q, {"type": "done", "passed": passed[0], "total": len(rows)})
        q.put(None)

    threading.Thread(target=run_all, daemon=True).start()

    def generate():
        yield ": " + " " * 2048 + "\n\n"
        while True:
            item = q.get()
            if item is None:
                break
            yield f"data: {json.dumps(item, ensure_ascii=False)}\n\n"

    return Response(
        stream_with_context(generate()),
        mimetype="text/event-stream",
        headers={"Cache-Control": "no-cache", "X-Accel-Buffering": "no"},
    )


@load_test_bp.route("/api/load-test/export", methods=["POST"])
def export_results():
    try:
        data = request.get_json(silent=True) or {}
        runs = data.get("runs", [])
        if runs and isinstance(runs[0], dict):
            runs = [runs]
        if not runs or all(len(r) == 0 for r in runs):
            return jsonify({"error": "No results to export"}), 400

        wb = Workbook()
        ws = wb.active
        ws.title = "Test Results"

        headers = [
            "Doctor ID",
            "Patient ID",
            "Audio ID",
            "Summary ID",
            "Audio Duration (s)",
            "Language",
            "Login Response Time (s)",
            "Profile Response Time (s)",
            "Patient Data Response Time (s)",
            "Audio Preprocessing Latency (s)",
            "STT Latency (s)",
            "Translation Latency (s)",
            "LLM Latency (s)",
            "Backend Processing Latency (s)",
            "Response Time (s)",
            "Clock Time",
            "Archival Upload Response Time (s)",
            "Summary Store Response Time (s)",
            "Status",
        ]
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        pass_fill = PatternFill(start_color="D1FAE5", end_color="D1FAE5", fill_type="solid")
        fail_fill = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")
        sep_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        sep_font = Font(bold=True, color="1F3864")

        def write_header(row_idx):
            for col_idx, header in enumerate(headers, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=header)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center")

        def _t(result, key):
            value = result.get(key)
            return round(value, 5) if value is not None else None

        def write_data_row(row_idx, result):
            is_fail = result.get("status") == "fail"

            def _or_dash(value):
                if value is not None:
                    return value
                return "--" if is_fail else None

            values = [
                result.get("doctor_id"),
                result.get("patient_id"),
                _or_dash(result.get("audio_id")),
                _or_dash(result.get("summary_id")),
                result.get("audio_duration"),
                result.get("language"),
                _or_dash(_t(result, "step1_time")),
                _or_dash(_t(result, "step1b_time")),
                _or_dash(_t(result, "patient_metadata_time")),
                _or_dash(_t(result, "audio_processing_time")),
                _or_dash(_t(result, "transcription_time")),
                _or_dash(_t(result, "translation_time")),
                _or_dash(_t(result, "llm_time")),
                _or_dash(_t(result, "flask_total_time")),
                _or_dash(_t(result, "transcribe_rtt") or _t(result, "user_percieved_summary_latency")),
                None,
                _or_dash(_t(result, "audio_upload_time")),
                _or_dash(_t(result, "summary_store_time")),
                None,
            ]
            clock_start = result.get("clock_start", "")
            clock_end = result.get("clock_end", "")
            values[15] = f"{clock_start} - {clock_end}" if clock_start and clock_end else ("--" if is_fail else None)

            row_fill = fail_fill if is_fail else pass_fill
            for col_idx, value in enumerate(values[:-1], 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.fill = row_fill

            status_cell = ws.cell(row=row_idx, column=19)
            if is_fail:
                status_cell.value = "Fail"
                status_cell.font = Font(bold=True, color="FF0000")
                status_cell.fill = fail_fill
            else:
                status_cell.value = "Pass"
                status_cell.font = Font(bold=True, color="006100")
                status_cell.fill = pass_fill

        def pid_sort_key(result):
            value = str(result.get("patient_id", ""))
            return int(value) if value.isdigit() else value

        cur_row = 1
        num_cols = len(headers)
        for round_idx, round_results in enumerate(runs):
            if not round_results:
                continue
            if round_idx > 0:
                cur_row += 1
                sep_cell = ws.cell(row=cur_row, column=1, value=f"Run {round_idx + 1}")
                sep_cell.fill = sep_fill
                sep_cell.font = sep_font
                for col in range(2, num_cols + 1):
                    ws.cell(row=cur_row, column=col).fill = sep_fill
                cur_row += 1

            write_header(cur_row)
            cur_row += 1
            for result in sorted(round_results, key=pid_sort_key):
                write_data_row(cur_row, result)
                cur_row += 1

        widths = {
            "A": 12, "B": 12, "C": 12, "D": 12, "E": 18, "F": 14,
            "G": 16, "H": 22, "I": 24, "J": 18, "K": 20, "L": 22,
            "M": 14, "N": 18, "O": 14, "P": 20, "Q": 24, "R": 22, "S": 10,
        }
        for col, width in widths.items():
            ws.column_dimensions[col].width = width

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        stamp = datetime.now().strftime("%Y-%m-%d")
        return send_file(
            buf,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            as_attachment=True,
            download_name=f"lt_results_{stamp}.xlsx",
        )
    except Exception as exc:
        return jsonify({"error": str(exc)}), 500
