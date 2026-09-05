"""Ground Truth vs Generated SOAP comparison report.

Walks the SOAP consult schema field-by-field, applies the eight validation
cases (match, both-null, missing, extra, type/format/value mismatch), and
renders a downloadable report for frontend display.
"""

from __future__ import annotations

import csv
import io
import json
import re
from datetime import datetime
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import cm
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

from medsum_testing.backend.services.gt_comparison_table import generated_soap
from medsum_testing.backend.services.medsum_api import SOAP_CONSULT_TEMPLATE

MISSING = object()

STATUS_PASS = "PASS"
STATUS_FAIL = "FAIL"
STATUS_PARTIAL = "PARTIAL PASS"

CAT_MATCH = "MATCH"
CAT_BOTH_NULL = "BOTH_NULL"
CAT_CORRECTLY_ABSENT = "CORRECTLY_ABSENT"
CAT_VALUE_MISMATCH = "VALUE_MISMATCH"
CAT_TYPE_MISMATCH = "TYPE_MISMATCH"
CAT_FORMAT_MISMATCH = "FORMAT_MISMATCH"
CAT_MISSING_FROM_GENERATED = "MISSING_FROM_GENERATED"
CAT_EXTRA_IN_GENERATED = "EXTRA_IN_GENERATED"

PASS_CATEGORIES = frozenset({CAT_MATCH, CAT_BOTH_NULL, CAT_CORRECTLY_ABSENT})
SCHEMA_PENALTY_CATEGORIES = frozenset(
    {CAT_TYPE_MISMATCH, CAT_MISSING_FROM_GENERATED, CAT_EXTRA_IN_GENERATED}
)

PASS_THRESHOLD = 80.0
PARTIAL_THRESHOLD = 60.0

ERROR_MARK_PASS = "✓"
ERROR_MARK_FAIL = "✗"

SUPPORTED_FORMATS = ("json", "csv", "excel", "html", "pdf")

TABLE_COLUMNS = (
    "FIELD NAME",
    "GROUND TRUTH VALUE",
    "GENERATED VALUE",
    "FOUND IN GROUND TRUTH",
    "ERROR FOUND",
    "ERROR DESCRIPTION",
    "STATUS",
)

MED_FIELDS = (
    ("drug_name", "drug_name"),
    ("dose", "dose"),
    ("schedule", "schedule"),
    ("duration", "duration"),
    ("instructions", "instruct"),
    ("snomed_ct_id", "snomed_id"),
)
MED_KEY_ALIASES = {
    "drug_name": ("drug_name", "name"),
    "dose": ("dose",),
    "schedule": ("schedule", "frequency"),
    "duration": ("duration",),
    "instructions": ("instructions", "instruct"),
    "snomed_ct_id": ("snomed_ct_id", "snomed_id"),
}
MED_LEAF_LABELS = {
    "drug_name": "Drug Name",
    "name": "Drug Name",
    "dose": "Dose",
    "schedule": "Schedule",
    "frequency": "Schedule",
    "duration": "Duration",
    "instructions": "Instructions",
    "instruct": "Instructions",
    "snomed_ct_id": "SNOMED CT ID",
    "snomed_id": "SNOMED CT ID",
}
FIELD_LABELS = {
    "chief_complaint": "Chief Complaint",
    "history_of_present_illness": "History of Present Illness",
    "past_medical_history": "Past Medical History",
    "medications": "Medications",
    "allergies": "Allergies",
    "social_history": "Social History",
    "family_history": "Family History",
    "blood_group": "Blood Group",
    "blood_pressure": "Blood Pressure",
    "heart_rate": "Heart Rate",
    "respiratory_rate": "Respiratory Rate",
    "temperature": "Temperature",
    "spo2": "SpO2",
    "heart": "Heart",
    "height": "Height",
    "weight": "Weight",
    "diagnosis": "Diagnosis",
    "type": "Type",
    "status": "Status",
    "reasoning": "Reasoning",
    "activity": "Activity",
    "investigations": "Investigations",
    "education": "Education",
    "follow_up": "Follow Up",
    "summary": "Summary",
    **MED_LEAF_LABELS,
}

SUBJECTIVE_FIELDS = tuple(SOAP_CONSULT_TEMPLATE["subjective"].keys())
VITALS_FIELDS = tuple(SOAP_CONSULT_TEMPLATE["objective"]["vitals"].keys())
PHYSICAL_EXAM_FIELDS = tuple(SOAP_CONSULT_TEMPLATE["objective"]["physical_exam"].keys())
ASSESSMENT_FIELDS = tuple(SOAP_CONSULT_TEMPLATE["assessment"].keys())
PLAN_OTHER_FIELDS = ("activity", "investigations", "education", "follow_up")

REPORT_SECTIONS = (
    ("subjective", "SUBJECTIVE SECTION", "subjective", SUBJECTIVE_FIELDS),
    ("objective_vitals", "OBJECTIVE SECTION → VITALS", "objective.vitals", VITALS_FIELDS),
    (
        "objective_physical_exam",
        "OBJECTIVE SECTION → PHYSICAL_EXAM",
        "objective.physical_exam",
        PHYSICAL_EXAM_FIELDS,
    ),
    ("assessment", "ASSESSMENT SECTION", "assessment", ASSESSMENT_FIELDS),
    ("plan_medications", "PLAN SECTION → MEDICATIONS", "plan.medications", ()),
    ("plan_other", "PLAN SECTION → OTHER", "plan", PLAN_OTHER_FIELDS),
    ("summary", "SUMMARY", "", ("summary",)),
)

_UNIT_RE = re.compile(
    r"(?ix)"
    r"(mm\s*hg|mmhg|bpm|°\s*[cf]|degrees?\s*[cf]|"
    r"breaths?\s*(?:per|/)\s*min(?:ute)?s?|breaths?/min|"
    r"beats?\s*(?:per|/)\s*min(?:ute)?s?|"
    r"\b(?:cm|kg|mg|mcg|ml|mmol|%|[cf])\b)"
)
_DISPLAY_LIMIT = 36


def _as_dict(value: Any) -> dict:
    return value if isinstance(value, dict) else {}


def _as_soap(value: Any) -> dict:
    if isinstance(value, dict):
        return value
    if isinstance(value, str) and value.strip():
        try:
            parsed = json.loads(value)
        except json.JSONDecodeError:
            return {}
        return parsed if isinstance(parsed, dict) else {}
    return {}


def _path_parts(path: str) -> list[Any]:
    parts: list[Any] = []
    for chunk in path.split("."):
        if not chunk:
            continue
        match = re.fullmatch(r"([^\[\]]+)\[(\d+)\]", chunk)
        if match:
            parts.append(match.group(1))
            parts.append(int(match.group(2)))
        else:
            parts.append(int(chunk) if chunk.isdigit() else chunk)
    return parts


def _join_path(prefix: str, key: str) -> str:
    return f"{prefix}.{key}" if prefix else key


def _get_at(node: Any, path: str) -> tuple[bool, Any]:
    current = node
    for part in _path_parts(path):
        if isinstance(part, int):
            if not isinstance(current, list) or part < 0 or part >= len(current):
                return False, MISSING
            current = current[part]
            continue
        if not isinstance(current, dict) or part not in current:
            return False, MISSING
        current = current[part]
    return True, current


def _is_blank(value: Any) -> bool:
    if value is MISSING or value is None:
        return True
    if isinstance(value, str):
        return not value.strip()
    return False


def _value_kind(value: Any) -> str | None:
    if value is MISSING or value is None:
        return None
    if isinstance(value, bool):
        return "boolean"
    if isinstance(value, (int, float)):
        return "number"
    if isinstance(value, str):
        return "string"
    if isinstance(value, list):
        return "array"
    if isinstance(value, dict):
        return "object"
    return type(value).__name__


def _title_case_key(key: str) -> str:
    return str(key or "").replace("_", " ").strip().title()


def human_field_name(field: str, path: str = "") -> str:
    """Readable labels such as 'Medication 1 Drug Name' / 'Chief Complaint'."""
    raw = str(field or path or "").strip()
    med = re.search(r"medications\[(\d+)\]\.([^.]+)$", raw)
    if not med:
        med = re.search(r"medications\[(\d+)\]\.([^.]+)$", str(path or ""))
    if med:
        index = int(med.group(1)) + 1
        leaf = med.group(2)
        label = MED_LEAF_LABELS.get(leaf) or FIELD_LABELS.get(leaf) or _title_case_key(leaf)
        return f"Medication {index} {label}"
    leaf = raw.rsplit(".", 1)[-1]
    leaf = re.sub(r"\[\d+\]$", "", leaf)
    if leaf in FIELD_LABELS:
        return FIELD_LABELS[leaf]
    if leaf in MED_LEAF_LABELS:
        return MED_LEAF_LABELS[leaf]
    return _title_case_key(leaf) or raw


def _data_type_label(
    *,
    gen_present: bool,
    gen_value: Any,
    gen_kind: str | None,
    gt_present: bool,
    gt_value: Any,
    gt_kind: str | None,
) -> str:
    if gen_present:
        if gen_value is None:
            return "null"
        return gen_kind or _value_kind(gen_value) or "null"
    if gt_present:
        if gt_value is None:
            return "null"
        return gt_kind or _value_kind(gt_value) or "null"
    return "null"


def _values_equal(left: Any, right: Any) -> bool:
    if isinstance(left, (int, float)) and isinstance(right, (int, float)):
        return float(left) == float(right)
    return left == right


def _core_text(value: Any) -> str:
    text = str(value).strip()
    text = _UNIT_RE.sub("", text)
    return re.sub(r"[\s,;:()]+", "", text).lower()


def _is_format_mismatch(left: Any, right: Any) -> bool:
    if isinstance(left, (list, dict)) or isinstance(right, (list, dict)):
        return False
    left_text = str(left).strip()
    right_text = str(right).strip()
    if left_text == right_text:
        return False
    if left_text.lower() == right_text.lower():
        return True
    left_core = _core_text(left)
    right_core = _core_text(right)
    if not left_core or not right_core or left_core != right_core:
        return False
    return True


def _quote_display(value: Any, *, present: bool, kind: str | None = None) -> str:
    if not present:
        return "[NULL/MISSING]"
    if value is None:
        return "null"
    shown_kind = kind or _value_kind(value)
    if shown_kind == "number":
        if isinstance(value, float) and value.is_integer():
            return f"{int(value)} (NUMBER)"
        return f"{value} (NUMBER)"
    if shown_kind == "boolean":
        return "true" if value else "false"
    if shown_kind in ("array", "object"):
        return json.dumps(value, ensure_ascii=False)
    if isinstance(value, str):
        if value == "":
            return '""'
        return json.dumps(value, ensure_ascii=False)
    return str(value)


def _truncate_display(text: str) -> tuple[str, bool]:
    raw = text or ""
    if len(raw) <= _DISPLAY_LIMIT:
        return raw, False
    return raw[: _DISPLAY_LIMIT - 2] + "..", True


def _in_gt_label(present: bool, value: Any) -> str:
    if not present:
        return "NO"
    if value is None:
        return "YES"
    return "YES"


def classify_field(
    *,
    gt_present: bool,
    gt_value: Any,
    gen_present: bool,
    gen_value: Any,
    in_schema: bool,
) -> dict[str, Any]:
    """Apply the eight SOAP GT-vs-generated validation cases."""
    gt_blank = (not gt_present) or _is_blank(gt_value)
    gen_blank = (not gen_present) or _is_blank(gen_value)
    in_gt = _in_gt_label(gt_present, gt_value if gt_present else MISSING)
    gt_kind = _value_kind(gt_value) if gt_present else None
    gen_kind = _value_kind(gen_value) if gen_present else None

    def _row(
        category: str,
        description: str,
        *,
        status: str | None = None,
    ) -> dict[str, Any]:
        passed = category in PASS_CATEGORIES
        return {
            "in_gt": in_gt,
            "in_gt_bool": gt_present,
            "in_schema": in_schema,
            "category": category,
            "error_found": ERROR_MARK_PASS if passed else ERROR_MARK_FAIL,
            "error_found_pass": passed,
            "error_description": description,
            "status": status or (STATUS_PASS if passed else STATUS_FAIL),
            "passed": passed,
            "gt_kind": gt_kind,
            "gen_kind": gen_kind,
        }

    if gt_blank and gen_blank:
        if gt_present or gen_present:
            if gt_present and gen_present and gt_value is None and gen_value is None:
                return _row(CAT_BOTH_NULL, "Both versions correctly null")
            if (
                gt_present
                and gen_present
                and isinstance(gt_value, str)
                and isinstance(gen_value, str)
                and gt_value == ""
                and gen_value == ""
            ):
                return _row(CAT_BOTH_NULL, "Both correctly empty")
            return _row(CAT_BOTH_NULL, "Both versions correctly null")
        return _row(CAT_CORRECTLY_ABSENT, "Field correctly absent in both versions")

    if not gt_blank and gen_blank:
        return _row(CAT_MISSING_FROM_GENERATED, "Required field missing from generated data")

    if gt_blank and not gen_blank:
        if not in_schema and not gt_present:
            return _row(
                CAT_EXTRA_IN_GENERATED,
                "Field exists in generated but not in ground truth or schema",
            )
        return _row(CAT_VALUE_MISMATCH, "Ground truth is null, but generated has value")

    assert not gt_blank and not gen_blank
    if gt_kind and gen_kind and gt_kind != gen_kind:
        return _row(
            CAT_TYPE_MISMATCH,
            f"Type mismatch: ground truth is {gt_kind}, generated is {gen_kind}",
        )
    if _values_equal(gt_value, gen_value):
        if isinstance(gt_value, list) and isinstance(gen_value, list):
            if not gt_value and not gen_value:
                return _row(CAT_MATCH, "Arrays match (both empty)")
            return _row(CAT_MATCH, "Arrays match exactly")
        return _row(CAT_MATCH, "Values match exactly")
    if _is_format_mismatch(gt_value, gen_value):
        return _row(CAT_FORMAT_MISMATCH, _format_description(gt_value, gen_value))
    return _row(
        CAT_VALUE_MISMATCH,
        f"Values do not match: expected {_quote_display(gt_value, present=True)}, "
        f"got {_quote_display(gen_value, present=True)}",
    )


def _format_description(gt_value: Any, gen_value: Any) -> str:
    gt_text = str(gt_value)
    gen_text = str(gen_value)
    if gt_text.lower() == gen_text.lower() and gt_text != gen_text:
        return "Format mismatch: letter casing differs"
    if _UNIT_RE.search(gt_text) and not _UNIT_RE.search(gen_text):
        return "Format mismatch: missing unit"
    if _UNIT_RE.search(gen_text) and not _UNIT_RE.search(gt_text):
        return "Format mismatch: extra unit text"
    return "Format mismatch: same value, different format"


def _display_row(
    field: str,
    gt_present: bool,
    gt_value: Any,
    gen_present: bool,
    gen_value: Any,
    classified: dict[str, Any],
) -> dict[str, Any]:
    gt_display = _quote_display(gt_value, present=gt_present, kind=classified.get("gt_kind"))
    gen_display = _quote_display(gen_value, present=gen_present, kind=classified.get("gen_kind"))
    gt_shown, gt_cut = _truncate_display(gt_display)
    gen_shown, gen_cut = _truncate_display(gen_display)
    description = classified["error_description"]
    if classified["category"] == CAT_MATCH and (gt_cut or gen_cut):
        description = "Values match (truncated for display)"
    return {
        "field": field,
        "display_field": human_field_name(field),
        "ground_truth": gt_shown,
        "ground_truth_full": gt_display,
        "generated": gen_shown,
        "generated_full": gen_display,
        "ground_truth_raw": None if not gt_present else gt_value,
        "generated_raw": None if not gen_present else gen_value,
        **classified,
        "error_description": description,
        "data_type": _data_type_label(
            gen_present=gen_present,
            gen_value=gen_value,
            gen_kind=classified.get("gen_kind"),
            gt_present=gt_present,
            gt_value=gt_value,
            gt_kind=classified.get("gt_kind"),
        ),
        "status_display": (
            f"{classified['status']} {ERROR_MARK_PASS}"
            if classified["passed"]
            else f"{classified['status']} {ERROR_MARK_FAIL}"
        ),
    }


def _compare_path(
    field: str,
    path: str,
    gt_soap: dict,
    gen_soap: dict,
    *,
    in_schema: bool,
) -> dict[str, Any]:
    gt_present, gt_value = _get_at(gt_soap, path)
    gen_present, gen_value = _get_at(gen_soap, path)
    classified = classify_field(
        gt_present=gt_present,
        gt_value=gt_value,
        gen_present=gen_present,
        gen_value=gen_value,
        in_schema=in_schema,
    )
    row = _display_row(field, gt_present, gt_value, gen_present, gen_value, classified)
    row["path"] = path
    return row


def _assessment_root(soap: dict) -> dict:
    raw = soap.get("assessment")
    if isinstance(raw, list):
        for item in raw:
            if isinstance(item, dict):
                return item
        return {}
    return raw if isinstance(raw, dict) else {}


def _med_list(value: Any) -> list | None:
    if isinstance(value, list):
        return value
    if isinstance(value, dict) and ("drug_name" in value or "dose" in value or "name" in value):
        return [value]
    return None


def _med_field_value(med: Any, canonical: str) -> tuple[bool, Any]:
    if not isinstance(med, dict):
        return False, MISSING
    for alias in MED_KEY_ALIASES[canonical]:
        if alias in med:
            return True, med[alias]
    return False, MISSING


def _medication_rows(gt_soap: dict, gen_soap: dict, seen: set[str]) -> list[dict[str, Any]]:
    gt_present, gt_val = _get_at(gt_soap, "plan.medications")
    gen_present, gen_val = _get_at(gen_soap, "plan.medications")
    gt_list = _med_list(gt_val) if gt_present else None
    gen_list = _med_list(gen_val) if gen_present else None

    if gt_list is None and gen_list is None:
        seen.add("plan.medications")
        return [_compare_path("medications", "plan.medications", gt_soap, gen_soap, in_schema=True)]

    gt_items = gt_list or []
    gen_items = gen_list or []
    count = max(len(gt_items), len(gen_items))
    if count == 0:
        seen.add("plan.medications")
        classified = classify_field(
            gt_present=gt_present,
            gt_value=gt_val if gt_present else MISSING,
            gen_present=gen_present,
            gen_value=gen_val if gen_present else MISSING,
            in_schema=True,
        )
        row = _display_row(
            "medications",
            gt_present,
            gt_val if gt_present else MISSING,
            gen_present,
            gen_val if gen_present else MISSING,
            classified,
        )
        row["path"] = "plan.medications"
        return [row]

    rows: list[dict[str, Any]] = []
    seen.add("plan.medications")
    for index in range(count):
        gt_med = gt_items[index] if index < len(gt_items) else MISSING
        gen_med = gen_items[index] if index < len(gen_items) else MISSING
        gt_med_present = isinstance(gt_med, dict)
        gen_med_present = isinstance(gen_med, dict)
        extra_keys: list[str] = []
        if gt_med_present:
            extra_keys.extend(gt_med.keys())
        if gen_med_present:
            extra_keys.extend(gen_med.keys())
        for canonical, display in MED_FIELDS:
            field = f"medications[{index}].{display}"
            path = f"plan.medications[{index}].{canonical}"
            gt_f_present, gt_f_val = (
                _med_field_value(gt_med, canonical) if gt_med_present else (False, MISSING)
            )
            gen_f_present, gen_f_val = (
                _med_field_value(gen_med, canonical) if gen_med_present else (False, MISSING)
            )
            classified = classify_field(
                gt_present=gt_f_present,
                gt_value=gt_f_val,
                gen_present=gen_f_present,
                gen_value=gen_f_val,
                in_schema=True,
            )
            row = _display_row(field, gt_f_present, gt_f_val, gen_f_present, gen_f_val, classified)
            row["path"] = path
            row["display_field"] = human_field_name(field, path)
            rows.append(row)
            seen.add(path)
            for alias in MED_KEY_ALIASES[canonical]:
                seen.add(f"plan.medications[{index}].{alias}")
        known = {alias for aliases in MED_KEY_ALIASES.values() for alias in aliases}
        for extra in dict.fromkeys(extra_keys):
            extra_path = f"plan.medications[{index}].{extra}"
            if extra in known or extra_path in seen:
                continue
            gt_x = extra in gt_med if gt_med_present else False
            gen_x = extra in gen_med if gen_med_present else False
            classified = classify_field(
                gt_present=gt_x,
                gt_value=gt_med.get(extra) if gt_x else MISSING,
                gen_present=gen_x,
                gen_value=gen_med.get(extra) if gen_x else MISSING,
                in_schema=False,
            )
            field = f"medications[{index}].{extra}"
            row = _display_row(
                field,
                gt_x,
                gt_med.get(extra) if gt_x else MISSING,
                gen_x,
                gen_med.get(extra) if gen_x else MISSING,
                classified,
            )
            row["path"] = extra_path
            rows.append(row)
            seen.add(extra_path)
    return rows


def _walk_extra_leaves(node: Any, prefix: str, out: list[tuple[str, bool, Any]]) -> None:
    if isinstance(node, dict):
        for key, val in node.items():
            path = _join_path(prefix, str(key))
            if isinstance(val, dict):
                _walk_extra_leaves(val, path, out)
                continue
            if isinstance(val, list) and val and isinstance(val[0], dict):
                for idx, item in enumerate(val):
                    _walk_extra_leaves(item, f"{path}[{idx}]", out)
                continue
            out.append((path, True, val))
        return
    if prefix:
        out.append((prefix, True, node))


def _extra_rows(gt_soap: dict, gen_soap: dict, seen: set[str]) -> list[dict[str, Any]]:
    leaves: dict[str, dict[str, Any]] = {}
    gt_leaves: list[tuple[str, bool, Any]] = []
    gen_leaves: list[tuple[str, bool, Any]] = []
    _walk_extra_leaves(gt_soap, "", gt_leaves)
    _walk_extra_leaves(gen_soap, "", gen_leaves)
    for path, present, value in gt_leaves:
        slot = leaves.setdefault(path, {"gt_p": False, "gt": MISSING, "gen_p": False, "gen": MISSING})
        slot["gt_p"] = present
        slot["gt"] = value
    for path, present, value in gen_leaves:
        slot = leaves.setdefault(path, {"gt_p": False, "gt": MISSING, "gen_p": False, "gen": MISSING})
        slot["gen_p"] = present
        slot["gen"] = value
    rows: list[dict[str, Any]] = []
    for path in sorted(leaves):
        if path in seen:
            continue
        if any(path == found or path.startswith(found + ".") or path.startswith(found + "[") for found in seen):
            continue
        if any(found.startswith(path + ".") or found.startswith(path + "[") for found in seen):
            continue
        slot = leaves[path]
        classified = classify_field(
            gt_present=slot["gt_p"],
            gt_value=slot["gt"],
            gen_present=slot["gen_p"],
            gen_value=slot["gen"],
            in_schema=False,
        )
        row = _display_row(path, slot["gt_p"], slot["gt"], slot["gen_p"], slot["gen"], classified)
        row["path"] = path
        rows.append(row)
        seen.add(path)
    return rows


def _section_for_path(path: str) -> str:
    if path == "summary" or path.startswith("summary."):
        return "summary"
    if path.startswith("plan.medications"):
        return "plan_medications"
    if path.startswith("plan."):
        return "plan_other"
    if path.startswith("objective.vitals"):
        return "objective_vitals"
    if path.startswith("objective.physical_exam"):
        return "objective_physical_exam"
    if path.startswith("objective."):
        return "objective_physical_exam"
    if path.startswith("assessment"):
        return "assessment"
    if path.startswith("subjective"):
        return "subjective"
    return "extra"


def _result_meta(result: dict | None) -> dict[str, str]:
    data = _as_dict(result)
    test_id = str(data.get("tc_ref") or data.get("test_case_id") or data.get("test_id") or "")
    name = str(
        data.get("audio_filename")
        or data.get("filename")
        or data.get("drive_audio_filename")
        or ""
    )
    if name.lower().endswith((".mp3", ".wav", ".m4a", ".ogg", ".webm")):
        name = name.rsplit(".", 1)[0]
    return {
        "test_case_id": test_id or "N/A",
        "test_name": name or "N/A",
        "test_id": str(data.get("test_id") or ""),
    }


def pct_rate(count: int, total: int) -> float:
    if not total:
        return 0.0
    return round(100.0 * count / total, 1)


def threshold_status(compliance: float) -> tuple[str, str]:
    """Map overall compliance onto PASS (≥80) / PARTIAL (60–80) / FAIL (<60)."""
    try:
        score = float(compliance)
    except (TypeError, ValueError):
        return STATUS_FAIL, "✗"
    if score >= PASS_THRESHOLD:
        return STATUS_PASS, "✓"
    if score >= PARTIAL_THRESHOLD:
        return STATUS_PARTIAL, "⚠"
    return STATUS_FAIL, "✗"


def flatten_report_rows(report: dict[str, Any] | None) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    for section in (report or {}).get("sections") or []:
        rows.extend(section.get("rows") or [])
    return rows


def _error_impact(category: str) -> str:
    if category in {CAT_TYPE_MISMATCH, CAT_MISSING_FROM_GENERATED, CAT_VALUE_MISMATCH}:
        return "HIGH"
    return "MEDIUM"


def _error_recommendation(row: dict[str, Any]) -> str:
    category = row.get("category")
    path = str(row.get("path") or row.get("field") or "")
    if category == CAT_TYPE_MISMATCH:
        if "vitals" in path or row.get("gt_kind") == "string":
            return "Ensure all vital values are returned as strings, not numbers"
        return (
            f"Keep {row.get('field')} as {row.get('gt_kind') or 'string'}, "
            f"not {row.get('gen_kind')}"
        )
    if category == CAT_FORMAT_MISMATCH:
        if "temperature" in path:
            return "Include temperature unit (°C) in value"
        if "vitals" in path:
            return "Include unit descriptions in vital measurements"
        return "Include unit descriptions in formatted values"
    if category == CAT_MISSING_FROM_GENERATED:
        return "Validate all fields against schema before API response"
    if category == CAT_EXTRA_IN_GENERATED:
        return "Do not emit fields that are not in the SOAP schema or ground truth"
    if category == CAT_VALUE_MISMATCH:
        return "Align generated values with ground truth"
    return _action_item(row)


def build_error_analysis(rows: list[dict[str, Any]] | None) -> list[dict[str, Any]]:
    errors: list[dict[str, Any]] = []
    for row in rows or []:
        if row.get("passed"):
            continue
        errors.append(
            {
                "field": row.get("path") or row.get("field"),
                "display_field": row.get("field"),
                "error_type": str(row.get("category") or "").replace("_", " "),
                "category": row.get("category"),
                "ground_truth": row.get("ground_truth_full") or row.get("ground_truth"),
                "generated": row.get("generated_full") or row.get("generated"),
                "description": row.get("error_description"),
                "impact": _error_impact(str(row.get("category") or "")),
                "recommendation": _error_recommendation(row),
            }
        )
    return errors


def _http_status(data: dict) -> str:
    flask_error = data.get("flask_error")
    if flask_error:
        return "500 Error"
    status = str(data.get("status") or "").lower()
    if status in {"failed", "error"}:
        return "500 Error"
    if data.get("transcription_result") or data.get("soap_generated"):
        return "200 OK"
    return "unavailable"


def _processing_time(data: dict) -> str:
    tr = _as_dict(data.get("transcription_result"))
    nested = tr.get("time") if isinstance(tr.get("time"), dict) else {}
    seconds = (
        tr.get("total-time")
        or tr.get("llm-time")
        or nested.get("total")
        or nested.get("llm")
        or data.get("total_test_time_seconds")
    )
    if seconds is None or seconds == "":
        return "unavailable"
    try:
        value = float(seconds)
    except (TypeError, ValueError):
        return str(seconds)
    if value >= 10:
        return f"{round(value, 1)}s"
    return f"{int(round(value * 1000))}ms"


def _response_size(payload: Any) -> str:
    raw = json.dumps(payload, ensure_ascii=False, default=str) if payload else ""
    nbytes = len(raw.encode("utf-8")) if raw else 0
    if nbytes >= 1024:
        return f"{nbytes / 1024:.1f} KB"
    return f"{nbytes} B"


def build_case_metadata(result: dict | None, generated: dict | None = None) -> dict[str, Any]:
    data = _as_dict(result)
    meta = _result_meta(data)
    endpoint = str(
        data.get("flask_transcribe_endpoint")
        or "/transcribe"
    )
    method = "POST"
    gt_file = (
        data.get("uploaded_ground_truth_filename")
        or data.get("drive_audio_filename")
        or data.get("soap_gt_filename")
        or ""
    )
    if not gt_file and meta["test_case_id"] not in {"", "N/A"}:
        gt_file = f"gt_{meta['test_case_id']}.json"
    gen_file = f"gen_{meta['test_case_id']}.json" if meta["test_case_id"] not in {"", "N/A"} else ""
    soap_json = generated if generated is not None else generated_soap(data)
    return {
        "test_id": meta["test_case_id"],
        "test_name": meta["test_name"],
        "harness_test_id": meta["test_id"],
        "timestamp": str(data.get("timestamp") or data.get("session_datetime") or ""),
        "api_method": f"{method} {endpoint}",
        "api_method_verb": method,
        "endpoint": endpoint,
        "backend_source": "Transcribe API → SOAP Generation",
        "api_request_id": str(data.get("session_id") or data.get("job_id") or data.get("test_id") or ""),
        "response_code": _http_status(data),
        "processing_time": _processing_time(data),
        "ground_truth_file": str(gt_file or "unavailable"),
        "generated_file": gen_file or "unavailable",
        "response_size": _response_size(soap_json),
        "language": str(data.get("language") or ""),
        "batch_id": str(data.get("batch_id") or data.get("batch_ref") or ""),
        "audio_filename": str(data.get("audio_filename") or data.get("filename") or ""),
    }


def _action_item(row: dict[str, Any]) -> str:
    field = row["field"]
    category = row["category"]
    if category == CAT_TYPE_MISMATCH:
        expected = row.get("gt_kind") or "string"
        return f"Fix {field} type (should be {expected} not {row.get('gen_kind')})"
    if category == CAT_FORMAT_MISMATCH:
        if "missing unit" in row["error_description"].lower():
            return f"Add units to {field}"
        return f"Align format of {field} with ground truth"
    if category == CAT_MISSING_FROM_GENERATED:
        return f"Add required field {field} to generated data"
    if category == CAT_EXTRA_IN_GENERATED:
        return f"Remove extra field {field} from generated data"
    if category == CAT_VALUE_MISMATCH:
        return f"Fix {field}: {row['error_description']}"
    return f"Fix {field}"


def _summary_block(rows: list[dict[str, Any]]) -> dict[str, Any]:
    total = len(rows)
    passed = sum(1 for row in rows if row["passed"])
    failed = total - passed
    compliance = round(100.0 * passed / total, 1) if total else 0.0
    if total == 0:
        overall = "N/A"
    elif failed == 0:
        overall = STATUS_PASS
    elif passed == 0:
        overall = STATUS_FAIL
    else:
        overall = STATUS_PARTIAL

    def _cat(name: str) -> list[str]:
        return [row["field"] for row in rows if row["category"] == name]

    exact = _cat(CAT_MATCH)
    both_null = _cat(CAT_BOTH_NULL) + _cat(CAT_CORRECTLY_ABSENT)
    type_m = _cat(CAT_TYPE_MISMATCH)
    format_m = _cat(CAT_FORMAT_MISMATCH)
    value_m = _cat(CAT_VALUE_MISMATCH)
    missing = _cat(CAT_MISSING_FROM_GENERATED)
    extra = _cat(CAT_EXTRA_IN_GENERATED)
    action_items = [_action_item(row) for row in rows if not row["passed"]]
    type_n = len(type_m)
    missing_n = len(missing)
    extra_n = len(extra)
    schema_penalties = type_n + missing_n + extra_n
    schema_accuracy = pct_rate(max(total - schema_penalties, 0), total)
    gt_match = compliance
    band, mark = threshold_status(compliance)
    return {
        "total_fields": total,
        "passed": passed,
        "failed": failed,
        "passed_percent": round(100.0 * passed / total, 1) if total else 0.0,
        "failed_percent": round(100.0 * failed / total, 1) if total else 0.0,
        "compliance_percent": compliance,
        "overall_status": overall,
        "schema_accuracy": schema_accuracy,
        "ground_truth_match": gt_match,
        "type_error_rate": pct_rate(type_n, total),
        "missing_field_rate": pct_rate(missing_n, total),
        "extra_field_rate": pct_rate(extra_n, total),
        "threshold_status": band,
        "threshold_mark": mark,
        "pass_threshold": PASS_THRESHOLD,
        "partial_threshold": PARTIAL_THRESHOLD,
        "error_breakdown": {
            "exact_matches": len(exact),
            "exact_match_fields": exact,
            "type_mismatches": len(type_m),
            "type_mismatch_fields": type_m,
            "format_mismatches": len(format_m),
            "format_mismatch_fields": format_m,
            "value_mismatches": len(value_m),
            "value_mismatch_fields": value_m,
            "missing_from_generated": len(missing),
            "missing_from_generated_fields": missing,
            "extra_in_generated": len(extra),
            "extra_in_generated_fields": extra,
            "missing_from_both": len(both_null),
            "missing_from_both_fields": both_null,
        },
        "action_items": action_items,
    }


def build_soap_gt_comparison_report(
    result: dict | None = None,
    *,
    ground_truth: dict | None = None,
    generated: dict | None = None,
    generated_at: str | None = None,
) -> dict[str, Any]:
    """Build the structured GT vs Generated SOAP comparison report."""
    data = _as_dict(result)
    gt_soap = _as_soap(ground_truth if ground_truth is not None else data.get("soap_ground_truth"))
    if generated is not None:
        gen_soap = _as_soap(generated)
    else:
        gen_soap = generated_soap(data)
    meta = _result_meta(data)
    stamp = generated_at or datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    seen: set[str] = set()
    buckets: dict[str, list[dict[str, Any]]] = {key: [] for key, *_ in REPORT_SECTIONS}
    buckets["extra"] = []

    for section_id, _label, prefix, fields in REPORT_SECTIONS:
        if section_id == "plan_medications":
            buckets[section_id].extend(_medication_rows(gt_soap, gen_soap, seen))
            continue
        soap_gt = gt_soap
        soap_gen = gen_soap
        path_prefix = prefix
        if section_id == "assessment":
            wrapped_gt = dict(gt_soap)
            wrapped_gen = dict(gen_soap)
            wrapped_gt["assessment"] = _assessment_root(gt_soap)
            wrapped_gen["assessment"] = _assessment_root(gen_soap)
            soap_gt, soap_gen = wrapped_gt, wrapped_gen
        for field in fields:
            path = _join_path(path_prefix, field) if path_prefix else field
            buckets[section_id].append(
                _compare_path(field, path, soap_gt, soap_gen, in_schema=True)
            )
            seen.add(path)
            if section_id == "assessment":
                for original in (gt_soap, gen_soap):
                    raw = original.get("assessment")
                    if isinstance(raw, list):
                        for idx in range(len(raw)):
                            seen.add(f"assessment[{idx}].{field}")

    for extra in _extra_rows(gt_soap, gen_soap, seen):
        buckets[_section_for_path(extra["path"])].append(extra)

    sections = []
    all_rows: list[dict[str, Any]] = []
    extra_section_added = False
    for section_id, label, _prefix, _fields in REPORT_SECTIONS:
        rows = buckets.get(section_id) or []
        if not rows:
            continue
        sections.append({"id": section_id, "label": label, "rows": rows})
        all_rows.extend(rows)
    if buckets["extra"]:
        extra_section_added = True
        sections.append(
            {"id": "extra", "label": "EXTRA FIELDS", "rows": buckets["extra"]}
        )
        all_rows.extend(buckets["extra"])
    summary = _summary_block(all_rows)
    header_status = (
        STATUS_PASS if summary["overall_status"] == STATUS_PASS else STATUS_FAIL
    )
    if summary["overall_status"] == STATUS_PARTIAL:
        header_status = STATUS_PARTIAL
    return {
        "title": "GROUND TRUTH vs GENERATED SOAP COMPARISON REPORT",
        "test_case_id": meta["test_case_id"],
        "test_name": meta["test_name"],
        "test_id": meta["test_id"],
        "status": header_status,
        "threshold_status": summary["threshold_status"],
        "threshold_mark": summary["threshold_mark"],
        "generated_at": stamp,
        "compliance_percent": summary["compliance_percent"],
        "metrics": {
            "schema_accuracy": summary["schema_accuracy"],
            "ground_truth_match": summary["ground_truth_match"],
            "type_error_rate": summary["type_error_rate"],
            "missing_field_rate": summary["missing_field_rate"],
            "extra_field_rate": summary["extra_field_rate"],
            "overall_compliance": summary["compliance_percent"],
        },
        "columns": list(TABLE_COLUMNS),
        "sections": sections,
        "summary": summary,
        "error_analysis": build_error_analysis(all_rows),
        "metadata": build_case_metadata(data, gen_soap),
        "generated_soap": gen_soap,
        "has_extra_section": extra_section_added,
    }


def _escape_html(text: Any) -> str:
    return (
        str(text or "")
        .replace("&", "&amp;")
        .replace("<", "&lt;")
        .replace(">", "&gt;")
        .replace('"', "&quot;")
    )


def render_text_report(report: dict[str, Any]) -> str:
    """ASCII layout matching the comparison-report spec example."""
    bar = "═" * 115
    thin = "─" * 115
    summary = report["summary"]
    lines = [
        bar,
        report["title"],
        (
            f"Test Case: {report['test_case_id']} | {report['test_name']}"
        ),
        (
            f"Generated: {report['generated_at']} | Overall Compliance: "
            f"{report['compliance_percent']}%"
        ),
        f"Status: {report['status']}",
        bar,
        "",
    ]
    header = (
        f"{'FIELD NAME':<24}| {'GROUND TRUTH':<24}| {'GENERATED':<21}| "
        f"{'IN GT':<6}| {'ERROR':<6}| {'ERROR DESCRIPTION':<36}| STATUS"
    )
    for section in report["sections"]:
        lines.append(f"{section['label']}:")
        lines.append(thin)
        lines.append(header)
        lines.append(thin)
        for row in section["rows"]:
            lines.append(
                f"{str(row['field'])[:24]:<24}| {str(row['ground_truth'])[:24]:<24}| "
                f"{str(row['generated'])[:21]:<21}| {row['in_gt']:<6}| "
                f"{row['error_found']:<6}| {str(row['error_description'])[:36]:<36}| "
                f"{row['status_display']}"
            )
        lines.append("")
    br = summary["error_breakdown"]
    lines.extend(
        [
            bar,
            "COMPARISON SUMMARY:",
            bar,
            "",
            f"Total Fields: {summary['total_fields']}",
            (
                f"Fields PASSED (Match or Correctly Null): {summary['passed']} "
                f"({summary['passed_percent']}%)"
            ),
            (
                f"Fields FAILED (Mismatch or Missing): {summary['failed']} "
                f"({summary['failed_percent']}%)"
            ),
            "",
            "Error Breakdown:",
            f"  - Exact Matches: {br['exact_matches']}",
            f"  - Type Mismatches: {br['type_mismatches']}"
            + (
                f" ({', '.join(br['type_mismatch_fields'])})"
                if br["type_mismatch_fields"]
                else ""
            ),
            f"  - Format Mismatches: {br['format_mismatches']}"
            + (
                f" ({', '.join(br['format_mismatch_fields'])})"
                if br["format_mismatch_fields"]
                else ""
            ),
            f"  - Value Mismatches: {br['value_mismatches']}",
            f"  - Missing from Generated: {br['missing_from_generated']}",
            f"  - Extra in Generated: {br['extra_in_generated']}",
            f"  - Missing from Both (Correct): {br['missing_from_both']}",
            "",
            f"OVERALL STATUS: {summary['overall_status']} ({summary['compliance_percent']}% compliance)",
            "",
        ]
    )
    if summary["action_items"]:
        lines.append("Action Items:")
        for item in summary["action_items"]:
            lines.append(f"  ✗ {item}")
        lines.append("")
    lines.append(bar)
    return "\n".join(lines) + "\n"


def render_json_report(report: dict[str, Any]) -> bytes:
    return json.dumps(report, ensure_ascii=False, indent=2, default=str).encode("utf-8")


def _flat_csv_rows(report: dict[str, Any]) -> list[list[str]]:
    rows = [
        [
            "Section",
            *TABLE_COLUMNS,
            "Category",
        ]
    ]
    for section in report["sections"]:
        for row in section["rows"]:
            rows.append(
                [
                    section["label"],
                    row["field"],
                    row["ground_truth_full"],
                    row["generated_full"],
                    row["in_gt"],
                    row["error_found"],
                    row["error_description"],
                    row["status"],
                    row["category"],
                ]
            )
    summary = report["summary"]
    rows.append([])
    rows.append(["COMPARISON SUMMARY"])
    rows.append(["Total Fields", str(summary["total_fields"])])
    rows.append(["Passed", str(summary["passed"]), f"{summary['passed_percent']}%"])
    rows.append(["Failed", str(summary["failed"]), f"{summary['failed_percent']}%"])
    rows.append(["Compliance", f"{summary['compliance_percent']}%"])
    rows.append(["Overall Status", summary["overall_status"]])
    for item in summary["action_items"]:
        rows.append(["Action Item", item])
    return rows


def render_csv_report(report: dict[str, Any]) -> bytes:
    buffer = io.StringIO()
    writer = csv.writer(buffer)
    for row in _flat_csv_rows(report):
        writer.writerow(row)
    return buffer.getvalue().encode("utf-8-sig")


def render_excel_report(report: dict[str, Any]) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "SOAP Comparison"
    header_fill = PatternFill("solid", fgColor="1E293B")
    header_font = Font(bold=True, color="FFFFFF", size=10)
    pass_fill = PatternFill("solid", fgColor="D1FAE5")
    fail_fill = PatternFill("solid", fgColor="FECACA")
    pass_font = Font(color="166534")
    fail_font = Font(color="991B1B")
    title_font = Font(bold=True, size=14, color="0F172A")
    thin = Border(
        left=Side(style="thin", color="CBD5E1"),
        right=Side(style="thin", color="CBD5E1"),
        top=Side(style="thin", color="CBD5E1"),
        bottom=Side(style="thin", color="CBD5E1"),
    )
    wrap = Alignment(wrap_text=True, vertical="top")

    ws.merge_cells("A1:H1")
    ws["A1"] = report["title"]
    ws["A1"].font = title_font
    ws.merge_cells("A2:H2")
    ws["A2"] = (
        f"Test Case ID: {report['test_case_id']} | Test Name: {report['test_name']} | "
        f"Status: {report['status']}"
    )
    ws.merge_cells("A3:H3")
    ws["A3"] = (
        f"Generated: {report['generated_at']} | Compliance: {report['compliance_percent']}%"
    )

    current = 5
    col_headers = ["Section", *TABLE_COLUMNS]
    for section in report["sections"]:
        ws.merge_cells(start_row=current, start_column=1, end_row=current, end_column=8)
        cell = ws.cell(row=current, column=1, value=section["label"])
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="334155")
        current += 1
        for col, header in enumerate(col_headers, start=1):
            cell = ws.cell(row=current, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.border = thin
        current += 1
        for row in section["rows"]:
            values = [
                section["label"],
                row["field"],
                row["ground_truth_full"],
                row["generated_full"],
                row["in_gt"],
                row["error_found"],
                row["error_description"],
                row["status_display"],
            ]
            fill = pass_fill if row["passed"] else fail_fill
            font = pass_font if row["passed"] else fail_font
            for col, value in enumerate(values, start=1):
                cell = ws.cell(row=current, column=col, value=value)
                cell.fill = fill
                cell.font = font
                cell.alignment = wrap
                cell.border = thin
            current += 1
        current += 1

    summary = report["summary"]
    ws.cell(row=current, column=1, value="COMPARISON SUMMARY").font = Font(bold=True, size=12)
    current += 1
    summary_rows = [
        ("Total Fields", summary["total_fields"]),
        ("Passed", f"{summary['passed']} ({summary['passed_percent']}%)"),
        ("Failed", f"{summary['failed']} ({summary['failed_percent']}%)"),
        ("Compliance", f"{summary['compliance_percent']}%"),
        ("Overall Status", summary["overall_status"]),
        ("Exact Matches", summary["error_breakdown"]["exact_matches"]),
        ("Type Mismatches", summary["error_breakdown"]["type_mismatches"]),
        ("Format Mismatches", summary["error_breakdown"]["format_mismatches"]),
        ("Value Mismatches", summary["error_breakdown"]["value_mismatches"]),
        ("Missing from Generated", summary["error_breakdown"]["missing_from_generated"]),
        ("Extra in Generated", summary["error_breakdown"]["extra_in_generated"]),
        ("Missing from Both (Correct)", summary["error_breakdown"]["missing_from_both"]),
    ]
    for label, value in summary_rows:
        ws.cell(row=current, column=1, value=label).font = Font(bold=True)
        ws.cell(row=current, column=2, value=value)
        current += 1
    current += 1
    if summary["action_items"]:
        ws.cell(row=current, column=1, value="Action Items").font = Font(bold=True)
        current += 1
        for item in summary["action_items"]:
            ws.cell(row=current, column=1, value="✗")
            ws.cell(row=current, column=2, value=item)
            current += 1

    widths = (28, 28, 32, 32, 18, 14, 48, 14)
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width
    ws.freeze_panes = "A5"
    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def _html_table(section: dict[str, Any]) -> str:
    body = []
    for row in section["rows"]:
        css = "pass" if row["passed"] else "fail"
        body.append(
            "<tr class='{css}'>"
            "<td>{field}</td><td>{gt}</td><td>{gen}</td><td>{in_gt}</td>"
            "<td>{err}</td><td>{desc}</td><td>{status}</td>"
            "</tr>".format(
                css=css,
                field=_escape_html(row["field"]),
                gt=_escape_html(row["ground_truth_full"]),
                gen=_escape_html(row["generated_full"]),
                in_gt=_escape_html(row["in_gt"]),
                err=_escape_html(row["error_found"]),
                desc=_escape_html(row["error_description"]),
                status=_escape_html(row["status_display"]),
            )
        )
    return (
        f"<h2>{_escape_html(section['label'])}</h2>"
        "<div class='table-wrap'><table>"
        "<thead><tr>"
        "<th>Field Name</th><th>Ground Truth Value</th><th>Generated Value</th>"
        "<th>In GT</th><th>Error Found</th><th>Error Description</th><th>Status</th>"
        "</tr></thead>"
        f"<tbody>{''.join(body)}</tbody></table></div>"
    )


def render_html_report(report: dict[str, Any]) -> bytes:
    summary = report["summary"]
    br = summary["error_breakdown"]
    sections = "".join(_html_table(section) for section in report["sections"])
    actions = "".join(f"<li>✗ {_escape_html(item)}</li>" for item in summary["action_items"])
    overall_css = {
        STATUS_PASS: "pass",
        STATUS_FAIL: "fail",
        STATUS_PARTIAL: "partial",
    }.get(summary["overall_status"], "")
    html = f"""<!DOCTYPE html>
<html lang="en">
<head>
  <meta charset="utf-8"/>
  <title>{_escape_html(report['title'])}</title>
  <style>
    body {{ font-family: Segoe UI, system-ui, sans-serif; margin: 24px; color: #0f172a; }}
    h1 {{ font-size: 20px; margin-bottom: 8px; }}
    .meta {{ color: #475569; margin-bottom: 20px; }}
    h2 {{ font-size: 14px; margin: 24px 0 8px; letter-spacing: .04em; }}
    .table-wrap {{ overflow-x: auto; }}
    table {{ border-collapse: collapse; width: 100%; font-size: 13px; }}
    th, td {{ border: 1px solid #e2e8f0; padding: 8px 10px; vertical-align: top; text-align: left; }}
    th {{ background: #1e293b; color: #fff; font-size: 11px; text-transform: uppercase; }}
    tr.pass td {{ background: #f0fdf4; }}
    tr.fail td {{ background: #fef2f2; }}
    .summary {{ background: #f8fafc; border: 1px solid #e2e8f0; padding: 16px; border-radius: 8px; }}
    .badge {{ display: inline-block; padding: 4px 10px; border-radius: 999px; font-weight: 600; }}
    .badge.pass {{ background: #d1fae5; color: #166534; }}
    .badge.fail {{ background: #fecaca; color: #991b1b; }}
    .badge.partial {{ background: #fef3c7; color: #92400e; }}
  </style>
</head>
<body>
  <h1>{_escape_html(report['title'])}</h1>
  <p class="meta">
    Test Case ID: {_escape_html(report['test_case_id'])} |
    Test Name: {_escape_html(report['test_name'])} |
    Status: {_escape_html(report['status'])}<br/>
    Generated: {_escape_html(report['generated_at'])} |
    Compliance: {report['compliance_percent']}%
  </p>
  {sections}
  <div class="summary">
    <h2>Comparison Summary</h2>
    <p><span class="badge {overall_css}">{_escape_html(summary['overall_status'])}</span>
       {summary['compliance_percent']}% compliance</p>
    <p>Total Fields: {summary['total_fields']}<br/>
       Fields PASSED: {summary['passed']} ({summary['passed_percent']}%)<br/>
       Fields FAILED: {summary['failed']} ({summary['failed_percent']}%)</p>
    <p>Error Breakdown:</p>
    <ul>
      <li>Exact Matches: {br['exact_matches']}</li>
      <li>Type Mismatches: {br['type_mismatches']}</li>
      <li>Format Mismatches: {br['format_mismatches']}</li>
      <li>Value Mismatches: {br['value_mismatches']}</li>
      <li>Missing from Generated: {br['missing_from_generated']}</li>
      <li>Extra in Generated: {br['extra_in_generated']}</li>
      <li>Missing from Both (Correct): {br['missing_from_both']}</li>
    </ul>
    {"<p>Action Items:</p><ul>" + actions + "</ul>" if actions else ""}
  </div>
</body>
</html>
"""
    return html.encode("utf-8")


def _escape_pdf(text: Any) -> str:
    return (
        str(text or "")
        .replace("&", "&amp;")
        .replace("<", "&lt;")
        .replace(">", "&gt;")
        .replace("\n", "<br/>")
    )


def render_pdf_report(report: dict[str, Any]) -> bytes:
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=landscape(A4),
        leftMargin=1.2 * cm,
        rightMargin=1.2 * cm,
        topMargin=1.2 * cm,
        bottomMargin=1.2 * cm,
    )
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        "SoapGtTitle",
        parent=styles["Heading1"],
        fontSize=14,
        textColor=colors.HexColor("#0f172a"),
        spaceAfter=6,
    )
    meta_style = ParagraphStyle(
        "SoapGtMeta",
        parent=styles["Normal"],
        fontSize=9,
        textColor=colors.HexColor("#334155"),
        spaceAfter=8,
    )
    section_style = ParagraphStyle(
        "SoapGtSection",
        parent=styles["Heading2"],
        fontSize=10,
        textColor=colors.HexColor("#1e293b"),
        spaceBefore=10,
        spaceAfter=4,
    )
    cell_style = ParagraphStyle(
        "SoapGtCell",
        parent=styles["Normal"],
        fontSize=7,
        leading=9,
    )
    header_style = ParagraphStyle(
        "SoapGtHead",
        parent=cell_style,
        fontName="Helvetica-Bold",
        textColor=colors.white,
        fontSize=7,
    )
    story = [
        Paragraph(_escape_pdf(report["title"]), title_style),
        Paragraph(
            _escape_pdf(
                f"Test Case ID: {report['test_case_id']} | Test Name: {report['test_name']} | "
                f"Status: {report['status']}"
            ),
            meta_style,
        ),
        Paragraph(
            _escape_pdf(
                f"Generated: {report['generated_at']} | Compliance: {report['compliance_percent']}%"
            ),
            meta_style,
        ),
    ]
    col_widths = [3.4 * cm, 4.2 * cm, 4.2 * cm, 1.8 * cm, 1.6 * cm, 6.4 * cm, 2.2 * cm]
    header = [Paragraph(_escape_pdf(col), header_style) for col in TABLE_COLUMNS]
    for section in report["sections"]:
        story.append(Paragraph(_escape_pdf(section["label"]), section_style))
        data = [header]
        row_colors = []
        for row in section["rows"]:
            data.append(
                [
                    Paragraph(_escape_pdf(row["field"]), cell_style),
                    Paragraph(_escape_pdf(row["ground_truth_full"]), cell_style),
                    Paragraph(_escape_pdf(row["generated_full"]), cell_style),
                    Paragraph(_escape_pdf(row["in_gt"]), cell_style),
                    Paragraph(_escape_pdf(row["error_found"]), cell_style),
                    Paragraph(_escape_pdf(row["error_description"]), cell_style),
                    Paragraph(_escape_pdf(row["status_display"]), cell_style),
                ]
            )
            row_colors.append(
                colors.HexColor("#d1fae5") if row["passed"] else colors.HexColor("#fecaca")
            )
        table = Table(data, colWidths=col_widths, repeatRows=1)
        style_cmds = [
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1e293b")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#cbd5e1")),
            ("LEFTPADDING", (0, 0), (-1, -1), 4),
            ("RIGHTPADDING", (0, 0), (-1, -1), 4),
            ("TOPPADDING", (0, 0), (-1, -1), 3),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
        ]
        for idx, fill in enumerate(row_colors, start=1):
            style_cmds.append(("BACKGROUND", (0, idx), (-1, idx), fill))
        table.setStyle(TableStyle(style_cmds))
        story.append(table)

    summary = report["summary"]
    br = summary["error_breakdown"]
    story.append(Spacer(1, 0.4 * cm))
    story.append(Paragraph("COMPARISON SUMMARY", section_style))
    story.append(
        Paragraph(
            _escape_pdf(
                f"Total Fields: {summary['total_fields']} | "
                f"PASSED: {summary['passed']} ({summary['passed_percent']}%) | "
                f"FAILED: {summary['failed']} ({summary['failed_percent']}%) | "
                f"OVERALL: {summary['overall_status']} ({summary['compliance_percent']}%)"
            ),
            meta_style,
        )
    )
    story.append(
        Paragraph(
            _escape_pdf(
                "Error Breakdown: "
                f"Exact Matches {br['exact_matches']}; "
                f"Type Mismatches {br['type_mismatches']}; "
                f"Format Mismatches {br['format_mismatches']}; "
                f"Value Mismatches {br['value_mismatches']}; "
                f"Missing from Generated {br['missing_from_generated']}; "
                f"Extra in Generated {br['extra_in_generated']}; "
                f"Missing from Both {br['missing_from_both']}"
            ),
            meta_style,
        )
    )
    if summary["action_items"]:
        items = "<br/>".join(f"✗ {_escape_pdf(item)}" for item in summary["action_items"])
        story.append(Paragraph(items, cell_style))
    doc.build(story)
    return buffer.getvalue()


_RENDERERS = {
    "json": (render_json_report, "application/json", "json"),
    "csv": (render_csv_report, "text/csv", "csv"),
    "excel": (
        render_excel_report,
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        "xlsx",
    ),
    "html": (render_html_report, "text/html", "html"),
    "pdf": (render_pdf_report, "application/pdf", "pdf"),
}


def render_soap_gt_comparison_report(report: dict[str, Any], fmt: str) -> tuple[bytes, str, str]:
    """Return (payload, mimetype, extension) for a supported download format."""
    key = (fmt or "json").lower()
    if key in ("xlsx", "xls"):
        key = "excel"
    if key not in _RENDERERS:
        raise ValueError(f"format must be one of {', '.join(SUPPORTED_FORMATS)}")
    renderer, mime, ext = _RENDERERS[key]
    return renderer(report), mime, ext
