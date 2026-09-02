"""PDF and Excel report generation."""

from __future__ import annotations

import io
import json
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import cm
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

from medsum_testing.backend.models.test_result import TestResult
from medsum_testing.backend.services.batch_report import (
    BATCH_REPORT_SECTIONS,
    build_batch_report,
)
from medsum_testing.backend.services.individual_report import (
    extra_report_fields,
    individual_report_fields,
)


# One PDF table row cannot be taller than the page frame (~686pt on A4
# with these margins). 8pt type at 11pt leading wraps ~80 chars/line, so
# ~1800 characters stays well under one frame. Longer values are split
# across continuation rows instead of overflowing (LayoutError 500).
PDF_CELL_CHAR_LIMIT = 1800
PDF_TABLE_ROWS_PER_BLOCK = 12


def _fmt(value: Any) -> str:
    if value is None:
        return "N/A"
    if isinstance(value, (dict, list)):
        return json.dumps(value, ensure_ascii=False, indent=2)
    return str(value)


def _escape_pdf(text: str) -> str:
    return (
        str(text or "")
        .replace("&", "&amp;")
        .replace("<", "&lt;")
        .replace(">", "&gt;")
        .replace("\n", "<br/>")
    )


def _chunks_for_pdf(text: str, limit: int = PDF_CELL_CHAR_LIMIT) -> list[str]:
    raw = str(text or "")
    if len(raw) <= limit:
        return [raw]
    parts: list[str] = []
    remaining = raw
    while remaining:
        if len(remaining) <= limit:
            parts.append(remaining)
            break
        window = remaining[:limit]
        cut = window.rfind("\n")
        if cut < limit // 4:
            cut = window.rfind(" ")
        if cut < limit // 4:
            cut = limit
        else:
            cut += 1
        parts.append(remaining[:cut])
        remaining = remaining[cut:]
    return parts or [""]


def _section_rows(result: TestResult) -> list[tuple[str, str]]:
    required = individual_report_fields(result)
    tc = result.transcription_comparison
    sc = result.summary_comparison
    mc = result.medication_comparison
    rc = result.regression_comparison

    rows = required + extra_report_fields(result) + [
        ("Test Case ID", result.test_case_id or "N/A"),
        ("Test ID", result.test_id),
        ("Patient ID", result.patient_id or "N/A"),
        ("Doctor ID", result.doctor_id or "N/A"),
        ("Session Date/Time", result.session_datetime or result.timestamp),
        ("Timestamp", result.timestamp),
        ("Language", result.language),
        ("Audio File", result.audio_filename),
        ("Audio Duration (s)", str(result.audio_duration_seconds)),
        ("AI Model", result.ai_model),
        ("Final Result", result.final_result),
        ("Accuracy Score", str(result.accuracy_score) if result.accuracy_score is not None else "N/A"),
        ("Accuracy Skipped", "Yes" if result.accuracy_skipped else "No"),
        ("Accuracy Skip Reason", result.accuracy_skip_reason or "N/A"),
        ("Retry Count", str(result.retry_count)),
        ("Errors", "; ".join(result.errors) if result.errors else "None"),
        ("Ground Truth Transcription", result.ground_truth_transcription or "N/A"),
        ("Generated Transcription", result.generated_transcription or "N/A"),
        ("Previous Transcription", result.previous_transcription or "N/A"),
        ("Generated Summary", _fmt(result.generated_summary)),
        ("Previous Summary", _fmt(result.previous_summary)),
        ("Text Translation", result.text_translation or "N/A"),
        ("Medications Before", _fmt(result.medications_before)),
        ("Medications After Normalization", _fmt(result.medications_after_normalization)),
        ("Medications Generated", _fmt(result.medications_generated)),
    ]

    if tc:
        rows.extend(
            [
                ("Transcription Similarity", str(tc.similarity_score or "N/A")),
                ("Transcription Severity", tc.severity),
                ("Medical Terminology Differences", "\n".join(tc.medical_differences) or "None"),
                ("General Transcription Differences", "\n".join(tc.general_differences) or "None"),
                ("Transcription Comparison Summary", tc.summary or "N/A"),
            ]
        )

    if sc:
        rows.extend(
            [
                ("Summary Similarity", str(sc.similarity_score or "N/A")),
                ("Summary Severity", sc.severity),
                ("Summary Differences", "\n".join(sc.medical_differences + sc.general_differences) or "None"),
                ("Summary Comparison Summary", sc.summary or "N/A"),
            ]
        )

    if mc:
        rows.extend(
            [
                ("Medication Similarity", str(mc.similarity_score or "N/A")),
                ("Medication Severity", mc.severity),
                ("Medications Added", "\n".join(mc.added) or "None"),
                ("Medications Removed", "\n".join(mc.removed) or "None"),
                ("Medications Changed", "\n".join(mc.changed) or "None"),
                ("Medication Differences", "\n".join(mc.medical_differences) or "None"),
                ("Medication Comparison Summary", mc.summary or "N/A"),
            ]
        )

    if rc and not rc.skipped:
        rows.extend(
            [
                ("Regression Similarity", str(rc.similarity_score or "N/A")),
                ("Regression Severity", rc.severity),
                ("Regression Differences", "\n".join(rc.medical_differences + rc.general_differences) or "None"),
                ("Regression Summary", rc.summary or "N/A"),
            ]
        )

    return rows


def generate_pdf(test_result: TestResult) -> bytes:
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, rightMargin=1.5 * cm, leftMargin=1.5 * cm)
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        "Title",
        parent=styles["Heading1"],
        fontSize=16,
        spaceAfter=12,
        textColor=colors.HexColor("#2563eb"),
    )
    label_style = ParagraphStyle(
        "Label",
        parent=styles["Normal"],
        fontName="Helvetica-Bold",
        fontSize=9,
        textColor=colors.HexColor("#374151"),
    )
    value_style = ParagraphStyle(
        "Value",
        parent=styles["Normal"],
        fontName="Helvetica",
        fontSize=8,
        leading=11,
    )

    story = [
        Paragraph("MEDSUM Accuracy Test Report", title_style),
        Spacer(1, 0.3 * cm),
    ]

    table_data = []
    for label, value in _section_rows(test_result):
        chunks = _chunks_for_pdf(value)
        for i, chunk in enumerate(chunks):
            table_data.append(
                [
                    Paragraph(label if i == 0 else "", label_style),
                    Paragraph(_escape_pdf(chunk), value_style),
                ]
            )

    kv_style = TableStyle(
        [
            ("BACKGROUND", (0, 0), (0, -1), colors.HexColor("#f0f1f5")),
            ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#d9dce6")),
            ("LEFTPADDING", (0, 0), (-1, -1), 6),
            ("RIGHTPADDING", (0, 0), (-1, -1), 6),
            ("TOPPADDING", (0, 0), (-1, -1), 4),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
        ]
    )
    for start in range(0, len(table_data), PDF_TABLE_ROWS_PER_BLOCK):
        block = table_data[start : start + PDF_TABLE_ROWS_PER_BLOCK]
        table = Table(block, colWidths=[5.5 * cm, 12 * cm], splitByRow=1)
        table.setStyle(kv_style)
        story.append(table)
        story.append(Spacer(1, 0.05 * cm))
    doc.build(story)
    return buffer.getvalue()


def generate_excel(test_result: TestResult) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Test Result"

    header_fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")

    ws.cell(row=1, column=1, value="Field").fill = header_fill
    ws.cell(row=1, column=1).font = header_font
    ws.cell(row=1, column=2, value="Value").fill = header_fill
    ws.cell(row=1, column=2).font = header_font

    for idx, (label, value) in enumerate(_section_rows(test_result), start=2):
        ws.cell(row=idx, column=1, value=label)
        cell = ws.cell(row=idx, column=2, value=value)
        cell.alignment = Alignment(wrap_text=True, vertical="top")

    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 80

    # Side-by-side transcription sheet
    ws2 = wb.create_sheet("Transcription Diff")
    ws2.cell(row=1, column=1, value="Ground Truth").fill = header_fill
    ws2.cell(row=1, column=1).font = header_font
    ws2.cell(row=1, column=2, value="Generated").fill = header_fill
    ws2.cell(row=1, column=2).font = header_font
    ws2.cell(row=2, column=1, value=test_result.ground_truth_transcription or "N/A")
    ws2.cell(row=2, column=2, value=test_result.generated_transcription or "N/A")
    ws2.column_dimensions["A"].width = 60
    ws2.column_dimensions["B"].width = 60

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def save_report_path(
    test_id: str,
    fmt: str,
    path: str,
    token: str,
    config: dict,
) -> None:
    """Update the run record in Django with the generated report path."""
    from datetime import datetime, timezone

    from medsum_testing.backend.services import medsum_api

    field = "report_pdf_path" if fmt == "pdf" else "report_excel_path"
    medsum_api.save_test_run(
        {
            "test_id": test_id,
            field: path,
            "report_generated_at": datetime.now(timezone.utc).isoformat(),
        },
        token,
        config,
    )


def _kv_rows(mapping: dict) -> list[tuple[str, str]]:
    rows = []
    for key, value in mapping.items():
        if key in {"status_counts", "execution_counts", "evaluation_counts"} and isinstance(value, dict):
            rows.append((key, ", ".join(f"{k}={v}" for k, v in value.items()) or "—"))
        elif key == "stage_averages" and isinstance(value, dict):
            rows.append((key, ", ".join(f"{k}={v}" for k, v in value.items())))
        elif isinstance(value, list):
            continue
        else:
            rows.append((key, _fmt(value)))
    return rows


def generate_batch_pdf(rows: list) -> bytes:
    report = build_batch_report(rows)
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, rightMargin=1.5 * cm, leftMargin=1.5 * cm)
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        "BatchTitle",
        parent=styles["Heading1"],
        fontSize=16,
        spaceAfter=12,
        textColor=colors.HexColor("#2563eb"),
    )
    heading = ParagraphStyle(
        "BatchH",
        parent=styles["Heading2"],
        fontSize=12,
        spaceBefore=10,
        spaceAfter=6,
    )
    label_style = ParagraphStyle(
        "BatchLabel",
        parent=styles["Normal"],
        fontName="Helvetica-Bold",
        fontSize=8,
    )
    value_style = ParagraphStyle(
        "BatchValue",
        parent=styles["Normal"],
        fontName="Helvetica",
        fontSize=8,
        leading=11,
    )

    def _cell(text: str, style) -> Paragraph:
        return Paragraph(_escape_pdf(str(text)[: PDF_CELL_CHAR_LIMIT * 2]), style)

    def _kv_table(pairs: list[tuple[str, str]]):
        rows = []
        for key, value in pairs:
            chunks = _chunks_for_pdf(value)
            for i, chunk in enumerate(chunks):
                rows.append([
                    _cell(key if i == 0 else "", label_style),
                    _cell(chunk, value_style),
                ])
        return rows

    story = [Paragraph(report["title"], title_style)]
    for section in BATCH_REPORT_SECTIONS:
        story.append(Paragraph(section, heading))
        body = report.get(section)
        if section == "Test Case Details":
            table_data = [[
                _cell("Test Case ID", label_style),
                _cell("Execution Status", label_style),
                _cell("SOAP Evaluation", label_style),
                _cell("Accuracy", label_style),
                _cell("Latency", label_style),
                _cell("Individual Report", label_style),
            ]]
            for case in body or []:
                table_data.append([
                    _cell(case.get("test_case_id", ""), value_style),
                    _cell(case.get("execution_status", ""), value_style),
                    _cell(case.get("soap_evaluation", ""), value_style),
                    _cell(case.get("accuracy", ""), value_style),
                    _cell(case.get("latency", ""), value_style),
                    _cell(case.get("individual_report", ""), value_style),
                ])
        elif isinstance(body, dict):
            pairs = list(_kv_rows(body))
            if body.get("per_case"):
                pairs.append((
                    "per_case",
                    "\n".join(
                        f"{c.get('test_case_id') or c.get('audio_file')}: "
                        f"{c.get('accuracy') or c.get('total_time')}"
                        for c in body["per_case"]
                    ),
                ))
            table_data = _kv_table(pairs)
        else:
            table_data = _kv_table([("", _fmt(body))])
        table = Table(
            table_data,
            colWidths=[4.5 * cm, 13 * cm] if section != "Test Case Details" else None,
            splitByRow=1,
        )
        table.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (0, -1), colors.HexColor("#f0f1f5")),
                    ("VALIGN", (0, 0), (-1, -1), "TOP"),
                    ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#d9dce6")),
                    ("LEFTPADDING", (0, 0), (-1, -1), 4),
                    ("RIGHTPADDING", (0, 0), (-1, -1), 4),
                    ("TOPPADDING", (0, 0), (-1, -1), 3),
                    ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
                ]
            )
        )
        story.append(table)
        story.append(Spacer(1, 0.2 * cm))
    doc.build(story)
    return buffer.getvalue()


def generate_batch_excel(rows: list) -> bytes:
    report = build_batch_report(rows)
    wb = Workbook()
    header_fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")

    first = True
    for section in BATCH_REPORT_SECTIONS:
        ws = wb.active if first else wb.create_sheet(section[:31])
        if first:
            ws.title = section[:31]
            first = False
        body = report.get(section)
        if section == "Test Case Details":
            headers = [
                "Test Case ID",
                "Execution Status",
                "SOAP Evaluation",
                "Accuracy",
                "Latency",
                "Audio File",
                "Individual Report",
            ]
            for col, title in enumerate(headers, start=1):
                cell = ws.cell(row=1, column=col, value=title)
                cell.fill = header_fill
                cell.font = header_font
            for idx, case in enumerate(body or [], start=2):
                ws.cell(row=idx, column=1, value=case.get("test_case_id"))
                ws.cell(row=idx, column=2, value=case.get("execution_status"))
                ws.cell(row=idx, column=3, value=case.get("soap_evaluation"))
                ws.cell(row=idx, column=4, value=case.get("accuracy"))
                ws.cell(row=idx, column=5, value=case.get("latency"))
                ws.cell(row=idx, column=6, value=case.get("audio_file"))
                ws.cell(row=idx, column=7, value=case.get("individual_report"))
        else:
            ws.cell(row=1, column=1, value="Field").fill = header_fill
            ws.cell(row=1, column=1).font = header_font
            ws.cell(row=1, column=2, value="Value").fill = header_fill
            ws.cell(row=1, column=2).font = header_font
            row_i = 2
            for key, value in _kv_rows(body or {}):
                ws.cell(row=row_i, column=1, value=key)
                ws.cell(row=row_i, column=2, value=value)
                row_i += 1
            if isinstance(body, dict) and body.get("per_case"):
                ws.cell(row=row_i, column=1, value="per_case")
                ws.cell(row=row_i, column=1).font = Font(bold=True)
                row_i += 1
                keys = list(body["per_case"][0].keys()) if body["per_case"] else []
                for col, key in enumerate(keys, start=1):
                    cell = ws.cell(row=row_i, column=col, value=key)
                    cell.fill = header_fill
                    cell.font = header_font
                row_i += 1
                for case in body["per_case"]:
                    for col, key in enumerate(keys, start=1):
                        ws.cell(row=row_i, column=col, value=case.get(key))
                    row_i += 1
        ws.column_dimensions["A"].width = 28
        ws.column_dimensions["B"].width = 50

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()
