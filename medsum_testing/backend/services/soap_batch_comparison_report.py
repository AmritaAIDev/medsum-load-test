"""SOAP batch test report: summary, latency, and per-case field comparisons.

CSV is summary + latency tables only. Excel / HTML / PDF / JSON include
per-case comparison tables and the final batch summary.
"""

from __future__ import annotations

import csv
import io
import json
import re
from datetime import datetime
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import cm
from reportlab.platypus import (
    KeepTogether,
    PageBreak,
    Paragraph,
    SimpleDocTemplate,
    Table,
    TableStyle,
)

from medsum_testing.backend.services.individual_report import as_result_dict
from medsum_testing.backend.services.latency_analysis import pick_transcribe_time
from medsum_testing.backend.services.soap_gt_comparison_report import (
    CAT_EXTRA_IN_GENERATED,
    CAT_FORMAT_MISMATCH,
    CAT_MISSING_FROM_GENERATED,
    CAT_TYPE_MISMATCH,
    CAT_VALUE_MISMATCH,
    PARTIAL_THRESHOLD,
    PASS_THRESHOLD,
    STATUS_FAIL,
    STATUS_PARTIAL,
    STATUS_PASS,
    SUPPORTED_FORMATS,
    build_soap_gt_comparison_report,
    flatten_report_rows,
    human_field_name,
    pct_rate,
    threshold_status,
)

TITLE = "SOAP BATCH TEST REPORT - MedSum Testing"

SUMMARY_COLUMNS = (
    "TEST ID",
    "ACCURACY",
    "PROCESSING TIME",
    "API STATUS",
    "MODEL CONFIG",
)

LATENCY_COLUMNS = (
    "TEST ID",
    "TRANSCRIPTION LATENCY",
    "TRANSLATION LATENCY",
    "LLM LATENCY",
    "TOTAL LATENCY",
)

COMPARISON_COLUMNS = (
    "FIELD NAME",
    "GROUND TRUTH SOAP",
    "GENERATED SOAP",
    "ERROR FOUND",
    "ERROR DESCRIPTION",
    "DATA TYPE",
)

ERROR_COLUMNS = (
    "ERROR #",
    "FIELD NAME",
    "ERROR TYPE",
    "GROUND TRUTH VALUE",
    "GENERATED VALUE",
    "IMPACT",
    "RECOMMENDATION",
)

BATCH_SECTIONS = (
    ("subjective", "SUBJECTIVE SECTION"),
    ("objective_vitals", "OBJECTIVE SECTION → VITALS"),
    ("objective_physical_exam", "OBJECTIVE SECTION → PHYSICAL EXAM"),
    ("assessment", "ASSESSMENT SECTION"),
    ("plan", "PLAN SECTION"),
    ("summary", "SUMMARY"),
)

_PASS_FILL = PatternFill("solid", fgColor="D1FAE5")
_FAIL_FILL = PatternFill("solid", fgColor="FECACA")
_HEADER_FILL = PatternFill("solid", fgColor="1E293B")
_SECTION_FILL = PatternFill("solid", fgColor="334155")
_AVG_FILL = PatternFill("solid", fgColor="E2E8F0")


def _as_dict(value: Any) -> dict:
    return value if isinstance(value, dict) else {}


def _mean(values: list[float]) -> float:
    if not values:
        return 0.0
    return round(sum(values) / len(values), 1)


def _pct_label(value: float) -> str:
    return f"{value}%"


def _format_ms(seconds: Any) -> str:
    if seconds is None or seconds == "":
        return "unavailable"
    try:
        value = float(seconds)
    except (TypeError, ValueError):
        return "unavailable"
    return f"{int(round(value * 1000))} ms"


def _format_timestamp(value: Any) -> str:
    raw = str(value or "").strip()
    if not raw:
        return ""
    cleaned = raw.replace("T", " ")
    if cleaned.endswith("Z"):
        cleaned = cleaned[:-1].strip() + " UTC"
    elif "UTC" not in cleaned.upper() and "+" not in cleaned:
        cleaned = f"{cleaned} UTC"
    return cleaned


def _model_config(data: dict[str, Any], metadata: dict[str, Any] | None = None) -> str:
    meta = metadata or {}
    model = str(
        data.get("ai_model_used")
        or data.get("ai_model")
        or data.get("llm_model")
        or data.get("model_type")
        or meta.get("model_config")
        or ""
    ).strip()
    temp = data.get("llm_temperature")
    if temp in (None, ""):
        temp = data.get("model_temperature")
    if model and temp not in (None, ""):
        return f"{model} | Temp: {temp}"
    return model or "-"


def _api_endpoint(metadata: dict[str, Any], data: dict[str, Any]) -> str:
    endpoint = str(
        metadata.get("endpoint")
        or data.get("flask_transcribe_endpoint")
        or "/transcribe"
    ).strip()
    verb = str(metadata.get("api_method_verb") or "POST").strip() or "POST"
    if endpoint.upper().startswith("POST ") or endpoint.upper().startswith("GET "):
        return endpoint
    return f"{verb} {endpoint}"


def _schema_rows(report: dict[str, Any]) -> list[dict[str, Any]]:
    rows = []
    for row in flatten_report_rows(report):
        if row.get("in_schema") is False or row.get("category") == CAT_EXTRA_IN_GENERATED:
            continue
        rows.append(row)
    return rows


def _row_data_type(row: dict[str, Any]) -> str:
    existing = str(row.get("data_type") or "").strip()
    if existing:
        return existing
    if row.get("gen_kind"):
        return str(row["gen_kind"])
    if row.get("generated_raw") is None and (
        row.get("ground_truth_raw") is None or row.get("gt_kind") is None
    ):
        return "null"
    return str(row.get("gt_kind") or "null")


def _display_name(row: dict[str, Any]) -> str:
    return str(
        row.get("display_field")
        or human_field_name(str(row.get("field") or ""), str(row.get("path") or ""))
    )


def _error_type(row: dict[str, Any]) -> str:
    category = str(row.get("category") or row.get("error_type") or "").strip()
    return category.replace(" ", "_").upper() if category else ""


def _error_recommendation(row: dict[str, Any]) -> str:
    category = row.get("category") or row.get("error_type")
    field = str(row.get("path") or row.get("field") or row.get("display_field") or "").lower()
    existing = str(row.get("recommendation") or "").strip()
    if category == CAT_TYPE_MISMATCH:
        return "Ensure all values as strings"
    if category == CAT_FORMAT_MISMATCH:
        if "temperature" in field:
            return "Include temperature unit"
        return "Include unit descriptions"
    if category == CAT_MISSING_FROM_GENERATED:
        return "Validate all fields against schema before API response"
    if category == CAT_EXTRA_IN_GENERATED:
        return "Do not emit fields that are not in the SOAP schema or ground truth"
    if category == CAT_VALUE_MISMATCH:
        return "Align generated values with ground truth"
    return existing or "Align generated SOAP with ground truth"


def _latency_seconds(data: dict[str, Any]) -> dict[str, float | None]:
    out: dict[str, float | None] = {}
    for key in ("transcription", "translation", "soap", "total_time"):
        raw = pick_transcribe_time(data, key)
        if raw is None or raw == "":
            out[key] = None
            continue
        try:
            out[key] = float(raw)
        except (TypeError, ValueError):
            out[key] = None
    return out


def _batch_latency_row(test_id: str, seconds: dict[str, float | None]) -> dict[str, Any]:
    return {
        "test_id": test_id,
        "transcription_latency": _format_ms(seconds.get("transcription")),
        "translation_latency": _format_ms(seconds.get("translation")),
        "llm_latency": _format_ms(seconds.get("soap")),
        "total_latency": _format_ms(seconds.get("total_time")),
        "transcription_seconds": seconds.get("transcription"),
        "translation_seconds": seconds.get("translation"),
        "llm_seconds": seconds.get("soap"),
        "total_seconds": seconds.get("total_time"),
    }


def _error_analysis_rows(report: dict[str, Any]) -> list[dict[str, Any]]:
    errors: list[dict[str, Any]] = []
    seen: set[tuple[str, str]] = set()
    for row in flatten_report_rows(report):
        if row.get("passed"):
            continue
        err = {
            "field": _display_name(row),
            "error_type": _error_type(row),
            "category": row.get("category"),
            "ground_truth": row.get("ground_truth_full") or row.get("ground_truth"),
            "generated": row.get("generated_full") or row.get("generated"),
            "impact": (
                "HIGH"
                if row.get("category")
                in {CAT_TYPE_MISMATCH, CAT_MISSING_FROM_GENERATED, CAT_VALUE_MISMATCH}
                else "MEDIUM"
            ),
            "recommendation": _error_recommendation(row),
        }
        key = (str(err["field"]), str(err["error_type"]))
        if key in seen:
            continue
        seen.add(key)
        errors.append(err)
    return errors


def _case_sections(report: dict[str, Any]) -> list[dict[str, Any]]:
    by_id = {section.get("id"): section for section in report.get("sections") or []}
    plan_rows = []
    for key in ("plan_medications", "plan_other", "plan"):
        plan_rows.extend((by_id.get(key) or {}).get("rows") or [])
    sections = []
    for section_id, label in BATCH_SECTIONS:
        if section_id == "plan":
            rows = [
                row
                for row in plan_rows
                if row.get("in_schema") is not False
                and row.get("category") != CAT_EXTRA_IN_GENERATED
            ]
        else:
            rows = [
                row
                for row in (by_id.get(section_id) or {}).get("rows") or []
                if row.get("in_schema") is not False
                and row.get("category") != CAT_EXTRA_IN_GENERATED
            ]
        if not rows:
            continue
        sections.append({"id": section_id, "label": label, "rows": rows})
    return sections


def case_summary_row(report: dict[str, Any], result: dict[str, Any] | None = None) -> dict[str, Any]:
    data = _as_dict(result)
    summary = _as_dict(report.get("summary"))
    meta = _as_dict(report.get("metadata"))
    schema = _schema_rows(report)
    total = len(schema) or int(summary.get("total_fields") or 0)
    passed = sum(1 for row in schema if row.get("passed")) if schema else int(summary.get("passed") or 0)
    failed = total - passed
    accuracy = round(100.0 * passed / total, 1) if total else 0.0
    band, mark = threshold_status(accuracy)
    seconds = _latency_seconds(data)
    processing = _format_ms(seconds.get("total_time"))
    if processing == "unavailable":
        processing = str(meta.get("processing_time") or "unavailable")
    return {
        "test_id": report.get("test_case_id") or meta.get("test_id") or "N/A",
        "test_name": report.get("test_name") or meta.get("test_name") or "",
        "accuracy": accuracy,
        "accuracy_display": _pct_label(accuracy),
        "processing_time": processing,
        "api_status": str(meta.get("response_code") or "unavailable"),
        "model_config": _model_config(data, meta),
        "api_endpoint": _api_endpoint(meta, data),
        "request_id": str(meta.get("api_request_id") or ""),
        "timestamp": _format_timestamp(meta.get("timestamp") or data.get("timestamp")),
        "fields_validated": total,
        "fields_passed": passed,
        "fields_failed": failed,
        "passed_percent": pct_rate(passed, total),
        "failed_percent": pct_rate(failed, total),
        "threshold_status": band,
        "threshold_mark": mark,
        "total_seconds": seconds.get("total_time"),
    }


def build_soap_batch_comparison_report(
    rows: list[Any] | None,
    *,
    generated_at: str | None = None,
    batch_id: str = "",
) -> dict[str, Any]:
    """Build the SOAP batch test report payload from result dicts."""
    items = [as_result_dict(row) for row in (rows or [])]
    stamp = generated_at or datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    cases: list[dict[str, Any]] = []
    latency_rows: list[dict[str, Any]] = []
    for index, item in enumerate(items, start=1):
        report = build_soap_gt_comparison_report(item, generated_at=stamp)
        summary_row = case_summary_row(report, item)
        seconds = _latency_seconds(item)
        latency_rows.append(_batch_latency_row(summary_row["test_id"], seconds))
        cases.append(
            {
                "index": index,
                "heading": (
                    f"TEST CASE: {summary_row['test_id']} | {summary_row['test_name']}"
                ).rstrip(" |"),
                "summary_row": summary_row,
                "sections": _case_sections(report),
                "error_analysis": _error_analysis_rows(report),
                "report": report,
            }
        )

    summary_rows = [case["summary_row"] for case in cases]
    n = len(summary_rows)
    avg_accuracy = _mean([r["accuracy"] for r in summary_rows])
    bands = [r["threshold_status"] for r in summary_rows]
    passed_n = sum(1 for b in bands if b == STATUS_PASS)
    partial_n = sum(1 for b in bands if b == STATUS_PARTIAL)
    failed_n = sum(1 for b in bands if b == STATUS_FAIL)

    def _avg_ms(key: str) -> str:
        nums = [row[key] for row in latency_rows if row.get(key) is not None]
        if not nums:
            return "unavailable"
        return _format_ms(sum(nums) / len(nums))

    avg_processing_nums = [r["total_seconds"] for r in summary_rows if r.get("total_seconds") is not None]
    avg_processing = (
        _format_ms(sum(avg_processing_nums) / len(avg_processing_nums))
        if avg_processing_nums
        else "unavailable"
    )
    statuses = [r["api_status"] for r in summary_rows]
    avg_status = statuses[0] if statuses and len(set(statuses)) == 1 else "-"

    averages = {
        "test_id": "BATCH AVG",
        "accuracy": avg_accuracy,
        "accuracy_display": _pct_label(avg_accuracy),
        "processing_time": avg_processing,
        "api_status": avg_status,
        "model_config": "-",
        "threshold_status": threshold_status(avg_accuracy)[0],
    }
    latency_average = {
        "test_id": "BATCH AVG",
        "transcription_latency": _avg_ms("transcription_seconds"),
        "translation_latency": _avg_ms("translation_seconds"),
        "llm_latency": _avg_ms("llm_seconds"),
        "total_latency": _avg_ms("total_seconds"),
    }

    all_errors: list[dict[str, Any]] = []
    for case in cases:
        all_errors.extend(case.get("error_analysis") or [])
    type_n = sum(1 for e in all_errors if e.get("category") == CAT_TYPE_MISMATCH)
    format_n = sum(1 for e in all_errors if e.get("category") == CAT_FORMAT_MISMATCH)
    missing_n = sum(1 for e in all_errors if e.get("category") == CAT_MISSING_FROM_GENERATED)
    extra_n = sum(1 for e in all_errors if e.get("category") == CAT_EXTRA_IN_GENERATED)
    value_n = sum(1 for e in all_errors if e.get("category") == CAT_VALUE_MISMATCH)

    return {
        "title": TITLE,
        "generated_at": stamp,
        "batch_id": batch_id or (items[0].get("batch_id") if items else "") or "",
        "total_test_cases": n,
        "overall_batch_accuracy": avg_accuracy,
        "pass_threshold": PASS_THRESHOLD,
        "partial_threshold": PARTIAL_THRESHOLD,
        "columns": list(SUMMARY_COLUMNS),
        "latency_columns": list(LATENCY_COLUMNS),
        "comparison_columns": list(COMPARISON_COLUMNS),
        "summary_rows": summary_rows,
        "averages": averages,
        "latency_rows": latency_rows,
        "latency_average": latency_average,
        "cases": cases,
        "batch_summary": {
            "total_test_cases": n,
            "passed": passed_n,
            "partial": partial_n,
            "failed": failed_n,
            "average_accuracy": avg_accuracy,
            "average_processing_time": avg_processing,
            "average_latency": latency_average["total_latency"],
            "total_errors": len(all_errors),
            "type_mismatches": type_n,
            "format_mismatches": format_n,
            "missing_fields": missing_n,
            "extra_fields": extra_n,
            "value_mismatches": value_n,
        },
    }


def _escape_html(text: Any) -> str:
    return (
        str(text or "")
        .replace("&", "&amp;")
        .replace("<", "&lt;")
        .replace(">", "&gt;")
        .replace('"', "&quot;")
    )


def _escape_pdf(text: Any) -> str:
    return (
        str(text or "")
        .replace("&", "&amp;")
        .replace("<", "&lt;")
        .replace(">", "&gt;")
        .replace("\n", "<br/>")
    )


def _summary_csv_rows(report: dict[str, Any]) -> list[list[str]]:
    rows = [list(SUMMARY_COLUMNS)]
    for item in report.get("summary_rows") or []:
        rows.append(
            [
                item.get("test_id") or "",
                item.get("accuracy_display") or _pct_label(item.get("accuracy") or 0),
                item.get("processing_time") or "",
                item.get("api_status") or "",
                item.get("model_config") or "",
            ]
        )
    avg = report.get("averages") or {}
    if avg:
        rows.append(
            [
                avg.get("test_id") or "BATCH AVG",
                avg.get("accuracy_display") or _pct_label(avg.get("accuracy") or 0),
                avg.get("processing_time") or "",
                avg.get("api_status") or "",
                avg.get("model_config") or "-",
            ]
        )
    return rows


def _latency_csv_rows(report: dict[str, Any]) -> list[list[str]]:
    rows = [list(LATENCY_COLUMNS)]
    for item in report.get("latency_rows") or []:
        rows.append(
            [
                item.get("test_id") or "",
                item.get("transcription_latency") or "",
                item.get("translation_latency") or "",
                item.get("llm_latency") or "",
                item.get("total_latency") or "",
            ]
        )
    avg = report.get("latency_average") or {}
    if avg:
        rows.append(
            [
                avg.get("test_id") or "BATCH AVG",
                avg.get("transcription_latency") or "",
                avg.get("translation_latency") or "",
                avg.get("llm_latency") or "",
                avg.get("total_latency") or "",
            ]
        )
    return rows


def render_batch_json(report: dict[str, Any]) -> bytes:
    return json.dumps(report, ensure_ascii=False, indent=2, default=str).encode("utf-8")


def render_batch_csv(report: dict[str, Any]) -> bytes:
    buffer = io.StringIO()
    writer = csv.writer(buffer)
    writer.writerow([report.get("title") or TITLE])
    writer.writerow(
        [
            f"Total Test Cases: {report.get('total_test_cases') or 0}",
            f"Report Generated: {report.get('generated_at') or ''}",
        ]
    )
    writer.writerow([])
    writer.writerow(["SUMMARY TABLE"])
    for row in _summary_csv_rows(report):
        writer.writerow(row)
    writer.writerow([])
    writer.writerow(["LATENCY ANALYSIS TABLE"])
    for row in _latency_csv_rows(report):
        writer.writerow(row)
    return buffer.getvalue().encode("utf-8-sig")


def _thin_border() -> Border:
    side = Side(style="thin", color="CBD5E1")
    return Border(left=side, right=side, top=side, bottom=side)


def _write_header_row(ws, row: int, headers: tuple[str, ...], border) -> None:
    header_font = Font(bold=True, color="FFFFFF", size=10)
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.fill = _HEADER_FILL
        cell.font = header_font
        cell.border = border


def _write_summary_sheet(ws, report: dict[str, Any]) -> None:
    title_font = Font(bold=True, size=14, color="0F172A")
    wrap = Alignment(wrap_text=True, vertical="top")
    border = _thin_border()
    ws.merge_cells("A1:E1")
    ws["A1"] = report.get("title") or TITLE
    ws["A1"].font = title_font
    ws.merge_cells("A2:E2")
    ws["A2"] = (
        f"Total Test Cases: {report.get('total_test_cases')} | "
        f"Report Generated: {report.get('generated_at')}"
    )
    current = 4
    _write_header_row(ws, current, SUMMARY_COLUMNS, border)
    current += 1
    items = list(report.get("summary_rows") or [])
    if report.get("averages"):
        items.append(report["averages"])
    for item in items:
        values = [
            item.get("test_id"),
            item.get("accuracy_display") or _pct_label(item.get("accuracy") or 0),
            item.get("processing_time"),
            item.get("api_status"),
            item.get("model_config"),
        ]
        fill = _AVG_FILL if item.get("test_id") == "BATCH AVG" else None
        for col, value in enumerate(values, start=1):
            cell = ws.cell(row=current, column=col, value=value)
            if fill:
                cell.fill = fill
            cell.alignment = wrap
            cell.border = border
            if col == 1 and item.get("test_id") == "BATCH AVG":
                cell.font = Font(bold=True)
        current += 1
    current += 2
    _write_batch_summary_block(ws, report, current)
    widths = (18, 14, 20, 16, 28)
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width
    ws.freeze_panes = "A5"


def _write_batch_summary_block(ws, report: dict[str, Any], start_row: int) -> int:
    final = report.get("batch_summary") or {}
    current = start_row
    ws.cell(row=current, column=1, value="FINAL BATCH SUMMARY").font = Font(bold=True, size=12)
    current += 1
    lines = [
        ("Total Test Cases", final.get("total_test_cases")),
        ("Test Cases Passed", final.get("passed")),
        ("Test Cases Partial", final.get("partial")),
        ("Test Cases Failed", final.get("failed")),
        ("Batch Average Accuracy", f"{final.get('average_accuracy')}%"),
        ("Batch Average Processing Time", final.get("average_processing_time")),
        ("Batch Average Latency", final.get("average_latency")),
        ("Critical Issues", f"{final.get('total_errors')} total"),
        ("Type Mismatches", final.get("type_mismatches")),
        ("Format Mismatches", final.get("format_mismatches")),
        ("Missing Fields", final.get("missing_fields")),
        ("Extra Fields", final.get("extra_fields")),
    ]
    for label, value in lines:
        ws.cell(row=current, column=1, value=label).font = Font(bold=True)
        ws.cell(row=current, column=2, value=value)
        current += 1
    return current


def _write_latency_sheet(ws, report: dict[str, Any]) -> None:
    wrap = Alignment(wrap_text=True, vertical="top")
    border = _thin_border()
    ws.merge_cells("A1:E1")
    ws["A1"] = "LATENCY ANALYSIS TABLE"
    ws["A1"].font = Font(bold=True, size=14)
    current = 3
    _write_header_row(ws, current, LATENCY_COLUMNS, border)
    current += 1
    items = list(report.get("latency_rows") or [])
    if report.get("latency_average"):
        items.append(report["latency_average"])
    for item in items:
        values = [
            item.get("test_id"),
            item.get("transcription_latency"),
            item.get("translation_latency"),
            item.get("llm_latency"),
            item.get("total_latency"),
        ]
        fill = _AVG_FILL if item.get("test_id") == "BATCH AVG" else None
        for col, value in enumerate(values, start=1):
            cell = ws.cell(row=current, column=col, value=value)
            if fill:
                cell.fill = fill
            cell.alignment = wrap
            cell.border = border
            if item.get("test_id") == "BATCH AVG":
                cell.font = Font(bold=True)
        current += 1
    widths = (18, 24, 22, 16, 16)
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width
    ws.freeze_panes = "A4"


def _safe_sheet_name(index: int, test_id: str) -> str:
    raw = f"{index}-{test_id or 'case'}"
    cleaned = re.sub(r"[\[\]\*\?:/\\]", "-", raw)[:31]
    return cleaned or f"Case {index}"


def _write_case_sheet(ws, case: dict[str, Any]) -> None:
    header_font = Font(bold=True, color="FFFFFF", size=9)
    wrap = Alignment(wrap_text=True, vertical="top")
    border = _thin_border()
    summary = case.get("summary_row") or {}
    ws.merge_cells("A1:F1")
    ws["A1"] = case["heading"]
    ws["A1"].font = Font(bold=True, size=13)
    current = 3
    ws.cell(row=current, column=1, value="METADATA").font = Font(bold=True)
    current += 1
    for label, value in (
        ("Test ID", summary.get("test_id")),
        ("Accuracy", summary.get("accuracy_display")),
        ("API Endpoint", summary.get("api_endpoint")),
        ("Request ID", summary.get("request_id")),
        ("Status Code", summary.get("api_status")),
        ("Model Config", summary.get("model_config")),
        ("Processing Time", summary.get("processing_time")),
        ("Timestamp", summary.get("timestamp")),
        ("Fields Validated", summary.get("fields_validated")),
        (
            "Fields Passed",
            f"{summary.get('fields_passed')} ({summary.get('passed_percent')}%)",
        ),
        (
            "Fields Failed",
            f"{summary.get('fields_failed')} ({summary.get('failed_percent')}%)",
        ),
    ):
        ws.cell(row=current, column=1, value=label).font = Font(bold=True)
        ws.cell(row=current, column=2, value=value)
        current += 1
    current += 1
    for section in case.get("sections") or []:
        ws.merge_cells(start_row=current, start_column=1, end_row=current, end_column=6)
        cell = ws.cell(row=current, column=1, value=section.get("label"))
        cell.fill = _SECTION_FILL
        cell.font = Font(bold=True, color="FFFFFF")
        current += 1
        for col, header in enumerate(COMPARISON_COLUMNS, start=1):
            cell = ws.cell(row=current, column=col, value=header)
            cell.fill = _HEADER_FILL
            cell.font = header_font
            cell.border = border
        current += 1
        for row in section.get("rows") or []:
            values = [
                _display_name(row),
                row.get("ground_truth_full") or row.get("ground_truth"),
                row.get("generated_full") or row.get("generated"),
                row.get("error_found"),
                row.get("error_description"),
                _row_data_type(row),
            ]
            fill = _PASS_FILL if row.get("passed") else _FAIL_FILL
            for col, value in enumerate(values, start=1):
                cell = ws.cell(row=current, column=col, value=value)
                cell.fill = fill
                cell.alignment = wrap
                cell.border = border
            current += 1
        current += 1
    ws.cell(row=current, column=1, value="ERROR ANALYSIS").font = Font(bold=True)
    current += 1
    for col, header in enumerate(ERROR_COLUMNS, start=1):
        cell = ws.cell(row=current, column=col, value=header)
        cell.fill = _HEADER_FILL
        cell.font = header_font
        cell.border = border
    current += 1
    errors = case.get("error_analysis") or []
    if not errors:
        ws.cell(row=current, column=1, value="No errors.")
        current += 1
    for idx, err in enumerate(errors, start=1):
        values = [
            idx,
            err.get("field"),
            err.get("error_type"),
            err.get("ground_truth"),
            err.get("generated"),
            err.get("impact"),
            err.get("recommendation"),
        ]
        for col, value in enumerate(values, start=1):
            cell = ws.cell(row=current, column=col, value=value)
            cell.fill = _FAIL_FILL
            cell.alignment = wrap
            cell.border = border
            cell.font = Font(color="991B1B")
        current += 1
    widths = (28, 32, 32, 14, 36, 14, 36)
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width


def render_batch_excel(report: dict[str, Any]) -> bytes:
    wb = Workbook()
    summary = wb.active
    summary.title = "Summary"
    _write_summary_sheet(summary, report)
    latency = wb.create_sheet("Latency")
    _write_latency_sheet(latency, report)
    used_names: set[str] = {"Summary", "Latency"}
    for case in report.get("cases") or []:
        name = _safe_sheet_name(case["index"], case["summary_row"]["test_id"])
        base = name
        suffix = 2
        while name in used_names:
            name = f"{base[:28]}-{suffix}"[:31]
            suffix += 1
        used_names.add(name)
        ws = wb.create_sheet(name)
        _write_case_sheet(ws, case)
    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def _html_table(headers: tuple[str, ...], body_rows: list[str], css: str = "") -> str:
    heads = "".join(f"<th>{_escape_html(col)}</th>" for col in headers)
    cls = f" class='{css}'" if css else ""
    return (
        "<div class='table-wrap'><table"
        f"{cls}>"
        f"<thead><tr>{heads}</tr></thead><tbody>{''.join(body_rows)}</tbody></table></div>"
    )


def _html_summary_table(report: dict[str, Any]) -> str:
    body = []
    items = list(report.get("summary_rows") or [])
    if report.get("averages"):
        items.append(report["averages"])
    for item in items:
        is_avg = item.get("test_id") == "BATCH AVG"
        body.append(
            f"<tr class='{'avg' if is_avg else ''}'>"
            f"<td>{_escape_html(item.get('test_id'))}</td>"
            f"<td>{_escape_html(item.get('accuracy_display') or _pct_label(item.get('accuracy') or 0))}</td>"
            f"<td>{_escape_html(item.get('processing_time'))}</td>"
            f"<td>{_escape_html(item.get('api_status'))}</td>"
            f"<td>{_escape_html(item.get('model_config'))}</td>"
            "</tr>"
        )
    return _html_table(SUMMARY_COLUMNS, body, "summary-table")


def _html_latency_table(report: dict[str, Any]) -> str:
    body = []
    items = list(report.get("latency_rows") or [])
    if report.get("latency_average"):
        items.append(report["latency_average"])
    for item in items:
        is_avg = item.get("test_id") == "BATCH AVG"
        body.append(
            f"<tr class='{'avg' if is_avg else ''}'>"
            f"<td>{_escape_html(item.get('test_id'))}</td>"
            f"<td>{_escape_html(item.get('transcription_latency'))}</td>"
            f"<td>{_escape_html(item.get('translation_latency'))}</td>"
            f"<td>{_escape_html(item.get('llm_latency'))}</td>"
            f"<td>{_escape_html(item.get('total_latency'))}</td>"
            "</tr>"
        )
    return _html_table(LATENCY_COLUMNS, body, "latency-table")


def _html_field_table(section: dict[str, Any]) -> str:
    rows = []
    for row in section.get("rows") or []:
        css = "pass" if row.get("passed") else "fail"
        rows.append(
            f"<tr class='{css}'>"
            f"<td>{_escape_html(_display_name(row))}</td>"
            f"<td>{_escape_html(row.get('ground_truth_full') or row.get('ground_truth'))}</td>"
            f"<td>{_escape_html(row.get('generated_full') or row.get('generated'))}</td>"
            f"<td>{_escape_html(row.get('error_found'))}</td>"
            f"<td>{_escape_html(row.get('error_description'))}</td>"
            f"<td>{_escape_html(_row_data_type(row))}</td>"
            "</tr>"
        )
    heads = "".join(f"<th>{_escape_html(col)}</th>" for col in COMPARISON_COLUMNS)
    return (
        f"<h4>{_escape_html(section.get('label'))}</h4>"
        "<div class='table-wrap'><table>"
        f"<thead><tr>{heads}</tr></thead>"
        f"<tbody>{''.join(rows)}</tbody></table></div>"
    )


def _html_error_table(errors: list[dict[str, Any]]) -> str:
    if not errors:
        return "<p>No errors.</p>"
    rows = []
    for idx, err in enumerate(errors, start=1):
        rows.append(
            "<tr class='fail error-row'>"
            f"<td>{idx}</td>"
            f"<td>{_escape_html(err.get('field'))}</td>"
            f"<td>{_escape_html(err.get('error_type'))}</td>"
            f"<td>{_escape_html(err.get('ground_truth'))}</td>"
            f"<td>{_escape_html(err.get('generated'))}</td>"
            f"<td>{_escape_html(err.get('impact'))}</td>"
            f"<td>{_escape_html(err.get('recommendation'))}</td>"
            "</tr>"
        )
    return _html_table(ERROR_COLUMNS, rows, "error-table")


def _html_case(case: dict[str, Any], *, open_first: bool) -> str:
    summary = case.get("summary_row") or {}
    open_attr = " open" if open_first else ""
    meta_rows = "".join(
        f"<tr><th>{_escape_html(label)}</th><td>{_escape_html(value)}</td></tr>"
        for label, value in (
            ("Test ID", summary.get("test_id")),
            ("Accuracy", summary.get("accuracy_display")),
            ("API Endpoint", summary.get("api_endpoint")),
            ("Request ID", summary.get("request_id")),
            ("Status Code", summary.get("api_status")),
            ("Model Config", summary.get("model_config")),
            ("Processing Time", summary.get("processing_time")),
            ("Timestamp", summary.get("timestamp")),
        )
    )
    sections = "".join(_html_field_table(s) for s in case.get("sections") or [])
    return f"""
<details class="case"{open_attr}>
  <summary>{_escape_html(case['heading'])}</summary>
  <h3>Metadata</h3>
  <table class="meta">{meta_rows}</table>
  <h3>Summary</h3>
  <p>Fields Validated: {summary.get('fields_validated')} ·
     Fields Passed: {summary.get('fields_passed')} ({summary.get('passed_percent')}%) ·
     Fields Failed: {summary.get('fields_failed')} ({summary.get('failed_percent')}%)</p>
  <h3>Comparison Tables</h3>
  {sections}
  <h3>Error Analysis</h3>
  {_html_error_table(case.get('error_analysis') or [])}
</details>
"""


def render_batch_html(report: dict[str, Any]) -> bytes:
    final = report.get("batch_summary") or {}
    cases = "".join(
        _html_case(case, open_first=(idx == 0))
        for idx, case in enumerate(report.get("cases") or [])
    )
    html = f"""<!DOCTYPE html>
<html lang="en">
<head>
  <meta charset="utf-8"/>
  <title>{_escape_html(report.get('title') or TITLE)}</title>
  <style>
    body {{ font-family: Segoe UI, system-ui, sans-serif; margin: 24px; color: #0f172a; }}
    h1 {{ font-size: 22px; margin-bottom: 4px; }}
    .sub {{ color: #475569; margin-bottom: 16px; }}
    .table-wrap {{ overflow-x: auto; margin: 12px 0 24px; }}
    table {{ border-collapse: collapse; width: 100%; font-size: 13px; }}
    th, td {{ border: 1px solid #e2e8f0; padding: 8px 10px; vertical-align: top; text-align: left; }}
    th {{ background: #1e293b; color: #fff; font-size: 11px; text-transform: uppercase; }}
    tr.pass td {{ background: #f0fdf4; }}
    tr.fail td {{ background: #fef2f2; color: #991b1b; }}
    tr.avg td {{ font-weight: 700; background: #e2e8f0; }}
    table.error-table tr.fail td {{ background: #fecaca; color: #991b1b; }}
    table.meta th {{ width: 220px; background: #f8fafc; color: #334155; text-transform: none; }}
    details.case {{ border: 1px solid #e2e8f0; border-radius: 8px; margin: 12px 0; padding: 8px 12px; background: #fff; }}
    details.case > summary {{ cursor: pointer; font-weight: 700; padding: 8px 0; }}
    .final {{ background: #f8fafc; border: 1px solid #e2e8f0; padding: 16px; border-radius: 8px; }}
  </style>
</head>
<body>
  <h1>{_escape_html(report.get('title') or TITLE)}</h1>
  <p class="sub">Total Test Cases: {report.get('total_test_cases')} |
     Report Generated: {_escape_html(report.get('generated_at'))}</p>
  <h2>Summary Table</h2>
  {_html_summary_table(report)}
  <h2>Latency Analysis Table</h2>
  {_html_latency_table(report)}
  <h2>Detailed Test Case Sections</h2>
  {cases}
  <div class="final">
    <h2>Final Batch Summary</h2>
    <p>Total Test Cases: {final.get('total_test_cases')}<br/>
       Test Cases Passed: {final.get('passed')}<br/>
       Test Cases Partial: {final.get('partial')}<br/>
       Test Cases Failed: {final.get('failed')}</p>
    <p>Batch Average Accuracy: {final.get('average_accuracy')}%<br/>
       Batch Average Processing Time: {_escape_html(final.get('average_processing_time'))}<br/>
       Batch Average Latency: {_escape_html(final.get('average_latency'))}</p>
    <p>Critical Issues: {final.get('total_errors')} total<br/>
       Type Mismatches: {final.get('type_mismatches')}<br/>
       Format Mismatches: {final.get('format_mismatches')}<br/>
       Missing Fields: {final.get('missing_fields')}<br/>
       Extra Fields: {final.get('extra_fields')}</p>
  </div>
</body>
</html>
"""
    return html.encode("utf-8")


def _pdf_styles() -> dict[str, ParagraphStyle]:
    styles = getSampleStyleSheet()
    return {
        "title": ParagraphStyle(
            "BatchTitle", parent=styles["Heading1"], fontSize=14, spaceAfter=4,
            textColor=colors.HexColor("#0f172a"),
        ),
        "sub": ParagraphStyle(
            "BatchSub", parent=styles["Normal"], fontSize=9,
            textColor=colors.HexColor("#334155"), spaceAfter=8,
        ),
        "h2": ParagraphStyle(
            "BatchH2", parent=styles["Heading2"], fontSize=11, spaceBefore=10,
            spaceAfter=6, textColor=colors.HexColor("#1e293b"),
        ),
        "cell": ParagraphStyle(
            "BatchCell", parent=styles["Normal"], fontSize=7, leading=9,
        ),
        "fail": ParagraphStyle(
            "BatchFail", parent=styles["Normal"], fontSize=7, leading=9,
            textColor=colors.HexColor("#991b1b"),
        ),
        "head": ParagraphStyle(
            "BatchHead", parent=styles["Normal"], fontName="Helvetica-Bold",
            fontSize=7, textColor=colors.white, leading=9,
        ),
    }


def _pdf_table(data: list[list], widths: list, fills: list, styles: dict) -> Table:
    table = Table(data, colWidths=widths, repeatRows=1)
    cmds = [
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1e293b")),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#cbd5e1")),
        ("LEFTPADDING", (0, 0), (-1, -1), 4),
        ("RIGHTPADDING", (0, 0), (-1, -1), 4),
        ("TOPPADDING", (0, 0), (-1, -1), 3),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
    ]
    for idx, fill in enumerate(fills, start=1):
        cmds.append(("BACKGROUND", (0, idx), (-1, idx), fill))
    table.setStyle(TableStyle(cmds))
    return table


def _pdf_summary_table(report: dict[str, Any], styles: dict) -> Table:
    header = [Paragraph(_escape_pdf(col), styles["head"]) for col in SUMMARY_COLUMNS]
    data = [header]
    fills = []
    items = list(report.get("summary_rows") or [])
    if report.get("averages"):
        items.append(report["averages"])
    for item in items:
        data.append(
            [
                Paragraph(_escape_pdf(item.get("test_id")), styles["cell"]),
                Paragraph(
                    _escape_pdf(item.get("accuracy_display") or _pct_label(item.get("accuracy") or 0)),
                    styles["cell"],
                ),
                Paragraph(_escape_pdf(item.get("processing_time")), styles["cell"]),
                Paragraph(_escape_pdf(item.get("api_status")), styles["cell"]),
                Paragraph(_escape_pdf(item.get("model_config")), styles["cell"]),
            ]
        )
        fills.append(
            colors.HexColor("#e2e8f0")
            if item.get("test_id") == "BATCH AVG"
            else colors.HexColor("#f8fafc")
        )
    widths = [4.4 * cm, 3.6 * cm, 4.4 * cm, 3.8 * cm, 8.0 * cm]
    return _pdf_table(data, widths, fills, styles)


def _pdf_latency_table(report: dict[str, Any], styles: dict) -> Table:
    header = [Paragraph(_escape_pdf(col), styles["head"]) for col in LATENCY_COLUMNS]
    data = [header]
    fills = []
    items = list(report.get("latency_rows") or [])
    if report.get("latency_average"):
        items.append(report["latency_average"])
    for item in items:
        data.append(
            [
                Paragraph(_escape_pdf(item.get("test_id")), styles["cell"]),
                Paragraph(_escape_pdf(item.get("transcription_latency")), styles["cell"]),
                Paragraph(_escape_pdf(item.get("translation_latency")), styles["cell"]),
                Paragraph(_escape_pdf(item.get("llm_latency")), styles["cell"]),
                Paragraph(_escape_pdf(item.get("total_latency")), styles["cell"]),
            ]
        )
        fills.append(
            colors.HexColor("#e2e8f0")
            if item.get("test_id") == "BATCH AVG"
            else colors.HexColor("#f8fafc")
        )
    widths = [4.4 * cm, 5.4 * cm, 5.2 * cm, 4.2 * cm, 4.2 * cm]
    return _pdf_table(data, widths, fills, styles)


def _pdf_field_table(section: dict[str, Any], styles: dict) -> Table:
    header = [Paragraph(_escape_pdf(col), styles["head"]) for col in COMPARISON_COLUMNS]
    data = [header]
    fills = []
    for row in section.get("rows") or []:
        data.append(
            [
                Paragraph(_escape_pdf(_display_name(row)), styles["cell"]),
                Paragraph(
                    _escape_pdf(row.get("ground_truth_full") or row.get("ground_truth")),
                    styles["cell"],
                ),
                Paragraph(
                    _escape_pdf(row.get("generated_full") or row.get("generated")),
                    styles["cell"],
                ),
                Paragraph(_escape_pdf(row.get("error_found")), styles["cell"]),
                Paragraph(_escape_pdf(row.get("error_description")), styles["cell"]),
                Paragraph(_escape_pdf(_row_data_type(row)), styles["cell"]),
            ]
        )
        fills.append(
            colors.HexColor("#d1fae5") if row.get("passed") else colors.HexColor("#fecaca")
        )
    widths = [4.2 * cm, 5.2 * cm, 5.2 * cm, 2.2 * cm, 5.4 * cm, 2.4 * cm]
    return _pdf_table(data, widths, fills, styles)


def _pdf_error_table(errors: list[dict[str, Any]], styles: dict) -> Table:
    header = [Paragraph(_escape_pdf(col), styles["head"]) for col in ERROR_COLUMNS]
    data = [header]
    fills = []
    for idx, err in enumerate(errors, start=1):
        data.append(
            [
                Paragraph(str(idx), styles["fail"]),
                Paragraph(_escape_pdf(err.get("field")), styles["fail"]),
                Paragraph(_escape_pdf(err.get("error_type")), styles["fail"]),
                Paragraph(_escape_pdf(err.get("ground_truth")), styles["fail"]),
                Paragraph(_escape_pdf(err.get("generated")), styles["fail"]),
                Paragraph(_escape_pdf(err.get("impact")), styles["fail"]),
                Paragraph(_escape_pdf(err.get("recommendation")), styles["fail"]),
            ]
        )
        fills.append(colors.HexColor("#fecaca"))
    widths = [1.6 * cm, 3.6 * cm, 3.4 * cm, 4.0 * cm, 4.0 * cm, 2.0 * cm, 5.6 * cm]
    return _pdf_table(data, widths, fills, styles)


def render_batch_pdf(report: dict[str, Any]) -> bytes:
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=landscape(A4),
        leftMargin=1.1 * cm,
        rightMargin=1.1 * cm,
        topMargin=1.1 * cm,
        bottomMargin=1.1 * cm,
    )
    styles = _pdf_styles()
    story: list[Any] = [
        Paragraph(_escape_pdf(report.get("title") or TITLE), styles["title"]),
        Paragraph(
            _escape_pdf(
                f"Total Test Cases: {report.get('total_test_cases')} | "
                f"Report Generated: {report.get('generated_at')}"
            ),
            styles["sub"],
        ),
        Paragraph("Summary Table", styles["h2"]),
        _pdf_summary_table(report, styles),
        Paragraph("Latency Analysis Table", styles["h2"]),
        _pdf_latency_table(report, styles),
    ]
    for case in report.get("cases") or []:
        story.append(PageBreak())
        summary = case.get("summary_row") or {}
        blocks: list[Any] = [
            Paragraph(_escape_pdf(case["heading"]), styles["h2"]),
            Paragraph(
                _escape_pdf(
                    f"Test ID: {summary.get('test_id')} | "
                    f"Accuracy: {summary.get('accuracy_display')} | "
                    f"API Endpoint: {summary.get('api_endpoint')} | "
                    f"Request ID: {summary.get('request_id')} | "
                    f"Status Code: {summary.get('api_status')} | "
                    f"Model Config: {summary.get('model_config')} | "
                    f"Processing Time: {summary.get('processing_time')} | "
                    f"Timestamp: {summary.get('timestamp')}"
                ),
                styles["sub"],
            ),
            Paragraph(
                _escape_pdf(
                    f"Fields Validated: {summary.get('fields_validated')} | "
                    f"Fields Passed: {summary.get('fields_passed')} "
                    f"({summary.get('passed_percent')}%) | "
                    f"Fields Failed: {summary.get('fields_failed')} "
                    f"({summary.get('failed_percent')}%)"
                ),
                styles["sub"],
            ),
        ]
        story.append(KeepTogether(blocks))
        for section in case.get("sections") or []:
            story.append(Paragraph(_escape_pdf(section.get("label")), styles["h2"]))
            story.append(_pdf_field_table(section, styles))
        errors = case.get("error_analysis") or []
        story.append(Paragraph("Error Analysis", styles["h2"]))
        if errors:
            story.append(_pdf_error_table(errors, styles))
        else:
            story.append(Paragraph("No errors.", styles["sub"]))

    final = report.get("batch_summary") or {}
    story.append(PageBreak())
    story.append(Paragraph("Final Batch Summary", styles["h2"]))
    story.append(
        Paragraph(
            _escape_pdf(
                f"Total Test Cases: {final.get('total_test_cases')} | "
                f"Test Cases Passed: {final.get('passed')} | "
                f"Test Cases Partial: {final.get('partial')} | "
                f"Test Cases Failed: {final.get('failed')}"
            ),
            styles["sub"],
        )
    )
    story.append(
        Paragraph(
            _escape_pdf(
                f"Batch Average Accuracy: {final.get('average_accuracy')}% | "
                f"Batch Average Processing Time: {final.get('average_processing_time')} | "
                f"Batch Average Latency: {final.get('average_latency')}"
            ),
            styles["sub"],
        )
    )
    story.append(
        Paragraph(
            _escape_pdf(
                f"Critical Issues: {final.get('total_errors')} total | "
                f"Type Mismatches: {final.get('type_mismatches')} | "
                f"Format Mismatches: {final.get('format_mismatches')} | "
                f"Missing Fields: {final.get('missing_fields')} | "
                f"Extra Fields: {final.get('extra_fields')}"
            ),
            styles["sub"],
        )
    )
    doc.build(story)
    return buffer.getvalue()


_RENDERERS = {
    "json": (render_batch_json, "application/json", "json"),
    "csv": (render_batch_csv, "text/csv", "csv"),
    "excel": (
        render_batch_excel,
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        "xlsx",
    ),
    "html": (render_batch_html, "text/html", "html"),
    "pdf": (render_batch_pdf, "application/pdf", "pdf"),
}


def render_soap_batch_comparison_report(
    report: dict[str, Any], fmt: str
) -> tuple[bytes, str, str]:
    key = (fmt or "json").lower()
    if key in ("xlsx", "xls"):
        key = "excel"
    if key not in _RENDERERS:
        raise ValueError(f"format must be one of {', '.join(SUPPORTED_FORMATS)}")
    renderer, mime, ext = _RENDERERS[key]
    return renderer(report), mime, ext
