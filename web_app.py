"""
Medisum Load Test Web UI
Run:  python web_app.py
Open: http://127.0.0.1:5050
"""

import base64
import io
import json
import queue
import re
import threading
import time
import uuid
import wave
from concurrent.futures import ThreadPoolExecutor, as_completed
from datetime import datetime

import requests
from flask import Flask, Response, request, stream_with_context, send_file
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

app = Flask(__name__)

# Global storage for test results (for export)
test_results = []

# ── Change these two lines to point to a different environment ────────────────
DJANGO_BASE_URL = "https://medsum.amritaai.org"
FLASK_BASE_URL  = "https://test-medsum.amritaai.org/transcribe"
# ─────────────────────────────────────────────────────────────────────────────

# ─────────────────────────────────────────────────────────────────────────────
# Templates
# ─────────────────────────────────────────────────────────────────────────────

SOAP_TEMPLATE = {
    "subjective": {
        "chief_complaint": "", "history_of_present_illness": "",
        "past_medical_history": "", "medications": "", "allergies": "",
        "social_history": "", "family_history": "",
    },
    "objective": {
        "vitals": {"blood_pressure": "", "heart_rate": "", "respiratory_rate": "", "temperature": ""},
        "physical_exam": {"heart": []},
    },
    "assessment": [],
    "plan": {
        "medications": [{"drug_name": "", "dose": "", "schedule": "", "duration": "", "instructions": "", "snomed_ct_id": ""}],
        "activity": "", "investigations": "", "education": "", "follow_up": "",
    },
    "summary": "",
}

DISCHARGE_TEMPLATE = {
    "Patient_Details": {"Name": "", "Age": "", "Gender": "", "Blood_Group": "", "ABHA_ID": ""},
    "Admission_Details": {"Date_of_Admission": "", "Date_of_Discharge": "", "Ward": "", "Bed_Number": "", "Consultant": "", "Department": ""},
    "Presenting_Complaints": "", "History_of_Presenting_Illness": "", "Past_Medical_History": "",
    "Personal_History": "", "Family_History": "",
    "Examination_Findings": {"General_Examination": "", "Systemic_Examination": ""},
    "Investigations": "",
    "Diagnosis": {"Primary_Diagnosis": "", "Secondary_Diagnosis": ""},
    "Course_in_the_Hospital_and_Discussion": "", "Operative_or_Procedure_Findings": "",
    "Prognosis_on_Discharge": "", "Diet_Recommendation": "", "Physiscal_Activity": "",
    "Discharge_Medication": {"Serial_Number": [], "Drug_Name": [], "Dose": [], "Schedule": [], "Duration": [], "Days": [], "Instructions": []},
    "Follow_Up": "", "summary": "",
}

KNOWN_META_KEYS = {
    "doctor_details", "patient_demographics", "session",
    "transcription", "transcription-time", "translation-time",
    "llm-time", "audio_length", "total-time", "medicines_backend",
}

# ─────────────────────────────────────────────────────────────────────────────
# HTML
# ─────────────────────────────────────────────────────────────────────────────

HTML = r"""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>Medisum Load Test</title>
<style>
  :root {
    --bg: #f5f6fa;
    --surface: #ffffff;
    --surface2: #f0f1f5;
    --border: #d9dce6;
    --accent: #2563eb;
    --accent2: #7c3aed;
    --green: #16a34a;
    --red: #dc2626;
    --yellow: #d97706;
    --orange: #ea580c;
    --text: #111827;
    --muted: #6b7280;
    --mono: 'Menlo', 'Consolas', monospace;
  }
  * { box-sizing: border-box; margin: 0; padding: 0; }
  body { background: var(--bg); color: var(--text); font-family: -apple-system, BlinkMacSystemFont, 'Segoe UI', sans-serif; font-size: 14px; min-height: 100vh; display: flex; flex-direction: column; }

  /* Header */
  header { background: var(--surface); border-bottom: 1px solid var(--border); padding: 14px 28px; display: flex; align-items: center; gap: 12px; position: sticky; top: 0; z-index: 100; }
  header h1 { font-size: 17px; font-weight: 700; }
  .badge { background: var(--accent2); color: #fff; font-size: 11px; font-weight: 600; padding: 2px 8px; border-radius: 20px; }
  .header-right { margin-left: auto; display: flex; align-items: center; gap: 10px; font-size: 13px; color: var(--muted); }

  /* Screens */
  .screen { display: none; flex: 1; padding: 24px 28px 0; max-width: 1200px; width: 100%; margin: 0 auto; }
  .screen.active { display: block; }

  /* Cards */
  .card { background: var(--surface); border: 1px solid var(--border); border-radius: 10px; padding: 18px; margin-bottom: 18px; }
  .card-title { font-size: 12px; font-weight: 700; text-transform: uppercase; letter-spacing: 1px; color: var(--muted); margin-bottom: 14px; display: flex; align-items: center; gap: 8px; }
  .dot { width: 6px; height: 6px; border-radius: 50%; background: var(--accent); flex-shrink: 0; }

  /* Fields */
  .field { margin-bottom: 12px; }
  .field:last-child { margin-bottom: 0; }
  .field label { display: block; font-size: 12px; color: var(--muted); margin-bottom: 4px; }
  .field input, .field select { width: 100%; background: var(--bg); border: 1px solid var(--border); border-radius: 6px; color: var(--text); padding: 7px 10px; font-size: 13px; outline: none; transition: border-color .15s; }
  .field input:focus, .field select:focus { border-color: var(--accent); }
  .field input::placeholder { color: #aaa; }
  select option { background: #fff; }

  /* Doctor cards grid */
  #doctors-list { display: grid; grid-template-columns: repeat(auto-fill, minmax(340px, 1fr)); gap: 14px; margin-bottom: 14px; }
  .doctor-card { background: var(--bg); border: 1px solid var(--border); border-radius: 8px; padding: 14px; position: relative; }
  .doctor-card-header { display: flex; align-items: center; gap: 8px; margin-bottom: 12px; font-size: 12px; font-weight: 600; color: var(--muted); }
  .doctor-num { background: var(--accent); color: #fff; width: 20px; height: 20px; border-radius: 50%; display: flex; align-items: center; justify-content: center; font-size: 11px; font-weight: 700; flex-shrink: 0; }
  .remove-btn { margin-left: auto; background: none; border: none; color: #ccc; cursor: pointer; font-size: 16px; line-height: 1; padding: 0; }
  .remove-btn:hover { color: var(--red); }
  .doctor-fields { display: grid; grid-template-columns: 1fr 1fr; gap: 10px; margin-bottom: 10px; }
  .add-btn { background: none; border: 1.5px dashed var(--border); border-radius: 8px; color: var(--muted); font-size: 13px; font-weight: 500; cursor: pointer; padding: 10px; width: 100%; text-align: center; transition: border-color .15s, color .15s; }
  .add-btn:hover { border-color: var(--accent); color: var(--accent); }

  /* Patient tag input */
  .tag-box { background: var(--bg); border: 1px solid var(--border); border-radius: 6px; padding: 5px 8px; display: flex; flex-wrap: wrap; gap: 5px; align-items: center; min-height: 36px; cursor: text; transition: border-color .15s; }
  .tag-box:focus-within { border-color: var(--accent); }
  .tag { background: #dbeafe; color: #1e40af; font-size: 12px; font-weight: 600; padding: 2px 8px; border-radius: 4px; display: flex; align-items: center; gap: 4px; }
  .tag button { background: none; border: none; color: #60a5fa; cursor: pointer; font-size: 13px; line-height: 1; padding: 0; }
  .tag button:hover { color: var(--red); }
  .tag-input { border: none; outline: none; background: transparent; font-size: 12px; color: var(--text); min-width: 80px; flex: 1; padding: 2px 0; }
  .tag-hint { font-size: 11px; color: #aaa; margin-top: 3px; }

  /* Advanced */
  .adv-toggle { background: none; border: none; color: var(--muted); font-size: 12px; cursor: pointer; display: flex; align-items: center; gap: 6px; padding: 0; margin-bottom: 12px; }
  .adv-toggle:hover { color: var(--text); }
  .adv-arrow { transition: transform .2s; display: inline-block; }
  .adv-toggle.open .adv-arrow { transform: rotate(90deg); }
  .adv-fields { display: none; }
  .adv-fields.open { display: block; }
  .adv-grid { display: grid; grid-template-columns: repeat(auto-fill, minmax(200px, 1fr)); gap: 12px; }

  /* Action buttons */
  .btn { border: none; border-radius: 7px; padding: 10px 24px; font-size: 13px; font-weight: 600; cursor: pointer; transition: opacity .15s; display: inline-flex; align-items: center; gap: 8px; }
  .btn:hover { opacity: 0.88; }
  .btn:disabled { opacity: 0.45; cursor: not-allowed; }
  .btn-primary { background: var(--accent); color: #fff; }
  .btn-secondary { background: var(--surface2); color: var(--text); border: 1px solid var(--border); }
  .btn-green { background: var(--green); color: #fff; }
  .btn-danger { background: var(--red); color: #fff; }
  .action-bar { display: flex; align-items: center; gap: 12px; margin-bottom: 20px; flex-wrap: wrap; }
  .spinner-sm { width: 14px; height: 14px; border: 2px solid rgba(255,255,255,0.3); border-top-color: #fff; border-radius: 50%; animation: spin .7s linear infinite; }
  @keyframes spin { to { transform: rotate(360deg); } }

  /* Stats bar */
  .stats-bar { background: var(--surface); border: 1px solid var(--border); border-radius: 8px; padding: 12px 18px; display: flex; gap: 28px; flex-wrap: wrap; margin-bottom: 18px; }
  .stat { display: flex; flex-direction: column; gap: 2px; }
  .stat-label { font-size: 10px; text-transform: uppercase; letter-spacing: 0.5px; color: var(--muted); font-weight: 600; }
  .stat-value { font-size: 22px; font-weight: 700; color: var(--text); font-variant-numeric: tabular-nums; }
  .stat-value.green { color: var(--green); }
  .stat-value.red { color: var(--red); }
  .stat-value.blue { color: var(--accent); }

  /* Preview / Run table */
  .table-wrap { background: var(--surface); border: 1px solid var(--border); border-radius: 10px; overflow: hidden; margin-bottom: 20px; }
  table { width: 100%; border-collapse: collapse; font-size: 13px; }
  th { background: var(--surface2); padding: 10px 14px; text-align: left; font-size: 11px; font-weight: 700; text-transform: uppercase; letter-spacing: 0.5px; color: var(--muted); border-bottom: 1px solid var(--border); white-space: nowrap; }
  td { padding: 10px 14px; border-bottom: 1px solid var(--border); vertical-align: middle; }
  tr:last-child td { border-bottom: none; }
  tr:hover td { background: var(--bg); }
  .mono { font-family: var(--mono); font-size: 12px; }

  /* Status badges */
  .badge-status { display: inline-flex; align-items: center; gap: 5px; font-size: 11px; font-weight: 700; padding: 3px 9px; border-radius: 20px; white-space: nowrap; }
  .badge-pending { background: #f3f4f6; color: var(--muted); }
  .badge-running { background: #dbeafe; color: #1d4ed8; }
  .badge-pass    { background: #dcfce7; color: #15803d; }
  .badge-fail    { background: #fee2e2; color: #dc2626; }
  .badge-skip    { background: #fef9c3; color: #92400e; }
  .dot-spin { width: 8px; height: 8px; border: 1.5px solid rgba(29,78,216,0.3); border-top-color: #1d4ed8; border-radius: 50%; animation: spin .7s linear infinite; }

  /* Error text */
  .err-text { font-size: 11px; color: var(--red); margin-top: 3px; font-family: var(--mono); }

  /* Expandable transcription/summary row */
  .expand-row td { padding: 0; background: #f8faff; border-bottom: 1px solid var(--border); }
  .expand-inner { display: grid; grid-template-columns: 1fr 1fr; gap: 0; }
  .expand-section { padding: 14px 18px; border-right: 1px solid var(--border); }
  .expand-section:last-child { border-right: none; }
  .expand-label { font-size: 10px; font-weight: 700; text-transform: uppercase; letter-spacing: 0.8px; color: var(--muted); margin-bottom: 8px; }
  .expand-text { font-family: var(--mono); font-size: 12px; line-height: 1.7; color: var(--text); white-space: pre-wrap; max-height: 200px; overflow-y: auto; }
  .view-btn { background: none; border: 1px solid var(--border); color: var(--accent); font-size: 11px; font-weight: 600; padding: 3px 10px; border-radius: 4px; cursor: pointer; white-space: nowrap; }
  .view-btn:hover { background: #eff6ff; }

  /* Progress bar */
  .progress-wrap { background: #e5e7eb; border-radius: 99px; height: 6px; flex: 1; min-width: 100px; overflow: hidden; }
  .progress-bar { height: 100%; background: var(--accent); border-radius: 99px; transition: width .4s ease; }

  /* Status bar */
  #status-bar { padding: 10px 14px; border-radius: 8px; margin-bottom: 14px; font-size: 13px; display: none; align-items: center; gap: 10px; }
  #status-bar.info    { background: #eff6ff; border: 1px solid #bfdbfe; color: #1d4ed8; }
  #status-bar.success { background: #f0fdf4; border: 1px solid #bbf7d0; color: #15803d; }
  #status-bar.error   { background: #fef2f2; border: 1px solid #fecaca; color: #dc2626; }

  /* Logs */
  #logs-panel { position: sticky; bottom: 0; background: var(--surface); border: 1px solid var(--border); border-bottom: none; border-radius: 10px 10px 0 0; overflow: hidden; margin-top: 24px; }
  .logs-header { background: var(--surface2); padding: 9px 14px; display: flex; align-items: center; justify-content: space-between; cursor: pointer; user-select: none; border-bottom: 1px solid var(--border); }
  .logs-header-left { display: flex; align-items: center; gap: 10px; font-size: 12px; font-weight: 600; color: var(--muted); }
  .logs-dot { width: 8px; height: 8px; border-radius: 50%; background: var(--muted); transition: background .3s; }
  .logs-dot.active { background: var(--green); animation: pulse 1s ease-in-out infinite; }
  @keyframes pulse { 0%,100%{opacity:1}50%{opacity:.4} }
  .logs-actions { display: flex; gap: 8px; }
  .logs-btn { background: none; border: 1px solid var(--border); color: var(--muted); font-size: 11px; padding: 3px 10px; border-radius: 4px; cursor: pointer; }
  .logs-btn:hover { color: var(--text); }
  #logs-body { height: 180px; overflow-y: auto; padding: 10px 14px; font-family: var(--mono); font-size: 12px; line-height: 1.6; transition: height .25s ease, padding .25s ease; }
  #logs-body.minimized { height: 0; padding: 0 14px; overflow: hidden; }
  .log-line { margin-bottom: 1px; word-break: break-all; color: var(--muted); }
  .log-line.ok   { color: #15803d; }
  .log-line.warn { color: #b45309; }
  .log-line.fail { color: #dc2626; }
  .log-line.step { color: #2563eb; }

  /* Per-patient audio rows inside doctor cards */
  .patient-audio-list { margin-top: 10px; display: flex; flex-direction: column; gap: 6px; }
  .patient-audio-row { background: var(--surface); border: 1px solid var(--border); border-radius: 6px; padding: 8px 12px; }
  .pat-audio-header { display: flex; align-items: center; gap: 8px; margin-bottom: 5px; }
  .pat-pid-badge { background: #dbeafe; color: #1e40af; font-size: 11px; font-weight: 700; padding: 2px 8px; border-radius: 4px; }
  .pat-upload-label { margin-left: auto; font-size: 11px; color: var(--accent); cursor: pointer; border: 1px dashed var(--border); padding: 2px 9px; border-radius: 4px; transition: border-color .15s; white-space: nowrap; }
  .pat-upload-label:hover { border-color: var(--accent); }
  .pat-file-list { display: flex; flex-wrap: wrap; gap: 5px; min-height: 18px; }
  .pat-file-chip { background: #f0fdf4; border: 1px solid #bbf7d0; border-radius: 4px; padding: 2px 7px; font-size: 11px; display: flex; align-items: center; gap: 4px; max-width: 240px; }
  .pat-file-chip span { overflow: hidden; text-overflow: ellipsis; white-space: nowrap; }
  .pat-file-chip button { background: none; border: none; color: #aaa; cursor: pointer; font-size: 12px; line-height: 1; padding: 0; flex-shrink: 0; }
  .pat-file-chip button:hover { color: var(--red); }
  .pat-no-audio { font-size: 11px; color: var(--muted); font-style: italic; }
  .pat-lang-row { display: flex; align-items: center; gap: 6px; margin-bottom: 5px; }
  .pat-lang-label { font-size: 11px; color: var(--muted); white-space: nowrap; }
  .pat-lang-select { font-size: 11px; background: var(--bg); border: 1px solid var(--border); border-radius: 4px; color: var(--text); padding: 3px 6px; outline: none; cursor: pointer; }
  .pat-lang-select:focus { border-color: var(--accent); }
</style>
</head>
<body>

<header>
  <h1>Medisum Load Test Console</h1>
  <span class="badge">Load Testing</span>
  <div class="header-right" id="header-progress" style="display:none;">
    <span id="hdr-count">0 / 0</span>
    <div class="progress-wrap" style="width:160px;">
      <div class="progress-bar" id="hdr-bar" style="width:0%"></div>
    </div>
  </div>
</header>

<!-- ═══════════════════════════ SCREEN 1: INPUT ═══════════════════════════ -->
<div class="screen active" id="s-input">

  <!-- Doctors -->
  <div class="card">
    <div class="card-title"><span class="dot"></span> Doctors</div>
    <div id="doctors-list"></div>
    <button type="button" class="add-btn" id="add-doctor-btn">+ Add Doctor</button>
  </div>

  <!-- Advanced -->
  <div class="card">
    <button type="button" class="adv-toggle" id="adv-toggle">
      <span class="adv-arrow">▶</span>&nbsp; Advanced Settings
    </button>
    <div class="adv-fields" id="adv-fields">
      <div class="adv-grid">
        <div class="field">
          <label>LLM</label>
          <select id="cfg-llm">
            <option value="OpenAI">OpenAI</option>
            <option value="Gemma">Gemma</option>
            <option value="Param">Param</option>
          </select>
        </div>
        <div class="field">
          <label>STT Model</label>
          <select id="cfg-stt-model">
            <option value="Bhasini">Bhasini</option>
            <option value="Bharatgen">Bharatgen</option>
          </select>
        </div>
        <div class="field">
          <label>Translate Model</label>
          <select id="cfg-translate-model">
            <option value="Bhasini">Bhasini</option>
            <option value="Bharatgen">Bharatgen</option>
          </select>
        </div>
        <div class="field">
          <label>Template Type</label>
          <select id="cfg-template-type">
            <option value="soap">SOAP</option>
            <option value="discharge">Discharge Summary</option>
          </select>
        </div>
        <div class="field">
          <label>Template ID</label>
          <input id="cfg-template-id" type="text" value="1">
        </div>
        <div class="field">
          <label>Audio Duration (s) — silent WAV fallback</label>
          <input id="cfg-duration" type="number" value="2" step="0.1" min="0.5">
        </div>
        <div class="field">
          <label>Django Base URL</label>
          <select id="cfg-django">
            <option value="https://test-medsum.amritaai.org">https://test-medsum.amritaai.org</option>
            <option value="https://medsum.bharatgen.dev">https://medsum.bharatgen.dev</option>
          </select>
        </div>
        <div class="field">
          <label>Flask Transcribe URL</label>
          <select id="cfg-flask">
            <option value="https://test-medsum.amritaai.org/transcribe">https://test-medsum.amritaai.org/transcribe</option>
            <option value="https://medsum.bharatgen.dev/transcribe">https://medsum.bharatgen.dev/transcribe</option>
          </select>
        </div>
      </div>
    </div>
  </div>

  <div id="status-bar"></div>
  <div style="margin-bottom:28px;">
    <button class="btn btn-primary" id="preview-btn" style="width:100%;justify-content:center;padding:12px;">
      Preview Test Matrix →
    </button>
  </div>
</div>

<!-- ═══════════════════════════ SCREEN 2: PREVIEW / RUN ═══════════════════════════ -->
<div class="screen" id="s-run">

  <div class="action-bar" id="action-bar">
    <button class="btn btn-secondary" id="back-btn">← Back</button>
    <button class="btn btn-green" id="run-btn">
      <span id="run-btn-text">Confirm &amp; Run</span>
      <div class="spinner-sm" id="run-spinner" style="display:none;"></div>
    </button>
    <button class="btn btn-secondary" id="new-test-btn" style="display:none;">↺ New Test</button>
    <button class="btn btn-primary" id="export-btn" style="display:none;">⬇ Export to Excel</button>
    <div class="progress-wrap" id="run-progress-wrap" style="display:none;flex:1;min-width:80px;">
      <div class="progress-bar" id="run-bar" style="width:0%"></div>
    </div>
    <span id="run-counter" style="font-size:13px;color:var(--muted);display:none;"></span>
  </div>

  <!-- Stats -->
  <div class="stats-bar" id="stats-bar">
    <div class="stat"><span class="stat-label">Doctors</span><span class="stat-value blue" id="stat-doctors">0</span></div>
    <div class="stat"><span class="stat-label">Patient Slots</span><span class="stat-value blue" id="stat-patients">0</span></div>
    <div class="stat"><span class="stat-label">Total Runs</span><span class="stat-value blue" id="stat-total">0</span></div>
    <div class="stat" id="stat-pass-wrap" style="display:none;"><span class="stat-label">Passed</span><span class="stat-value green" id="stat-pass">0</span></div>
    <div class="stat" id="stat-fail-wrap" style="display:none;"><span class="stat-label">Failed</span><span class="stat-value red" id="stat-fail">0</span></div>
  </div>

  <!-- Table -->
  <div class="table-wrap">
    <table id="run-table">
      <thead>
        <tr>
          <th>#</th>
          <th>Doctor Phone</th>
          <th>Patient ID</th>
          <th>Audio File</th>
          <th>Status</th>
          <th>audio_id</th>
          <th>summary_id</th>
          <th>Details</th>
        </tr>
      </thead>
      <tbody id="run-tbody"></tbody>
    </table>
  </div>

</div>

<!-- Logs panel -->
<div id="logs-panel">
  <div class="logs-header" id="logs-header">
    <div class="logs-header-left">
      <div class="logs-dot" id="logs-dot"></div>
      <span>Logs</span>
      <span id="log-count" style="font-weight:400;">0 lines</span>
    </div>
    <div class="logs-actions">
      <button class="logs-btn" id="logs-clear" type="button">Clear</button>
      <button class="logs-btn" id="logs-toggle" type="button">Minimize</button>
    </div>
  </div>
  <div id="logs-body"></div>
</div>

<script>
// ── State ────────────────────────────────────────────────────────────────────
let doctorCount    = 0;
let cardIdCounter  = 0;  // stable per-card ID (never decrements on remove)
let matrix         = [];   // [{rowId, doctorIdx, phone, patientId, audioName, audioIdx}]
let runStats       = { total: 0, passed: 0, failed: 0 };
let testResults    = [];   // [{doctor_id, patient_id, audio_id, ...timings}]

// ── Doctor cards ─────────────────────────────────────────────────────────────
function addDoctor() {
  const idx    = doctorCount++;
  const cardId = cardIdCounter++;
  const card   = document.createElement('div');
  card.className = 'doctor-card';
  card.dataset.idx    = idx;
  card.dataset.cardId = cardId;
  card.innerHTML = `
    <div class="doctor-card-header">
      <span class="doctor-num">${idx + 1}</span>
      <span>Doctor ${idx + 1}</span>
      <button class="remove-btn" onclick="removeDoctor(this)" title="Remove">×</button>
    </div>
    <div class="doctor-fields">
      <div class="field">
        <label>Phone Number</label>
        <input class="doc-phone" type="text" placeholder="9876543210" required>
      </div>
      <div class="field">
        <label>Password</label>
        <input class="doc-password" type="password" placeholder="••••••••" required>
      </div>
    </div>
    <div class="field">
      <label>Patient IDs</label>
      <div class="tag-box" onclick="this.querySelector('.tag-input').focus()">
        <input class="tag-input" type="text" placeholder="Type ID → Enter">
      </div>
      <div class="tag-hint">Press Enter or comma after each patient ID</div>
    </div>
    <div class="patient-audio-list"></div>`;

  const tagInput = card.querySelector('.tag-input');
  tagInput.addEventListener('keydown', e => {
    if (e.key === 'Enter' || e.key === ',') {
      e.preventDefault();
      addTag(tagInput, card);
    } else if (e.key === 'Backspace' && !tagInput.value) {
      const box  = tagInput.closest('.tag-box');
      const tags = box.querySelectorAll('.tag');
      if (tags.length) {
        const lastTag = tags[tags.length - 1];
        removePatientAudioRow(card, lastTag.dataset.patientId);
        lastTag.remove();
      }
    }
  });
  tagInput.addEventListener('blur', () => { if (tagInput.value.trim()) addTag(tagInput, card); });

  document.getElementById('doctors-list').appendChild(card);
  renumberDoctors();
}

function removeDoctor(btn) {
  btn.closest('.doctor-card').remove();
  renumberDoctors();
}

function renumberDoctors() {
  document.querySelectorAll('.doctor-card').forEach((c, i) => {
    c.dataset.idx = i;
    c.querySelector('.doctor-num').textContent = i + 1;
    c.querySelector('.doctor-card-header span:nth-child(2)').textContent = `Doctor ${i + 1}`;
  });
}

function addTag(input, card) {
  const val = input.value.replace(/,/g, '').trim();
  if (!val || isNaN(Number(val))) { input.value = ''; return; }
  const box = input.closest('.tag-box');
  if ([...box.querySelectorAll('.tag span')].some(s => s.textContent === val)) { input.value = ''; return; }
  const tag = document.createElement('span');
  tag.className = 'tag';
  tag.dataset.patientId = val;
  tag.innerHTML = `<span>${val}</span><button type="button" onclick="removeTag(this)">×</button>`;
  box.insertBefore(tag, input);
  input.value = '';
  addPatientAudioRow(card, val);
}

function removeTag(btn) {
  const tag  = btn.parentElement;
  const pid  = tag.dataset.patientId;
  const card = tag.closest('.doctor-card');
  removePatientAudioRow(card, pid);
  tag.remove();
}

function removePatientAudioRow(card, pid) {
  const row = card.querySelector(`.patient-audio-row[data-patient-id="${pid}"]`);
  if (row) row.remove();
}

function addPatientAudioRow(card, patientId) {
  const list = card.querySelector('.patient-audio-list');
  const row  = document.createElement('div');
  row.className = 'patient-audio-row';
  row.dataset.patientId = patientId;
  row._files = [];
  row.innerHTML = `
    <div class="pat-audio-header">
      <span class="pat-pid-badge">Patient ${patientId}</span>
      <label class="pat-upload-label">
        + Upload Audio
        <input type="file" class="pat-audio-input" accept="audio/*,.wav,.mp3,.m4a" multiple style="display:none">
      </label>
    </div>
    <div class="pat-lang-row">
      <span class="pat-lang-label">Language:</span>
      <select class="pat-lang-select">
        <option value="en">English (en)</option>
        <option value="hi">Hindi (hi)</option>
        <option value="ta">Tamil (ta)</option>
        <option value="te">Telugu (te)</option>
        <option value="kn">Kannada (kn)</option>
        <option value="ml">Malayalam (ml)</option>
        <option value="bn">Bengali (bn)</option>
        <option value="mr">Marathi (mr)</option>
      </select>
    </div>
    <div class="pat-file-list"><span class="pat-no-audio">No audio — silent WAV will be used</span></div>`;

  row.querySelector('.pat-audio-input').addEventListener('change', function() {
    for (const f of this.files) {
      if (!row._files.some(x => x.name === f.name)) row._files.push(f);
    }
    this.value = '';
    renderPatientFiles(row);
  });

  list.appendChild(row);
}

function renderPatientFiles(row) {
  const list = row.querySelector('.pat-file-list');
  if (!row._files.length) {
    list.innerHTML = '<span class="pat-no-audio">No audio — silent WAV will be used</span>';
    return;
  }
  list.innerHTML = row._files.map((f, i) => `
    <div class="pat-file-chip">
      <span title="${f.name}">🎵 ${f.name}</span>
      <button type="button" onclick="removePatientFile(this, ${i})">×</button>
    </div>`).join('');
}

function removePatientFile(btn, fileIdx) {
  const row = btn.closest('.patient-audio-row');
  row._files.splice(fileIdx, 1);
  renderPatientFiles(row);
}

function getDoctors() {
  return [...document.querySelectorAll('.doctor-card')].map(card => ({
    phone:      card.querySelector('.doc-phone').value.trim(),
    password:   card.querySelector('.doc-password').value.trim(),
    patientIds: [...card.querySelectorAll('.tag span')].map(s => s.textContent),
  }));
}

document.getElementById('add-doctor-btn').addEventListener('click', addDoctor);

// ── Advanced toggle ──────────────────────────────────────────────────────────
document.getElementById('adv-toggle').addEventListener('click', function() {
  this.classList.toggle('open');
  document.getElementById('adv-fields').classList.toggle('open');
});

// ── Logs ─────────────────────────────────────────────────────────────────────
const logsBody   = document.getElementById('logs-body');
const logsToggle = document.getElementById('logs-toggle');
const logsDot    = document.getElementById('logs-dot');
const logCountEl = document.getElementById('log-count');
let logsMin = false, logCount = 0;

function toggleLogs() {
  logsMin = !logsMin;
  logsBody.classList.toggle('minimized', logsMin);
  logsToggle.textContent = logsMin ? 'Expand' : 'Minimize';
}
document.getElementById('logs-header').addEventListener('click', toggleLogs);
logsToggle.addEventListener('click', e => { e.stopPropagation(); toggleLogs(); });
document.getElementById('logs-clear').addEventListener('click', e => {
  e.stopPropagation(); logsBody.innerHTML = '';
  logCount = 0; logCountEl.textContent = '0 lines';
});

function appendLog(msg) {
  logCount++;
  logCountEl.textContent = logCount + ' lines';
  const d = document.createElement('div');
  d.className = 'log-line';
  const m = msg.trim();
  if      (m.includes('OK]'))   d.classList.add('ok');
  else if (m.includes('WARN'))  d.classList.add('warn');
  else if (m.includes('FAIL') || m.includes('ERROR')) d.classList.add('fail');
  else if (m.includes('[Step')) d.classList.add('step');
  d.textContent = msg;
  logsBody.appendChild(d);
  if (!logsMin) logsBody.scrollTop = logsBody.scrollHeight;
}

// ── Status bar (screen 1) ────────────────────────────────────────────────────
const statusBar = document.getElementById('status-bar');
function setStatus(msg, type) { statusBar.textContent = msg; statusBar.className = type; statusBar.style.display = 'flex'; }
function clearStatus() { statusBar.style.display = 'none'; }

// ── Validation & matrix build ────────────────────────────────────────────────
function buildMatrix() {
  matrix = [];
  let rowNum = 1;
  document.querySelectorAll('.doctor-card').forEach((card, di) => {
    const phone = card.querySelector('.doc-phone').value.trim();
    card.querySelectorAll('.patient-audio-row').forEach(patRow => {
      const pid   = patRow.dataset.patientId;
      const files = patRow._files || [];
      const names = files.length ? files.map(f => f.name) : ['(silent WAV)'];
      names.forEach((aName, ai) => {
        matrix.push({ rowId: `d${di}_p${pid}_a${ai}`, rowNum: rowNum++, doctorIdx: di, phone, patientId: pid, audioName: aName, audioIdx: ai });
      });
    });
  });
  return matrix;
}

// ── Preview button ────────────────────────────────────────────────────────────
document.getElementById('preview-btn').addEventListener('click', () => {
  clearStatus();
  const doctors = getDoctors();

  // Validate
  if (!doctors.length) return setStatus('Add at least one doctor.', 'error');
  for (const [i, d] of doctors.entries()) {
    if (!d.phone)  return setStatus(`Doctor ${i+1}: phone number is required.`, 'error');
    if (!d.password) return setStatus(`Doctor ${i+1}: password is required.`, 'error');
    if (!d.patientIds.length) return setStatus(`Doctor ${i+1}: add at least one patient ID.`, 'error');
  }

  const mx = buildMatrix();
  const totalPatients = doctors.reduce((s, d) => s + d.patientIds.length, 0);

  // Update stats
  document.getElementById('stat-doctors').textContent  = doctors.length;
  document.getElementById('stat-patients').textContent = totalPatients;
  document.getElementById('stat-total').textContent    = mx.length;

  // Build table
  const tbody = document.getElementById('run-tbody');
  tbody.innerHTML = mx.map(r => `
    <tr id="row-${r.rowId}">
      <td class="mono" style="color:var(--muted)">${r.rowNum}</td>
      <td class="mono">${r.phone}</td>
      <td class="mono">${r.patientId}</td>
      <td style="max-width:200px;overflow:hidden;text-overflow:ellipsis;white-space:nowrap;" title="${r.audioName}">${r.audioName}</td>
      <td><span class="badge-status badge-pending">Pending</span></td>
      <td class="mono" id="aid-${r.rowId}">—</td>
      <td class="mono" id="sid-${r.rowId}">—</td>
      <td id="det-${r.rowId}" style="font-size:12px;color:var(--muted);">—</td>
    </tr>`).join('');

  // Reset run UI
  runStats = { total: mx.length, passed: 0, failed: 0 };
  document.getElementById('stat-pass-wrap').style.display = 'none';
  document.getElementById('stat-fail-wrap').style.display = 'none';
  document.getElementById('run-btn').style.display  = 'inline-flex';
  document.getElementById('run-btn-text').textContent = 'Confirm & Run';
  document.getElementById('run-btn').disabled = false;
  document.getElementById('back-btn').style.display = 'inline-flex';
  document.getElementById('new-test-btn').style.display = 'none';
  document.getElementById('export-btn').style.display = 'none';
  document.getElementById('run-progress-wrap').style.display = 'none';
  document.getElementById('run-counter').style.display = 'none';
  document.getElementById('header-progress').style.display = 'none';
  testResults = [];

  showScreen('s-run');
});

// ── Back button ───────────────────────────────────────────────────────────────
document.getElementById('back-btn').addEventListener('click', () => showScreen('s-input'));
document.getElementById('new-test-btn').addEventListener('click', () => {
  logsDot.classList.remove('active');
  showScreen('s-input');
});

// ── Export button ──────────────────────────────────────────────────────────────
document.getElementById('export-btn').addEventListener('click', async () => {
  if (testResults.length === 0) {
    alert('No test results to export.');
    return;
  }
  try {
    const response = await fetch('/export', {
      method: 'POST',
      headers: { 'Content-Type': 'application/json' },
      body: JSON.stringify({ results: testResults })
    });
    if (!response.ok) {
      throw new Error(`HTTP ${response.status}`);
    }
    const blob = await response.blob();
    const url = window.URL.createObjectURL(blob);
    const a = document.createElement('a');
    a.href = url;
    a.download = `medsum-results-${new Date().toISOString().split('T')[0]}.xlsx`;
    document.body.appendChild(a);
    a.click();
    window.URL.revokeObjectURL(url);
    document.body.removeChild(a);
    appendLog('✓ Excel file exported successfully');
  } catch (err) {
    appendLog(`✗ Export failed: ${err.message}`);
    alert(`Export failed: ${err.message}`);
  }
});

// ── Screen switch ─────────────────────────────────────────────────────────────
function showScreen(id) {
  document.querySelectorAll('.screen').forEach(s => s.classList.remove('active'));
  document.getElementById(id).classList.add('active');
  window.scrollTo(0, 0);
}

// ── Run ───────────────────────────────────────────────────────────────────────
document.getElementById('run-btn').addEventListener('click', async () => {
  const runBtn    = document.getElementById('run-btn');
  const runSpinner = document.getElementById('run-spinner');
  runBtn.disabled = true;
  document.getElementById('run-btn-text').textContent = 'Running…';
  runSpinner.style.display = 'inline-block';
  document.getElementById('back-btn').style.display = 'none';
  document.getElementById('run-progress-wrap').style.display = 'flex';
  document.getElementById('run-counter').style.display = 'inline';
  document.getElementById('header-progress').style.display = 'flex';
  document.getElementById('stat-pass-wrap').style.display = '';
  document.getElementById('stat-fail-wrap').style.display = '';

  logsDot.classList.add('active');
  if (logsMin) toggleLogs();

  runStats.passed = 0; runStats.failed = 0;
  updateProgress();

  // Build form data
  const fd = new FormData();
  fd.append('doctors_json', JSON.stringify(getDoctors()));
  // Per-patient audio files and language: paudio_{di}_{pid}_{ai}, plang_{di}_{pid}
  document.querySelectorAll('.doctor-card').forEach((card, di) => {
    card.querySelectorAll('.patient-audio-row').forEach(patRow => {
      const pid  = patRow.dataset.patientId;
      const lang = patRow.querySelector('.pat-lang-select').value;
      fd.append(`plang_${di}_${pid}`, lang);
      (patRow._files || []).forEach((f, ai) => fd.append(`paudio_${di}_${pid}_${ai}`, f, f.name));
    });
  });
  fd.append('llm',           document.getElementById('cfg-llm').value);
  fd.append('stt_model',     document.getElementById('cfg-stt-model').value);
  fd.append('translate_model', document.getElementById('cfg-translate-model').value);
  fd.append('template_type', document.getElementById('cfg-template-type').value);
  fd.append('template_id',   document.getElementById('cfg-template-id').value);
  fd.append('duration',      document.getElementById('cfg-duration').value);
  fd.append('django_url',    document.getElementById('cfg-django').value);
  fd.append('flask_url',     document.getElementById('cfg-flask').value);

  try {
    const resp = await fetch('/run', { method: 'POST', body: fd });
    if (!resp.ok) {
      const errText = await resp.text();
      appendLog(`[ERROR] Server returned ${resp.status}: ${errText}`);
      throw new Error(`HTTP ${resp.status}`);
    }
    const reader = resp.body.getReader();
    const decoder = new TextDecoder();
    let buf = '';

    while (true) {
      const { done, value } = await reader.read();
      if (done) break;
      buf += decoder.decode(value, { stream: true });
      const lines = buf.split('\n');
      buf = lines.pop();
      for (const line of lines) {
        if (!line.startsWith('data:')) continue;
        const raw = line.slice(5).trim();
        if (!raw) continue;
        let evt;
        try { evt = JSON.parse(raw); } catch { continue; }
        handleEvent(evt);
      }
    }
  } catch (err) {
    appendLog('[ERROR] ' + err.message);
  }

  runBtn.disabled = false;
  runSpinner.style.display = 'none';
  document.getElementById('run-btn-text').textContent = 'Run Again';
  document.getElementById('new-test-btn').style.display = 'inline-flex';
  document.getElementById('export-btn').style.display = 'inline-flex';
  logsDot.classList.remove('active');
});

function handleEvent(evt) {
  if (evt.type === 'log') {
    appendLog(evt.message);
  } else if (evt.type === 'row_update') {
    updateRow(evt);
  } else if (evt.type === 'done') {
    appendLog(`── All done: ${evt.passed}/${evt.total} passed ──`);
  }
}

function updateRow(evt) {
  const row = document.getElementById('row-' + evt.row_id);
  if (!row) return;
  const statusCell = row.querySelector('.badge-status');
  const aidCell    = document.getElementById('aid-' + evt.row_id);
  const sidCell    = document.getElementById('sid-' + evt.row_id);
  const detCell    = document.getElementById('det-' + evt.row_id);

  statusCell.className = 'badge-status badge-' + evt.status;
  if (evt.status === 'running') {
    statusCell.innerHTML = '<div class="dot-spin"></div> Running';
  } else if (evt.status === 'pass') {
    statusCell.textContent = '✓ Pass';
    runStats.passed++;
    if (aidCell) aidCell.textContent = evt.audio_id ?? '—';
    if (sidCell) sidCell.textContent = evt.summary_id ?? '—';
    if (detCell) {
      const timeStr = evt.total_time ? `${evt.total_time}s` : '';
      detCell.innerHTML = `${timeStr ? `<span style="color:var(--muted);font-size:11px;">${timeStr}</span> ` : ''}<button class="view-btn" onclick="toggleExpand('${evt.row_id}')">View ▾</button>`;
    }

    // Store result for export
    if (evt.timing_data) {
      testResults.push(evt.timing_data);
    }

    // Insert expandable sub-row after this row
    const expandRow = document.createElement('tr');
    expandRow.id = 'expand-' + evt.row_id;
    expandRow.className = 'expand-row';
    expandRow.style.display = 'none';
    const transcription = evt.transcription || '(empty)';
    const summary = typeof evt.summary === 'object'
      ? JSON.stringify(evt.summary, null, 2)
      : (evt.summary || '(empty)');
    expandRow.innerHTML = `<td colspan="8"><div class="expand-inner">
      <div class="expand-section">
        <div class="expand-label">🎙 Transcription</div>
        <div class="expand-text">${escHtml(transcription)}</div>
      </div>
      <div class="expand-section">
        <div class="expand-label">📋 Summary</div>
        <div class="expand-text">${escHtml(summary)}</div>
      </div>
    </div></td>`;
    row.insertAdjacentElement('afterend', expandRow);
    updateProgress();
  } else if (evt.status === 'fail') {
    statusCell.textContent = '✗ Fail';
    runStats.failed++;
    if (detCell) { detCell.innerHTML = `<span class="err-text">${evt.error || 'Unknown error'}</span>`; }
    updateProgress();
  }

  document.getElementById('stat-pass').textContent = runStats.passed;
  document.getElementById('stat-fail').textContent = runStats.failed;
}

function updateProgress() {
  const done  = runStats.passed + runStats.failed;
  const pct   = runStats.total ? Math.round(done / runStats.total * 100) : 0;
  document.getElementById('run-bar').style.width  = pct + '%';
  document.getElementById('hdr-bar').style.width  = pct + '%';
  document.getElementById('run-counter').textContent = `${done} / ${runStats.total}`;
  document.getElementById('hdr-count').textContent   = `${done} / ${runStats.total}`;
}

// ── Expand / collapse transcription+summary ───────────────────────────────────
function toggleExpand(rowId) {
  const expandRow = document.getElementById('expand-' + rowId);
  if (!expandRow) return;
  const visible = expandRow.style.display !== 'none';
  expandRow.style.display = visible ? 'none' : 'table-row';
  // Update button label
  const btn = document.querySelector(`#row-${rowId} .view-btn`);
  if (btn) btn.textContent = visible ? 'View ▾' : 'Hide ▴';
}

function escHtml(s) {
  return s.replace(/&/g,'&amp;').replace(/</g,'&lt;').replace(/>/g,'&gt;');
}

// ── Init ──────────────────────────────────────────────────────────────────────
addDoctor();  // start with one doctor row
</script>
</body>
</html>"""


# ─────────────────────────────────────────────────────────────────────────────
# Helpers
# ─────────────────────────────────────────────────────────────────────────────

def _make_wav(duration: float = 2.0, rate: int = 16000) -> bytes:
    buf = io.BytesIO()
    with wave.open(buf, "wb") as wf:
        wf.setnchannels(1); wf.setsampwidth(2); wf.setframerate(rate)
        wf.writeframes(b"\x00\x00" * int(rate * duration))
    return buf.getvalue()


def _extract_summary(body: dict) -> str:
    text = body.get("summary", "")
    if not text:
        for k in ("medical_summary", "subjective", "Diagnosis",
                  "Course_in_the_Hospital_and_Discussion", "Medicines"):
            v = body.get(k)
            if v:
                text = v if isinstance(v, str) else json.dumps(v, ensure_ascii=False)
                break
    if not text:
        content = {k: v for k, v in body.items() if k not in KNOWN_META_KEYS}
        text = json.dumps(content, ensure_ascii=False)
    return text


# ─────────────────────────────────────────────────────────────────────────────
# Per-doctor worker
# ─────────────────────────────────────────────────────────────────────────────

def _run_patient(doctor_idx, patient_id, doctor_id, auth_token,
                 doctor_name, department, hospital,
                 audios, cfg, q, doctor_times=None):
    """
    Runs all audios for one patient sequentially (step 4→5→6 must be in order).
    Each call gets its own requests.Session — safe to run in parallel with other patients.
    """
    def log(msg):          q.put({"type": "log",        "message": msg})
    def row(row_id, **kw): q.put({"type": "row_update", "row_id": row_id, **kw})

    django    = cfg["django_url"].rstrip("/")
    flask_url = cfg["flask_url"].rstrip("/")
    if not flask_url.endswith("/transcribe"):
        flask_url += "/transcribe"

    language        = cfg.get("patient_languages", {}).get((doctor_idx, str(patient_id)), cfg.get("language", "en"))
    llm             = cfg["llm"]
    stt_model       = cfg["stt_model"]
    translate_model = cfg["translate_model"]
    template_id     = cfg["template_id"]
    template        = SOAP_TEMPLATE if cfg["template_type"] == "soap" else DISCHARGE_TEMPLATE
    audio_duration  = float(cfg["duration"])

    # Own session per patient thread — not shared with other threads
    sess = requests.Session()
    sess.headers["Authorization"] = f"Bearer {auth_token}"

    # Fetch patient details
    patient_name = age = gender = ""
    start_time = time.time()
    patient_data_time = 0
    try:
        rpt = sess.get(f"{django}/api/patient-data/{patient_id}/", timeout=15)
        if rpt.status_code == 200:
            pd = rpt.json()
            patient_name = pd.get("patient_name", "")
            age          = str(pd.get("age", ""))
            gender       = pd.get("gender", "")
            patient_data_time = time.time() - start_time
            log(f"  [Dr {doctor_idx+1}|Pt {patient_id}] Patient data OK: {patient_name} | age={age} | gender={gender}  time={patient_data_time:.2f}s")
        else:
            patient_data_time = time.time() - start_time
            log(f"  [Dr {doctor_idx+1}|Pt {patient_id}] WARN: patient-data returned {rpt.status_code}  time={patient_data_time:.2f}s")
    except Exception as e:
        patient_data_time = time.time() - start_time
        log(f"  [Dr {doctor_idx+1}|Pt {patient_id}] WARN: {e}  time={patient_data_time:.2f}s")

    # Audios are sequential per patient (upload → transcribe → store must be in order)
    for audio_idx, (audio_name, audio_bytes) in enumerate(audios):
        row_id = f"d{doctor_idx}_p{patient_id}_a{audio_idx}"
        row(row_id, status="running")
        log(f"  [Dr {doctor_idx+1}|Pt {patient_id}] Audio {audio_idx+1}/{len(audios)}: {audio_name}")

        try:
            # Step 4 — transcribe (audio as base64 → AI pipeline)
            log(f"    [Dr {doctor_idx+1}|Pt {patient_id}] [Step 4] Transcribing ({language}/{llm})...")
            start_time = time.time()
            r4 = requests.post(flask_url, json={
                "audio_base64":      base64.b64encode(audio_bytes).decode(),
                "doctor_name":       doctor_name,
                "doctor_department": department,
                "hospital_name":     hospital,
                "patient_id":        str(patient_id),
                "patient_name":      patient_name,
                "age":               str(age),
                "gender":            gender,
                "template":          json.dumps(template),
                "template_id":       template_id,
                "language":          language,
                "llm":               llm,
                "stt_model":         stt_model,
                "translate_model":   translate_model,
            }, timeout=180)
            if r4.status_code != 200:
                raise RuntimeError(f"transcribe {r4.status_code}: {r4.text[:200]}")
            b4 = r4.json()
            total_time    = b4.get("total-time")
            transcription = b4.get("transcription", "")
            summary_text  = _extract_summary(b4)
            step4_time    = time.time() - start_time
            # Use actual audio length returned by Flask; fall back to configured duration
            audio_length  = b4.get("audio_length") or audio_duration
            # Compute audio_processing_time = flask_total - (stt + translation + llm)
            _ft  = b4.get("total-time") or 0
            _stt = b4.get("transcription-time") or 0
            _tr  = b4.get("translation-time") or 0
            _llm = b4.get("llm-time") or 0
            audio_processing_time = round(_ft - (_stt + _tr + _llm), 5) if _ft else None
            log(f"    [Dr {doctor_idx+1}|Pt {patient_id}] [Step 4 OK] total-time={total_time}s  audio_length={audio_length}s  step-time={step4_time:.2f}s")

            # Step 5 — upload audio recording for archival (after transcription completes)
            client_sid = str(uuid.uuid4())
            log(f"    [Dr {doctor_idx+1}|Pt {patient_id}] [Step 5] Uploading audio for archival...")
            start_time = time.time()
            r5 = sess.post(f"{django}/api/audio-data/",
                data={"user_id": str(doctor_id), "patient_id": str(patient_id),
                      "language": language, "file_duration": str(audio_duration), "session_id": client_sid},
                files={"audio": (audio_name, audio_bytes, "audio/wav")}, timeout=30)
            if r5.status_code != 201:
                raise RuntimeError(f"audio-data {r5.status_code}: {r5.text[:200]}")
            b5 = r5.json()
            if "audio_id" not in b5:
                raise RuntimeError(f"audio_id missing: {b5}")
            audio_id   = b5["audio_id"]
            session_id = b5.get("session_id") or client_sid
            step5_time = time.time() - start_time
            log(f"    [Dr {doctor_idx+1}|Pt {patient_id}] [Step 5 OK] audio_id={audio_id}  time={step5_time:.2f}s")

            # Step 6 — store summary
            log(f"    [Dr {doctor_idx+1}|Pt {patient_id}] [Step 6] Storing summary...")
            start_time = time.time()
            r6 = sess.post(f"{django}/api/summary-data/", json={
                "user_id": doctor_id, "patient_id": int(patient_id),
                "audio_id": audio_id, "session_id": session_id,
                "summary": summary_text, "summary_length": len(summary_text),
                "transcription": transcription, "template_id": template_id,
                "original_summary": summary_text, "is_approved": "No",
                "is_update": "False", "google_doc_link": None,
            }, timeout=30)
            if r6.status_code != 201:
                raise RuntimeError(f"summary-data {r6.status_code}: {r6.text[:200]}")
            summary_id = r6.json().get("summary_id")
            step6_time = time.time() - start_time
            log(f"    [Dr {doctor_idx+1}|Pt {patient_id}] [Step 6 OK] summary_id={summary_id}  time={step6_time:.2f}s")

            # Collect timing data for export
            timing_data = {
                "doctor_id":      doctor_id,
                "patient_id":     patient_id,
                "audio_id":       audio_id,
                "summary_id":     summary_id,
                "audio_duration": audio_length,  # actual length from Flask, not config default
                "language":       language,
            }
            if doctor_times:
                timing_data.update(doctor_times)
            timing_data.update({
                "patient_metadata_time":  round(patient_data_time, 2),
                "transcribe_rtt":         round(step4_time, 2),
                "audio_upload_time":      round(step5_time, 2),
                "summary_store_time":     round(step6_time, 2),
                "transcription_time":     b4.get("transcription-time"),
                "translation_time":       b4.get("translation-time"),
                "llm_time":               b4.get("llm-time"),
                "flask_total_time":       b4.get("total-time"),
                "audio_processing_time":  audio_processing_time,
            })

            row(row_id, status="pass", audio_id=audio_id, summary_id=summary_id,
                total_time=round(total_time, 1) if total_time else None,
                transcription=transcription,
                summary=summary_text,
                timing_data=timing_data)

        except Exception as e:
            log(f"    [Dr {doctor_idx+1}|Pt {patient_id}] [FAIL] {e}")
            row(row_id, status="fail", error=str(e))


def _run_doctor(doctor_idx, doctor_cfg, patient_audios, cfg, q):
    """
    Parallelism model:
      • Login + profile fetch  → sequential  (must complete before anything else)
      • Patients               → parallel    (each gets its own thread + session)
      • Audios within patient  → sequential  (step 4→5→6 must be in order)
    """
    def log(msg):          q.put({"type": "log",        "message": msg})
    def row(row_id, **kw): q.put({"type": "row_update", "row_id": row_id, **kw})

    django      = cfg["django_url"].rstrip("/")
    phone       = doctor_cfg["phone"]
    password    = doctor_cfg["password"]
    patient_ids = doctor_cfg["patientIds"]
    duration    = float(cfg["duration"])

    def get_audios(pid):
        result = patient_audios.get((doctor_idx, str(pid)), [])
        return result if result else [("silent.wav", _make_wav(duration))]

    # ── Step 1: Login (sequential — need token before anything else) ──────────
    log(f"  [Dr {doctor_idx+1}] [Step 1] Login  phone={phone}")
    start_time = time.time()
    step1_time = 0
    try:
        r = requests.post(f"{django}/api/login/",
                          json={"phone_number": phone, "password": password}, timeout=30)
        if r.status_code != 200:
            msg = f"Login failed {r.status_code}: {r.text[:150]}"
            log(f"  [Dr {doctor_idx+1}] [Step 1 FAIL] {msg}")
            for pid in patient_ids:
                for ai in range(len(get_audios(pid))):
                    row(f"d{doctor_idx}_p{pid}_a{ai}", status="fail", error=msg)
            return
        body      = r.json()
        doctor_id = body["user"]["id"]
        token     = body["access"]
        step1_time = time.time() - start_time
        log(f"  [Dr {doctor_idx+1}] [Step 1 OK] doctor_id={doctor_id}  free_minutes={body['user'].get('free_minutes')}  time={step1_time:.2f}s")
    except Exception as e:
        for pid in patient_ids:
            for ai in range(len(get_audios(pid))):
                row(f"d{doctor_idx}_p{pid}_a{ai}", status="fail", error=str(e))
        return

    # ── Step 1b: Doctor profile (sequential — shared across all patients) ─────
    doctor_name = department = hospital = ""
    start_time = time.time()
    step1b_time = 0
    try:
        rp = requests.get(f"{django}/api/user/update/{doctor_id}/",
                          headers={"Authorization": f"Bearer {token}"}, timeout=15)
        if rp.status_code == 200:
            prof        = rp.json()
            doctor_name = f"Dr {prof.get('firstname','')} {prof.get('lastname','')}".strip()
            department  = prof.get("department", "")
            hospital    = prof.get("hospital_name", "")
            step1b_time = time.time() - start_time
            log(f"  [Dr {doctor_idx+1}] [Step 1b OK] {doctor_name} | {department} | {hospital}  time={step1b_time:.2f}s")
        else:
            step1b_time = time.time() - start_time
            log(f"  [Dr {doctor_idx+1}] [Step 1b WARN] profile {rp.status_code} — using empty strings  time={step1b_time:.2f}s")
    except Exception as e:
        step1b_time = time.time() - start_time
        log(f"  [Dr {doctor_idx+1}] [Step 1b WARN] {e}  time={step1b_time:.2f}s")

    # Prepare doctor times for passing to patient threads
    doctor_times = {
        "step1_time": round(step1_time, 2),
        "step1b_time": round(step1b_time, 2),
    }

    # ── Patients: all run in parallel, each with its own session ──────────────
    log(f"  [Dr {doctor_idx+1}] Launching {len(patient_ids)} patient thread(s) in parallel...")
    with ThreadPoolExecutor(max_workers=len(patient_ids)) as pool:
        futures = {
            pool.submit(
                _run_patient,
                doctor_idx, pid, doctor_id, token,
                doctor_name, department, hospital,
                get_audios(pid), cfg, q, doctor_times
            ): pid
            for pid in patient_ids
        }
        for future in as_completed(futures):
            try:
                future.result()
            except Exception as e:
                pid = futures[future]
                log(f"  [Dr {doctor_idx+1}|Pt {pid}] Unhandled exception: {e}")


# ─────────────────────────────────────────────────────────────────────────────
# Routes
# ─────────────────────────────────────────────────────────────────────────────

@app.route("/")
def index():
    return (HTML
            .replace("__DJANGO_BASE_URL__", DJANGO_BASE_URL)
            .replace("__FLASK_BASE_URL__",  FLASK_BASE_URL))


@app.route("/run", methods=["POST"])
def run():
    try:
        doctors_raw = request.form.get("doctors_json", "[]")
        doctors     = json.loads(doctors_raw)
    except Exception as e:
        return {"error": f"Bad doctors_json: {e}"}, 400

    try:
        dur_raw  = request.form.get("duration", "")
        duration = float(dur_raw) if dur_raw.strip() else 2.0
    except Exception:
        duration = 2.0

    cfg = {
        "django_url":    request.form.get("django_url",    DJANGO_BASE_URL),
        "flask_url":     request.form.get("flask_url",     FLASK_BASE_URL),
        "language":      "en",  # fallback only — per-patient language overrides this
        "llm":           request.form.get("llm",           "OpenAI"),
        "stt_model":     request.form.get("stt_model",     "Bhasini"),
        "translate_model": request.form.get("translate_model", "Bhasini"),
        "template_type": request.form.get("template_type", "soap"),
        "template_id":   request.form.get("template_id",   "1"),
        "duration":      duration,
    }

    # Collect per-patient audio files: paudio_{di}_{pid}_{ai}
    try:
        patient_audios = {}
        for key in request.files:
            m = re.match(r'^paudio_(\d+)_(\w+)_(\d+)$', key)
            if m:
                di, pid, ai = int(m.group(1)), m.group(2), int(m.group(3))
                f = request.files[key]
                patient_audios.setdefault((di, pid), []).append((ai, f.filename, f.read()))
        for k in patient_audios:
            patient_audios[k].sort(key=lambda x: x[0])
            patient_audios[k] = [(name, data) for _, name, data in patient_audios[k]]
    except Exception as e:
        return {"error": f"Audio file read error: {e}"}, 500

    # Collect per-patient languages: plang_{di}_{pid}
    patient_languages = {}
    for key in request.form:
        m = re.match(r'^plang_(\d+)_(\w+)$', key)
        if m:
            di, pid = int(m.group(1)), m.group(2)
            patient_languages[(di, pid)] = request.form[key]
    cfg["patient_languages"] = patient_languages

    if not doctors:
        return {"error": "No doctors provided"}, 400

    def _get_audios(di, pid):
        result = patient_audios.get((di, str(pid)), [])
        return result if result else [("silent.wav", _make_wav(duration))]

    total = sum(
        len(_get_audios(di, pid))
        for di, doc in enumerate(doctors)
        for pid in doc.get("patientIds", [])
    )
    q          = queue.Queue()
    done_count = [0]
    passed     = [0]

    def run_all():
        total_patients = sum(len(d.get("patientIds", [])) for d in doctors)
        q.put({"type": "log", "message":
               f"  Starting load test: {len(doctors)} doctor(s) × {total_patients} patient slot(s) = {total} runs"})
        q.put({"type": "log", "message":
               f"  Concurrency model: doctors=parallel  patients=parallel  audios=sequential per patient"})

        with ThreadPoolExecutor(max_workers=len(doctors)) as pool:
            futures = [pool.submit(_run_doctor, di, doc, patient_audios, cfg, q)
                       for di, doc in enumerate(doctors)]
            for f in as_completed(futures):
                try:
                    f.result()
                except Exception as e:
                    q.put({"type": "log", "message": f"  [ERROR] Doctor thread: {e}"})

        q.put({"type": "done", "passed": passed[0], "total": total})
        q.put(None)

    threading.Thread(target=run_all, daemon=True).start()

    def generate():
        # Flush ngrok / reverse-proxy buffers (2 KB padding comment)
        yield ": " + " " * 2048 + "\n\n"
        while True:
            item = q.get()
            if item is None:
                break
            # Track pass count for done event
            if item.get("type") == "row_update" and item.get("status") == "pass":
                passed[0] += 1
            yield f"data: {json.dumps(item, ensure_ascii=False)}\n\n"

    return Response(
        stream_with_context(generate()),
        mimetype="text/event-stream",
        headers={"Cache-Control": "no-cache", "X-Accel-Buffering": "no"},
    )


@app.route("/export", methods=["POST"])
def export():
    try:
        data = request.get_json()
        results = data.get("results", [])
        
        if not results:
            return {"error": "No results to export"}, 400
        
        # Create a new workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Test Results"
        
        # Define headers — col order must match data rows below
        headers = [
            "Doctor ID",                          # 1
            "Patient ID",                         # 2
            "Audio ID",                           # 3
            "Summary ID",                         # 4
            "Audio Duration (s)",                 # 5
            "Language",                           # 6
            "Login Time (s)",                     # 7
            "Doctor Profile Time (s)",            # 8
            "Patient Metadata Time (s)",          # 9
            "Audio Processing Time (s)",          # 10
            "Audio Upload Time (s)",              # 11
            "Summary Store Time (s)",             # 12
            "STT Time (s)",                       # 13
            "Translation Time (s)",               # 14
            "LLM Time (s)",                       # 15
            "Flask Total Time (s)",               # 16
            "User Perceived Summary Latency (s)", # 17
            
        ]

        # Style the header row
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")

        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")

        def _t(result, key):
            v = result.get(key)
            return round(v, 5) if v is not None else None

        # Add data rows
        for row_idx, result in enumerate(results, 2):
            ws.cell(row=row_idx, column=1).value  = result.get("doctor_id")
            ws.cell(row=row_idx, column=2).value  = result.get("patient_id")
            ws.cell(row=row_idx, column=3).value  = result.get("audio_id")
            ws.cell(row=row_idx, column=4).value  = result.get("summary_id")
            ws.cell(row=row_idx, column=5).value  = result.get("audio_duration")
            ws.cell(row=row_idx, column=6).value  = result.get("language")
            ws.cell(row=row_idx, column=7).value  = _t(result, "step1_time")
            ws.cell(row=row_idx, column=8).value  = _t(result, "step1b_time")
            ws.cell(row=row_idx, column=9).value  = _t(result, "patient_metadata_time")
            ws.cell(row=row_idx, column=10).value = _t(result, "audio_processing_time")
            ws.cell(row=row_idx, column=11).value = _t(result, "audio_upload_time")
            ws.cell(row=row_idx, column=12).value = _t(result, "summary_store_time")
            ws.cell(row=row_idx, column=13).value = _t(result, "transcription_time")
            ws.cell(row=row_idx, column=14).value = _t(result, "translation_time")
            ws.cell(row=row_idx, column=15).value = _t(result, "llm_time")
            ws.cell(row=row_idx, column=16).value = _t(result, "flask_total_time")
            ws.cell(row=row_idx, column=17).value = _t(result, "transcribe_rtt")
            # ws.cell(row=row_idx, column=18).value = _t(result, "user_percieved_summary_latency")

        # Adjust column widths
        ws.column_dimensions['A'].width = 12
        ws.column_dimensions['B'].width = 12
        ws.column_dimensions['C'].width = 12
        ws.column_dimensions['D'].width = 12
        ws.column_dimensions['E'].width = 18
        ws.column_dimensions['F'].width = 14
        ws.column_dimensions['G'].width = 16
        ws.column_dimensions['H'].width = 22
        ws.column_dimensions['I'].width = 24
        ws.column_dimensions['J'].width = 18
        ws.column_dimensions['K'].width = 20
        ws.column_dimensions['L'].width = 22
        ws.column_dimensions['M'].width = 14
        ws.column_dimensions['N'].width = 18
        ws.column_dimensions['O'].width = 14
        ws.column_dimensions['P'].width = 20
        ws.column_dimensions['Q'].width = 24
        # ws.column_dimensions['R'].width = 30
        
        # Save to bytes
        excel_buffer = io.BytesIO()
        wb.save(excel_buffer)
        excel_buffer.seek(0)
        
        return send_file(
            excel_buffer,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            as_attachment=True,
            download_name=f"medsum-results-{datetime.now().strftime('%Y-%m-%d_%H-%M-%S')}.xlsx"
        )
    
    except Exception as e:
        return {"error": str(e)}, 500


if __name__ == "__main__":
    import os
    port = int(os.environ.get("PORT", 5050))
    print(f"\n  Medisum Load Test Console")
    print(f"  Open: http://127.0.0.1:{port}\n")
    app.run(host="0.0.0.0", port=port, debug=False, threaded=True)
